import math

import openpyxl
from openpyxl.styles import Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.pagebreak import Break


class DeskLabelGenerator:
    """
    桌角纸批量生成器。
    支持数据导入、自动分页、考场超员截断等功能。
    """

    def __init__(self, config):
        self.config = config
        # ==========================================
        # 样式定义
        # ==========================================
        self.font_style = Font(name="宋体", size=11)
        self.alignment_style = Alignment(wrap_text=True, vertical="center", horizontal="left")  # 左对齐
        self.thin_border = Border(
            left=Side(style="thin"),
            right=Side(style="thin"),
            top=Side(style="thin"),
            bottom=Side(style="thin"),
        )

        # 根据布局设置尺寸
        # A4 纵向宽度容量约 100 字符单位
        self.COL_WIDTH = 100 / self.config.layout_cols

        # A4 纵向可用高度点数 (保留上下边距后)
        # 以前 6行用 90 (总 540)，这里沿用 540 总高度作为基准，或者稍微增加以填满页面
        # 540 / 7 ≈ 77
        self.ROW_HEIGHT = 560 / self.config.layout_rows

    def _get_seat_mapping(self, rows, cols, pattern, start_pos):
        """
        计算座位号到坐标的映射
        返回: dict {seat_index_0_based: (r, c)}
        """
        mapping = {}
        current_seat = 0  # 0-based index

        custom_counts = self.config.custom_col_counts

        # 辅助函数：判断该位置是否有效
        def is_valid_pos(r, c):
            if custom_counts:
                if c < len(custom_counts):
                    return r < custom_counts[c]
            return True

        # 无论 start_pos 是左手位还是右手位，生成文件时都强制以“右手位”（左上角）作为起始位
        # 这样打印出来的纸张，座位1总是在左上角，符合剪裁习惯
        def get_actual_col(logic_col):
            # if start_pos == "left":
            #     return cols - 1 - logic_col
            return logic_col

        if pattern == "Z型横排":
            for r in range(rows):
                for c in range(cols):
                    actual_c = get_actual_col(c)
                    if is_valid_pos(r, actual_c):
                        mapping[current_seat] = (r, actual_c)
                        current_seat += 1

        elif pattern == "S型横排":
            for r in range(rows):
                is_even_row = r % 2 == 0

                if is_even_row:
                    # L -> R
                    for c in range(cols):
                        actual_c = get_actual_col(c)
                        if is_valid_pos(r, actual_c):
                            mapping[current_seat] = (r, actual_c)
                            current_seat += 1
                else:
                    # R -> L
                    for c in range(cols - 1, -1, -1):
                        actual_c = get_actual_col(c)
                        if is_valid_pos(r, actual_c):
                            mapping[current_seat] = (r, actual_c)
                            current_seat += 1

        elif pattern == "Z型竖排":
            for c in range(cols):
                for r in range(rows):
                    actual_c = get_actual_col(c)
                    if is_valid_pos(r, actual_c):
                        mapping[current_seat] = (r, actual_c)
                        current_seat += 1

        elif pattern == "S型竖排":
            for c in range(cols):
                is_even_col = c % 2 == 0
                if is_even_col:  # Down
                    actual_c = get_actual_col(c)
                    for r in range(rows):
                        if is_valid_pos(r, actual_c):
                            mapping[current_seat] = (r, actual_c)
                            current_seat += 1
                else:  # Up
                    actual_c = get_actual_col(c)
                    for r in range(rows - 1, -1, -1):
                        if is_valid_pos(r, actual_c):
                            mapping[current_seat] = (r, actual_c)
                            current_seat += 1

        return mapping

    def generate(self, progress_callback=None):
        """
        批量生成桌角纸。

        Args:
            progress_callback (func): 进度回调函数 callback(current_step, total_steps)
        """
        output_path = self.config.output_path
        student_data = self.config.student_data_list

        wb = openpyxl.Workbook()

        # 模式判断：是有数据生成还是空白生成
        if student_data:
            # 1. 生成总表
            ws_total = wb.active
            ws_total.title = "总表"
            self._generate_sheet_content(ws_total, student_data, progress_callback)
            self._setup_page_settings(ws_total)

            # 2. 按考场生成分表
            # 先按考场分组
            rooms_map = {}
            # 保持原始顺序
            room_order = []
            seen_rooms = set()

            for item in student_data:
                room = item.get("考场", "未命名考场")
                if room not in rooms_map:
                    rooms_map[room] = []

                rooms_map[room].append(item)

                if room not in seen_rooms:
                    room_order.append(room)
                    seen_rooms.add(room)

            for room_name in room_order:
                room_data = rooms_map[room_name]

                # 创建 Sheet
                # 名称过滤，防止非法字符
                safe_name = (
                    str(room_name)[:30]
                    .replace(":", "")
                    .replace("\\", "")
                    .replace("/", "")
                    .replace("?", "")
                    .replace("*", "")
                    .replace("[", "")
                    .replace("]", "")
                )
                ws_room = wb.create_sheet(title=safe_name)

                # 分表不传 progress_callback，以免进度条乱跳，只在生成总表时更新进度
                self._generate_sheet_content(ws_room, room_data, None)
                self._setup_page_settings(ws_room)

        else:
            ws = wb.active
            ws.title = "桌角纸（批量打印）"
            self._generate_empty(ws, progress_callback)
            self._setup_page_settings(ws)

        # 保存
        wb.save(output_path)
        return output_path

    def _setup_page_settings(self, ws):
        """统一页面设置"""
        ws.page_setup.orientation = "portrait"  # 纵向
        ws.page_setup.paperSize = 9  # A4

        # 页边距
        ws.page_margins.left = 0.2
        ws.page_margins.right = 0.2
        ws.page_margins.top = 0.2
        ws.page_margins.bottom = 0.2
        ws.page_margins.header = 0.0
        ws.page_margins.footer = 0.0

        # 打印居中
        ws.print_options.horizontalCentered = True
        ws.print_options.verticalCentered = True

    def _generate_sheet_content(self, ws, data_list, progress_callback):
        """根据数据生成单个Sheet的内容"""
        # 1. 按考场分组
        # 即使 data_list 只有一个考场的数据，这里逻辑也适用

        rooms_map = {}  # room_name -> [students]
        for item in data_list:
            # 统一使用 '考场' 字段作为分组依据，与 generate 中的逻辑保持一致
            room = item.get("考场", "Unknown")
            if room not in rooms_map:
                rooms_map[room] = []
            rooms_map[room].append(item)

        total_rooms = len(rooms_map)
        current_row_base = 1

        for idx, (room, students) in enumerate(rooms_map.items()):
            if progress_callback:
                progress_callback(idx, total_rooms)

            # 计算该考场需要分几页
            capacity = self.config.layout_rows * self.config.layout_cols

            # 分块
            chunks = [students[i : i + capacity] for i in range(0, len(students), capacity)]

            for chunk_idx, chunk in enumerate(chunks):
                # 生成这一页
                self._fill_page_grid(ws, chunk, current_row_base)

                # 更新基准行
                current_row_base += self.config.layout_rows

                # 插入分页符
                # 逻辑：只要不是(最后一个房间 的 最后一页)，就加分页符
                is_last_room = idx == total_rooms - 1
                is_last_chunk = chunk_idx == len(chunks) - 1

                if not (is_last_room and is_last_chunk):
                    ws.row_breaks.append(Break(id=current_row_base - 1))

    def _generate_empty(self, ws, progress_callback):
        """生成空白模板"""
        total_count = self.config.total_count
        layout_rows = self.config.layout_rows
        layout_cols = self.config.layout_cols

        # 构造伪数据
        fake_data = []
        for i in range(total_count):
            fake_data.append({"考生姓名": "", "考生考号": "", "考场": "", "考场号": "", "座位号": ""})

        # 当作一个大考场处理
        capacity = layout_rows * layout_cols
        chunks = [fake_data[i : i + capacity] for i in range(0, len(fake_data), capacity)]

        total_chunks = len(chunks)
        current_row_base = 1

        for i, chunk in enumerate(chunks):
            if progress_callback:
                progress_callback(i, total_chunks)

            self._fill_page_grid(ws, chunk, current_row_base)
            current_row_base += layout_rows

            # 只有中间需要分页符
            if i < total_chunks - 1:
                ws.row_breaks.append(Break(id=current_row_base - 1))

    def _fill_page_grid(self, ws, students, start_row):
        """填充单页网格"""
        layout_rows = self.config.layout_rows
        layout_cols = self.config.layout_cols

        pattern = getattr(self.config, "layout_pattern", "S型横排")
        start_pos = getattr(self.config, "start_pos", "left")
        seat_mapping = self._get_seat_mapping(layout_rows, layout_cols, pattern, start_pos)

        # 设置行高 (整页)
        for r in range(layout_rows):
            ws.row_dimensions[start_row + r].height = self.ROW_HEIGHT

        # 设置列宽 (只需设置一次，但重复设置也没事)
        for c in range(layout_cols):
            col_letter = get_column_letter(c + 1)
            ws.column_dimensions[col_letter].width = self.COL_WIDTH

        pos_to_student = {}
        for idx, student in enumerate(students):
            pos = seat_mapping.get(idx)
            if not pos:
                continue
            pos_to_student[pos] = student

        for r in range(layout_rows):
            for c in range(layout_cols):
                cell_row = start_row + r
                cell_col = c + 1
                cell = ws.cell(row=cell_row, column=cell_col)

                student = pos_to_student.get((r, c))
                if student:
                    name = student.get("考生姓名") or student.get("姓名") or ""
                    num = student.get("考生考号") or student.get("考号") or ""
                    room_name = student.get("考场", "")
                    room_num = student.get("考场号", "")
                    seat = student.get("座位号", "")
                    content = f"姓名：{name}\n考号：{num}\n考场：{room_name}\n考场号：{room_num}\n座位号：{seat}"
                else:
                    content = "姓名：\n考号：\n考场：\n考场号：\n座位号："

                cell.value = content
                cell.font = self.font_style
                cell.alignment = self.alignment_style
                cell.border = self.thin_border
