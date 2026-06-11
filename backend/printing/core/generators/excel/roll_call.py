from __future__ import annotations

from collections import OrderedDict

import openpyxl
from openpyxl.styles import Alignment, Border, Font, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.pagebreak import Break

from backend.printing.core.generators.roll_call_text import format_class_name
from backend.printing.core.seat_layout import get_seat_mapping, mirror_layout_start_pos, normalize_layout


class RollCallGenerator:
    A4_WIDTH_PT = 595.28
    A4_HEIGHT_PT = 841.89
    PAGE_MARGIN_IN = 12 / 25.4
    HEADER_HEIGHT_PT = 77
    FULL_FOOTER_RATIO = 0.20
    FOOTER_GAP_PT = 28

    def __init__(self, config):
        self.config = config
        thin = Side(style="thin", color="666666")
        self.thin = thin
        self.border = Border(left=thin, right=thin, top=thin, bottom=thin)

    @staticmethod
    def _sheet_name(name: str, used: set[str]) -> str:
        base = str(name or "科目").translate(str.maketrans({c: "" for c in ':\\/?*[]'}))[:28] or "科目"
        candidate = base
        index = 2
        while candidate in used:
            candidate = f"{base[:25]}-{index}"
            index += 1
        used.add(candidate)
        return candidate

    def generate(self, progress_callback=None):
        groups = self.config.groups or []
        if not groups:
            raise ValueError("没有可生成的点名表数据")

        by_subject = OrderedDict()
        for group in groups:
            by_subject.setdefault(group.get("subject") or "科目", []).append(group)

        wb = openpyxl.Workbook()
        wb.remove(wb.active)
        used = set()
        total = len(groups)
        done = 0
        for subject, subject_groups in by_subject.items():
            ws = wb.create_sheet(self._sheet_name(subject, used))
            max_cols = max(normalize_layout(group.get("seatLayout"))["layoutCols"] for group in subject_groups)
            landscape = self._is_landscape(max_cols)
            max_end_row = 1
            current_row = 1
            for index, group in enumerate(subject_groups):
                end_row = self._draw_page(ws, current_row, group, max_cols, landscape)
                max_end_row = max(max_end_row, end_row)
                done += 1
                if progress_callback:
                    progress_callback(done, total)
                if index < len(subject_groups) - 1:
                    ws.row_breaks.append(Break(id=end_row))
                current_row = end_row + 1

            ws.page_setup.orientation = "landscape" if landscape else "portrait"
            ws.page_setup.paperSize = ws.PAPERSIZE_A4
            ws.page_setup.fitToWidth = 1
            ws.page_setup.fitToHeight = 0
            ws.page_setup.scale = None
            ws.sheet_properties.pageSetUpPr.fitToPage = True
            ws.page_margins.left = self.PAGE_MARGIN_IN
            ws.page_margins.right = self.PAGE_MARGIN_IN
            ws.page_margins.top = self.PAGE_MARGIN_IN
            ws.page_margins.bottom = self.PAGE_MARGIN_IN
            ws.page_margins.header = 0
            ws.page_margins.footer = 0
            ws.print_options.horizontalCentered = True
            ws.print_options.verticalCentered = False
            ws.sheet_view.showGridLines = False
            ws.print_area = f"A1:{get_column_letter(max_cols)}{max_end_row}"

        wb.save(self.config.output_path)
        return self.config.output_path

    def _is_landscape(self, max_cols: int) -> bool:
        mode = str(getattr(self.config, "orientation", "auto") or "auto")
        if mode == "landscape":
            return True
        if mode == "portrait":
            return False
        return max_cols >= 8

    def _page_content_size(self, landscape: bool) -> tuple[float, float]:
        page_width = self.A4_HEIGHT_PT if landscape else self.A4_WIDTH_PT
        page_height = self.A4_WIDTH_PT if landscape else self.A4_HEIGHT_PT
        margin_points = self.PAGE_MARGIN_IN * 72
        return page_width - margin_points * 2, page_height - margin_points * 2

    def _apply_box_border(self, ws, min_row: int, max_row: int, min_col: int, max_col: int) -> None:
        for row in range(min_row, max_row + 1):
            for col in range(min_col, max_col + 1):
                ws.cell(row, col).border = Border(
                    left=self.thin if col == min_col else None,
                    right=self.thin if col == max_col else None,
                    top=self.thin if row == min_row else None,
                    bottom=self.thin if row == max_row else None,
                )

    def _draw_page(self, ws, start_row: int, group: dict, sheet_cols: int, landscape: bool) -> int:
        layout = normalize_layout(mirror_layout_start_pos(group.get("seatLayout"), self.config.mirror_view))
        rows = layout["layoutRows"]
        cols = layout["layoutCols"]
        mapping = get_seat_mapping(layout)
        subject = str(group.get("subject") or "")
        room_name = str(group.get("roomName") or "")
        room_no = str(group.get("roomNo") or "")
        students = group.get("students") or []
        by_seat = {int(item["seatNo"]): item for item in students}
        font_size = max(7, min(11, 12 - max(0, cols - 6) - max(0, rows - 7) // 2))

        usable_width, usable_height = self._page_content_size(landscape)
        footer_region_height = usable_height * self.FULL_FOOTER_RATIO if self.config.template_mode == "full" else 12
        grid_height = max(rows * 38, usable_height - self.HEADER_HEIGHT_PT - footer_region_height)

        title_row = start_row
        ws.merge_cells(start_row=title_row, start_column=1, end_row=title_row, end_column=sheet_cols)
        ws.cell(title_row, 1, self.config.exam_name).font = Font(name="宋体", size=18, bold=True)
        ws.cell(title_row, 1).alignment = Alignment(horizontal="center", vertical="center")
        ws.row_dimensions[title_row].height = 30

        info_row = title_row + 1
        ws.merge_cells(start_row=info_row, start_column=1, end_row=info_row, end_column=sheet_cols)
        ws.cell(info_row, 1, f"学校：{self.config.school_name}    科目：{subject}    考场：{room_name}    考场号：{room_no}    人数：{len(students)}")
        ws.cell(info_row, 1).font = Font(name="宋体", size=10)
        ws.cell(info_row, 1).alignment = Alignment(horizontal="center", vertical="center")
        ws.row_dimensions[info_row].height = 22

        sign_row = info_row + 1
        ws.merge_cells(start_row=sign_row, start_column=1, end_row=sign_row, end_column=sheet_cols)
        sign_text = "主监考（签名）：________________    副监考（签名）：________________" if self.config.template_mode == "full" else ""
        ws.cell(sign_row, 1, sign_text)
        ws.cell(sign_row, 1).alignment = Alignment(horizontal="right", vertical="center")
        ws.cell(sign_row, 1).font = Font(name="宋体", size=10)
        ws.row_dimensions[sign_row].height = 25 if sign_text else 8

        grid_start = sign_row + 1
        row_height = grid_height / max(1, rows)
        col_width = max(8, min(24, usable_width / max(1, sheet_cols) / 5.25))
        valid_positions = set(mapping.values())
        pos_to_seat = {position: seat for seat, position in mapping.items()}
        for col in range(1, sheet_cols + 1):
            ws.column_dimensions[get_column_letter(col)].width = col_width

        for row in range(rows):
            ws.row_dimensions[grid_start + row].height = row_height
            for col in range(sheet_cols):
                cell = ws.cell(grid_start + row, col + 1)
                if col >= cols or (row, col) not in valid_positions:
                    continue
                seat = pos_to_seat[(row, col)]
                student = by_seat.get(seat)
                if student:
                    lines = [f"{seat}. {student.get('name', '')}"]
                    if self.config.show_exam_no:
                        lines.append(str(student.get("examNo") or ""))
                    if self.config.show_class and student.get("className"):
                        lines.append(format_class_name(student.get("className")))
                    if self.config.show_checkbox:
                        lines.append("□ 缺考")
                    cell.value = "\n".join(lines)
                else:
                    cell.value = f"{seat}.\n" + ("□ 缺考" if self.config.show_checkbox else "")
                cell.font = Font(name="宋体", size=font_size)
                cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
                cell.border = self.border

        footer_row = grid_start + rows
        if self.config.template_mode == "full":
            ws.row_dimensions[footer_row].height = self.FOOTER_GAP_PT
            box_start = footer_row + 1
            box_end = box_start + 2
            box_height = max(72, footer_region_height - self.FOOTER_GAP_PT)
            ws.merge_cells(start_row=box_start, start_column=1, end_row=box_end, end_column=max(1, sheet_cols * 2 // 3))
            note_end = max(1, sheet_cols * 2 // 3)
            self._apply_box_border(ws, box_start, box_end, 1, note_end)
            note = ws.cell(box_start, 1, self.config.notes_title)
            note.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
            note.font = Font(name="宋体", size=10)
            instruction_start = max(2, sheet_cols * 2 // 3 + 1)
            ws.merge_cells(start_row=box_start, start_column=instruction_start, end_row=box_end, end_column=sheet_cols)
            instruction = ws.cell(box_start, instruction_start, self.config.instructions)
            instruction.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
            instruction.font = Font(name="宋体", size=9)
            for row in range(box_start, box_end + 1):
                ws.row_dimensions[row].height = box_height / 3
            return box_end
        ws.row_dimensions[footer_row].height = footer_region_height
        return footer_row
