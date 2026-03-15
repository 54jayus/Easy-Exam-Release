import openpyxl
from openpyxl.styles import Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.pagebreak import Break


class CornerPaperGenerator:
    """
    负责生成台角纸 Excel 文件的核心逻辑类。
    """

    # 模板尺寸常量
    TEMPLATE_HEIGHT = 14  # 单个模板占用的行数
    TEMPLATE_WIDTH = 6  # 单个模板占用的列数 (A-F 共6列)
    TEMPLATES_PER_ROW = 3  # 每一行放置的模板数量

    # ==========================================
    # 样式定义
    # ==========================================
    # 1. 边框样式
    THIN_BORDER = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )

    DASH_DOT_SIDE = Side(style="dashDot")

    # 2. 对齐样式
    CENTER_ALIGN = Alignment(horizontal="center", vertical="center")

    # 3. 字体样式
    FONT_TITLE = Font(name="宋体", size=14, bold=True)  # 大标题
    FONT_HEADER = Font(name="宋体", size=10, bold=True)  # 表头标签
    FONT_NORMAL = Font(name="宋体", size=10, bold=False)  # 普通内容

    def __init__(self, config):
        self.config = config
        self.subjects = config.subjects

        # 动态计算模板高度
        # 4行表头 + 科目数量 + 2行间隔(包含裁剪线位置)
        self.TEMPLATE_HEIGHT = 4 + len(self.subjects) + 2

        # 动态计算每页纵向模板数量 (Templates Per Column per Page)
        subject_count = len(self.subjects)
        if subject_count <= 3:
            self.TEMPLATES_PER_COL_PAGE = 5
        elif 4 <= subject_count <= 5:
            self.TEMPLATES_PER_COL_PAGE = 4
        elif 6 <= subject_count <= 9:
            self.TEMPLATES_PER_COL_PAGE = 3
        else:  # >= 10
            self.TEMPLATES_PER_COL_PAGE = 2

    def draw_template(self, ws, start_row, start_col, data=None):
        """
        在指定位置绘制一个完整的台角纸模板。
        使用相对坐标逻辑。
        """

        # 检测是否为高考模式数据
        is_gaokao_mode = data and "科目数据" in data

        if is_gaokao_mode:
            # 高考模式：使用科目数据数组
            subject_data_list = data.get("科目数据", [])
            kaochang = data.get("考场", "")
            kaochang_no = data.get("考场号", "")
            seat_no = data.get("座位号", "")
            # 高考模式下，学生信息在科目数据中
            name = ""
            exam_no = ""
            class_student = ""
        else:
            # 普通模式：保持原有逻辑
            kaochang = data.get("考场", "") if data else ""
            kaochang_no = data.get("考场号", "") if data else ""
            seat_no = data.get("座位号", "") if data else ""
            name = (data.get("考生姓名") or data.get("姓名") or "") if data else ""
            exam_no = (data.get("考生考号") or data.get("考号") or "") if data else ""

            class_no = data.get("班级", "") if data else ""
            student_no = data.get("学号", "") if data else ""

            # 拼接班级学号 (x班x号 格式)
            class_student = ""
            # 优先使用已经处理好的 '考生班级学号' 字段
            if data and "考生班级学号" in data:
                class_student = data["考生班级学号"]
            else:
                # 回退逻辑
                if data:
                    try:
                        c_str = str(class_no)
                        s_str = str(student_no)
                        class_student = f"{c_str}班{s_str}号"
                    except:
                        class_student = f"{class_no}班{student_no}号"

        # --- 辅助函数：相对坐标获取单元格 ---
        def cell(r_offset, c_offset):
            return ws.cell(row=start_row + r_offset, column=start_col + c_offset)

        # 设置行高 (13.5磅)
        for r_offset in range(self.TEMPLATE_HEIGHT):
            ws.row_dimensions[start_row + r_offset].height = 13.5

        # --- 1. 绘制标题行 (第1行, Offset 0) ---
        start_cell_str = cell(0, 0).coordinate
        end_cell_str = cell(0, 3).coordinate
        ws.merge_cells(f"{start_cell_str}:{end_cell_str}")

        c = cell(0, 0)
        c.value = self.config.title
        c.font = self.FONT_TITLE
        c.alignment = self.CENTER_ALIGN

        for c_off in range(4):
            cell(0, c_off).border = self.THIN_BORDER

        # --- 2. 绘制考场信息行 (第2行, Offset 1) ---
        data_row_2 = [
            (0, "考场", self.FONT_HEADER),
            (1, kaochang, self.FONT_NORMAL),
            (2, "考场号", self.FONT_HEADER),
            (3, kaochang_no, self.FONT_NORMAL),
        ]
        for c_off, val, font in data_row_2:
            c = cell(1, c_off)
            c.value = val
            c.font = font
            c.alignment = self.CENTER_ALIGN
            c.border = self.THIN_BORDER

        # --- 3. 绘制座位号行 (第3行, Offset 2) ---
        for c_off in range(2):
            cell(2, c_off).border = self.THIN_BORDER

        data_row_3 = [
            (2, "座位号", self.FONT_HEADER),
            (3, seat_no, self.FONT_NORMAL),
        ]
        for c_off, val, font in data_row_3:
            c = cell(2, c_off)
            c.value = val
            c.font = font
            c.alignment = self.CENTER_ALIGN
            c.border = self.THIN_BORDER

        # --- 4. 绘制列表头 (第4行, Offset 3) ---
        data_row_4 = [
            (0, "科目", self.FONT_HEADER),
            (1, "考生姓名", self.FONT_HEADER),
            (2, "考生考号", self.FONT_HEADER),
            (3, "考生班级学号", self.FONT_HEADER),
        ]
        for c_off, val, font in data_row_4:
            c = cell(3, c_off)
            c.value = val
            c.font = font
            c.alignment = self.CENTER_ALIGN
            c.border = self.THIN_BORDER

        # --- 5. 绘制科目数据行 (第5-12行, Offset 4-11) ---
        # 使用配置的科目列表
        subjects = self.subjects

        for i, subject in enumerate(subjects):
            r_current = 4 + i

            if is_gaokao_mode:
                # 高考模式：每个科目显示不同的学生信息
                if i < len(subject_data_list):
                    subject_item = subject_data_list[i]
                    subject_name = subject_item.get("科目", "")
                    student_name = subject_item.get("考生姓名", "")
                    student_exam_no = subject_item.get("考生考号", "")
                    student_class_student = subject_item.get("考生班级学号", "")
                else:
                    subject_name = subject
                    student_name = ""
                    student_exam_no = ""
                    student_class_student = ""
            else:
                # 普通模式：所有科目显示相同的学生信息
                subject_name = subject
                student_name = name
                student_exam_no = exam_no
                student_class_student = class_student

            # 第1列: 科目名称
            c = cell(r_current, 0)
            c.value = subject_name
            c.font = self.FONT_NORMAL
            c.alignment = self.CENTER_ALIGN
            c.border = self.THIN_BORDER

            # 第2-4列: 学生信息
            placeholders = [student_name, student_exam_no, student_class_student]
            for p_idx, p_text in enumerate(placeholders):
                c = cell(r_current, p_idx + 1)
                c.value = p_text
                c.font = self.FONT_NORMAL
                c.alignment = self.CENTER_ALIGN
                c.border = self.THIN_BORDER

        # --- 6. 绘制裁剪线 ---
        # A. 横向裁剪线
        cut_line_row_offset = 4 + len(subjects)
        for c_off in range(6):
            c = cell(cut_line_row_offset, c_off)
            current_border = c.border
            new_border = Border(
                left=current_border.left,
                right=current_border.right,
                top=current_border.top,
                bottom=self.DASH_DOT_SIDE,
            )
            c.border = new_border

        # B. 纵向裁剪线
        cut_line_col_offset = 4
        for r_off in range(self.TEMPLATE_HEIGHT):
            c = cell(r_off, cut_line_col_offset)
            current_border = c.border
            new_border = Border(
                left=current_border.left,
                right=self.DASH_DOT_SIDE,
                top=current_border.top,
                bottom=current_border.bottom,
            )
            c.border = new_border

        # --- 7. 特殊行高设置 ---
        ws.row_dimensions[start_row].height = 18.75
        ws.row_dimensions[start_row + cut_line_row_offset].height = 5
        ws.row_dimensions[start_row + cut_line_row_offset + 1].height = 5

    def generate_sheet(
        self,
        ws,
        data_list,
        start_idx=0,
        progress_callback=None,
        total_steps=1,
        include_room_in_footer=True,
        force_break_by_room=False,
    ):
        """
        在给定的 Worksheet 上绘制一批数据。
        """
        num_templates = len(data_list)
        
        # 如果需要强制分页，我们不能简单地使用 total_grid_rows
        # 我们需要迭代数据并计算位置
        
        current_grid_row = 0
        current_col = 0
        last_room = None
        
        # 记录上一次分页的行号（绝对行号），用于计算相对偏移
        last_break_row_idx = 0
        
        # 用于进度条
        # 我们不能简单地使用 grid_row，因为现在是非线性的
        # 我们每处理 TEMPLATES_PER_ROW 个数据，或者换行时，更新一次进度?
        # 为了简单，我们每次绘制都检查一下进度
        
        for idx, data in enumerate(data_list):
            # 报告进度
            if idx % self.TEMPLATES_PER_ROW == 0 and progress_callback:
                 current_step = start_idx + (current_grid_row) # 粗略估计
                 if current_step < total_steps:
                     progress_callback(current_step, total_steps)

            # 检查考场变化 (强制分页)
            room = data.get("考场") if data else None
            
            # 如果不是第一个数据，且开启了强制分页，且考场发生了变化
            if force_break_by_room and idx > 0 and room != last_room:
                # 换行
                if current_col > 0:
                    current_grid_row += 1
                    current_col = 0
                
                # 插入分页符 (在当前行之前)
                # current_grid_row 现在指向新的一行
                # 计算该行的起始像素行号
                current_start_row = current_grid_row * self.TEMPLATE_HEIGHT + 1
                ws.row_breaks.append(Break(id=current_start_row - 1))
                
                # 重置分页计数器 (模拟)
                last_break_row_idx = current_grid_row
            
            # 更新 last_room
            if data:
                last_room = room

            # 插入自然分页符
            # 逻辑：当前行号 - 上次分页行号 >= 每页行数
            rows_since_break = current_grid_row - last_break_row_idx
            
            if rows_since_break > 0 and rows_since_break % self.TEMPLATES_PER_COL_PAGE == 0 and current_col == 0:
                 current_start_row = current_grid_row * self.TEMPLATE_HEIGHT + 1
                 ws.row_breaks.append(Break(id=current_start_row - 1))
                 last_break_row_idx = current_grid_row

            start_row = current_grid_row * self.TEMPLATE_HEIGHT + 1
            start_col = current_col * self.TEMPLATE_WIDTH + 1

            self.draw_template(ws, start_row, start_col, data)
            
            current_col += 1
            if current_col >= self.TEMPLATES_PER_ROW:
                current_col = 0
                current_grid_row += 1

        # 2. 列宽设置
        col_widths = [10.625, 8.625, 12.625, 11.625, 1.275, 1.275]
        # 计算用到的最大列数
        # 由于我们可能换行了，列数最大也就是 TEMPLATES_PER_ROW
        max_grid_cols = self.TEMPLATES_PER_ROW 
        
        # 优化：如果数据很少不足一行，不需要设置那么多列宽
        if num_templates < self.TEMPLATES_PER_ROW and current_grid_row == 0:
             max_grid_cols = num_templates

        for grid_c in range(max_grid_cols):
            base_col_idx = grid_c * self.TEMPLATE_WIDTH
            for local_c, width in enumerate(col_widths):
                col_letter = get_column_letter(base_col_idx + local_c + 1)
                ws.column_dimensions[col_letter].width = width

        # 3. 页面设置
        ws.page_setup.paperSize = 9  # A4
        ws.page_setup.orientation = "landscape"

        ws.page_margins.left = 0.2
        ws.page_margins.right = 0
        ws.page_margins.top = 0.1
        ws.page_margins.bottom = 0.06
        ws.page_margins.header = 0
        ws.page_margins.footer = 0

        # 4. 页脚设置
        if include_room_in_footer:
            first_kaochang = ""
            for d in data_list:
                if d and d.get("考场"):
                    first_kaochang = d.get("考场")
                    break
            footer_text = f"&\"宋体\"&8第 &P 页，共 &N 页，当前考场：{first_kaochang}"
        else:
            footer_text = f"&\"宋体\"&8第 &P 页，共 &N 页"

        ws.oddFooter.center.text = footer_text
        ws.evenFooter.center.text = footer_text

    def generate(self, progress_callback=None):
        """
        批量生成台角纸。
        如果提供了 student_data_list，将按 '考场' 分 Sheet。
        """
        num_templates = self.config.num_templates
        output_path = self.config.output_path
        student_data_list = self.config.student_data_list

        wb = openpyxl.Workbook()

        # 如果是空白模板生成模式
        if not student_data_list:
            ws = wb.active
            ws.title = "空白模板"
            # 构造虚拟空数据
            dummy_data = [None] * num_templates
            self.generate_sheet(
                ws,
                dummy_data,
                0,
                progress_callback,
                (num_templates + 2) // 3,
                include_room_in_footer=False,
            )

        else:
            # 按考场分组
            exam_rooms = {}
            # 保持原始顺序
            room_order = []

            for d in student_data_list:
                room = d.get("考场", "未命名考场")
                if room not in exam_rooms:
                    exam_rooms[room] = []
                    room_order.append(room)
                exam_rooms[room].append(d)

            # 计算总步数用于进度条
            # 总表步数 + 各分表步数
            total_templates = len(student_data_list)
            total_rows_summary = (total_templates + self.TEMPLATES_PER_ROW - 1) // self.TEMPLATES_PER_ROW

            # 简单估算：总表占一半工作量，分表占一半工作量。或者直接按总行数累加。
            # 为了简单，我们认为总表是一次完整遍历，分表也是一次完整遍历。
            # 所以总步数 = total_rows_summary * 2 (近似)
            total_steps = total_rows_summary * 2

            processed_steps = 0

            # 1. 生成总表
            ws_total = wb.active
            ws_total.title = "总表"
            
            # 排序：先按考场排序，以支持强制分页
            # 辅助函数：尝试将字符串转换为数字进行排序
            def safe_int_sort_key(val):
                if isinstance(val, int):
                    return val
                if isinstance(val, str) and val.isdigit():
                    return int(val)
                # 尝试从字符串中提取数字 (例如 "001" -> 1)
                import re
                match = re.search(r"(\d+)", str(val))
                if match:
                    return int(match.group(1))
                return 0

            # 按考场号排序，如果考场号相同则按考场名，再按座位号
            student_data_list_sorted = sorted(
                student_data_list, 
                key=lambda x: (
                    safe_int_sort_key(x.get("考场号", 0)), 
                    x.get("考场", ""),
                    safe_int_sort_key(x.get("座位号", 0))
                )
            )

            self.generate_sheet(
                ws_total,
                student_data_list_sorted,
                processed_steps,
                progress_callback,
                total_steps,
                include_room_in_footer=False,
                force_break_by_room=True, # 开启强制分页
            )
            processed_steps += total_rows_summary

            # 2. 生成分考场表
            for room_name in room_order:
                # 创建新 Sheet，名称限制为 31 字符，且不能包含非法字符
                safe_name = (
                    str(room_name)[:30]
                    .replace(":", "")
                    .replace("\\", "")
                    .replace("/", "")
                    .replace("?", "")
                    .replace("*", "")
                    .replace("[", "")
                    .replace("]", "")
                )
                ws = wb.create_sheet(title=safe_name)

                room_data = exam_rooms[room_name]
                self.generate_sheet(
                    ws,
                    room_data,
                    processed_steps,
                    progress_callback,
                    total_steps,
                    include_room_in_footer=True,
                )

                # 更新进度计数
                rows_in_room = (len(room_data) + self.TEMPLATES_PER_ROW - 1) // self.TEMPLATES_PER_ROW
                processed_rows_step = rows_in_room
                processed_steps += processed_rows_step

        # 保存文件
        wb.save(output_path)
        return output_path

