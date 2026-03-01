
import openpyxl
from openpyxl.styles import Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.pagebreak import Break

class ExamBagLabelGenerator:
    """
    试卷袋标签生成器
    按照 3x3 格式批量生成试卷袋标签
    """
    def __init__(self, config):
        self.config = config
        
        # 样式定义
        self.font_bold = Font(name="宋体", size=14, bold=True)
        self.font_normal = Font(name="宋体", size=14, bold=False)
        self.alignment_style = Alignment(wrap_text=True, vertical="center", horizontal="left")
        self.thin_border = Border(
            left=Side(style="thin"),
            right=Side(style="thin"),
            top=Side(style="thin"),
            bottom=Side(style="thin"),
        )
        
        # 布局参数
        self.layout_rows = self.config.layout_rows # 3
        self.layout_cols = self.config.layout_cols # 3
        
        # 尺寸设置
        # 模板中 A=47.8, B=13, C=13 (很不均匀)
        # 这里为了通用性，采用均匀分布，尽量填满A4纸
        # A4 宽约 80-90 字符宽度。30 * 3 = 90
        self.col_width = 32
        self.row_height = 250 # 模板中的高度
        
    def generate(self, progress_callback=None):
        output_path = self.config.output_path
        data_list = self.config.student_data_list or []
        
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "试卷袋标签"
        
        # 页面设置
        self._setup_page_settings(ws)
        
        if not data_list:
            # 如果没有数据，生成一页空白
            self._fill_page_grid(ws, [], 1)
            wb.save(output_path)
            return output_path
            
        # 按科目分组
        subjects_map = {}
        # 保持科目顺序
        subject_order = []
        seen_subjects = set()
        
        for item in data_list:
            subj = item['subject']
            if subj not in subjects_map:
                subjects_map[subj] = []
                
            subjects_map[subj].append(item)
            
            if subj not in seen_subjects:
                subject_order.append(subj)
                seen_subjects.add(subj)
                
        current_row_base = 1
        total_subjects = len(subject_order)
        
        for idx, subj in enumerate(subject_order):
            if progress_callback:
                progress_callback(idx, total_subjects)
                
            items = subjects_map[subj]
            
            # 分页分块 (每页 9 个)
            capacity = self.layout_rows * self.layout_cols
            chunks = [items[i:i + capacity] for i in range(0, len(items), capacity)]
            
            for chunk_idx, chunk in enumerate(chunks):
                self._fill_page_grid(ws, chunk, current_row_base)
                
                current_row_base += self.layout_rows
                
                # 分页符逻辑：
                # 1. 同一科目内的分页
                # 2. 不同科目之间的强制分页
                # 只要不是(最后一个科目的 最后一页)，就加分页符
                is_last_subject = (idx == total_subjects - 1)
                is_last_chunk_in_subject = (chunk_idx == len(chunks) - 1)
                
                if not (is_last_subject and is_last_chunk_in_subject):
                    ws.row_breaks.append(Break(id=current_row_base - 1))
                    
        wb.save(output_path)
        return output_path

    def _setup_page_settings(self, ws):
        ws.page_setup.orientation = "portrait"
        ws.page_setup.paperSize = 9 # A4
        
        # 边距 (参考模板)
        ws.page_margins.left = 0.2
        ws.page_margins.right = 0.2
        ws.page_margins.top = 0.1
        ws.page_margins.bottom = 0.1
        ws.page_margins.header = 0.0
        ws.page_margins.footer = 0.0
        
        ws.print_options.horizontalCentered = True
        ws.print_options.verticalCentered = True

    def _fill_page_grid(self, ws, items, start_row):
        """填充单页网格"""
        # 设置行高
        for r in range(self.layout_rows):
            ws.row_dimensions[start_row + r].height = self.row_height
            
        # 设置列宽
        for c in range(self.layout_cols):
            col_letter = get_column_letter(c + 1)
            ws.column_dimensions[col_letter].width = self.col_width
            
        # 填充单元格
        for i in range(self.layout_rows * self.layout_cols):
            r = i // self.layout_cols
            c = i % self.layout_cols
            
            cell_row = start_row + r
            cell_col = c + 1
            cell = ws.cell(row=cell_row, column=cell_col)
            
            # 获取数据
            item = items[i] if i < len(items) else None
            
            if item:
                # 构造文本
                # 学校：xxx学校
                # 科目：语文
                # 考场：高二1班（42人）
                # 应到：
                # 实到：
                # 监考教师：
                # 考试情况：
                content = (
                    f"学校：{self.config.school_name}\n\n"
                    f"科目：{item['subject']}\n\n"
                    f"考场：{item['room']}（{item['count']}人）\n\n"
                    f"应到：\n\n"
                    f"实到：\n\n"
                    f"监考教师：\n\n"
                    f"考试情况："
                )
            else:
                # 空白标签（可选，这里留空或填默认格式）
                content = ""
            
            cell.value = content
            cell.font = self.font_bold
            cell.alignment = self.alignment_style
            cell.border = self.thin_border
