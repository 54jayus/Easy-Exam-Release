import openpyxl
import pandas as pd

class DataLoader:
    """
    负责从 Excel 读取考生数据并进行校验
    """
    
    @staticmethod
    def get_headers(file_path):
        """
        获取 Excel 文件的表头（第一行）
        """
        try:
            wb = openpyxl.load_workbook(file_path, read_only=True, data_only=True)
            ws = wb.active
            if ws.max_row < 1:
                return []
            
            # 读取第一行
            headers = []
            for cell in ws[1]:
                headers.append(cell.value)
            
            wb.close()
            # 过滤 None
            return [str(h) for h in headers if h is not None]
        except Exception as e:
            raise Exception(f"读取表头失败: {str(e)}")

    @staticmethod
    def load_data(file_path, mapping):
        """
        根据映射读取数据
        
        Args:
            file_path: Excel 文件路径
            mapping: 字段映射字典 {'目标字段': 'Excel列名'}
            
        Returns:
            list: 包含考生数据的字典列表
        """
        # 目标字段及其类型要求
        # type: 'str' (字符串, 保留原样), 'int' (数值)
        FIELD_CONFIG = {
            '考场号': 'str',
            '考场': 'str',
            '座位号': 'str',
            '考生姓名': 'str',
            '考生考号': 'str',
            '班级': 'int',
            '学号': 'int'
        }

        try:
            wb = openpyxl.load_workbook(file_path, read_only=True, data_only=True)
            ws = wb.active
            
            # 1. 构建 列名 -> 列索引 的映射
            header_map = {}
            headers = []
            for cell in ws[1]:
                headers.append(cell.value)
                
            for col_idx, header in enumerate(headers):
                if header:
                    header_map[str(header)] = col_idx # 0-based index

            # 2. 检查映射列是否存在
            col_indices = {}
            for target_field, excel_col in mapping.items():
                if excel_col not in header_map:
                    raise Exception(f"找不到列: {excel_col}")
                col_indices[target_field] = header_map[excel_col]

            data_list = []
            
            # 3. 逐行读取数据 (从第2行开始)
            # 使用 iter_rows 提高性能
            rows = ws.iter_rows(min_row=2, values_only=True)
            
            for row_idx, row_values in enumerate(rows, start=2):
                row_data = {}
                has_error = False
                
                # 检查是否为空行 (所有映射列都为空)
                all_empty = True
                for field in FIELD_CONFIG.keys():
                    col_idx = col_indices[field]
                    if col_idx < len(row_values) and row_values[col_idx] is not None:
                        all_empty = False
                        break
                
                if all_empty:
                    continue # 跳过空行

                for field, dtype in FIELD_CONFIG.items():
                    col_idx = col_indices[field]
                    val = row_values[col_idx] if col_idx < len(row_values) else None

                    # 非空校验
                    if val is None or str(val).strip() == "":
                        raise Exception(f"第 {row_idx} 行数据错误: '{field}' 不能为空")

                    # 类型处理
                    try:
                        if dtype == 'str':
                            # 强制转字符串，如果是浮点数形式的整数（如 1.0），尝试转回整数再转字符串
                            # 但 openpyxl data_only=True 读取时，如果是数字格式，会读成 int/float
                            if isinstance(val, float) and val.is_integer():
                                clean_val = str(int(val))
                            else:
                                clean_val = str(val)
                            row_data[field] = clean_val
                            
                        elif dtype == 'int':
                            # 转数值
                            clean_val = int(float(val)) # 处理 "1.0" 或 1.0
                            row_data[field] = clean_val
                            
                    except ValueError:
                        raise Exception(f"第 {row_idx} 行数据错误: '{field}' 格式不正确 (值: {val})")

                data_list.append(row_data)

            wb.close()
            
            if not data_list:
                raise Exception("未读取到有效数据")
                
            return data_list

        except Exception as e:
            # 捕获所有错误并向外抛出，由 UI 层处理显示
            raise e

    @staticmethod
    def load_exam_bag_data(file_path):
        """
        读取试卷袋标签数据
        格式：透视表（第一列为考场，后续列为科目，值为人数）
        返回：按科目分组的列表
        """
        try:
            df = pd.read_excel(file_path)

            if df.columns.size < 2:
                raise Exception("数据文件列数太少，无法解析")

            room_col = df.columns[0]
            subject_cols = list(df.columns[1:])

            result: list[dict] = []
            for subj in subject_cols:
                subject_name = "".join(str(subj).split())
                for _, row in df.iterrows():
                    room_val = row.get(room_col)
                    count_val = row.get(subj)

                    if room_val is None or str(room_val).strip() == "":
                        continue

                    if count_val is None or str(count_val).strip() == "":
                        continue

                    count = pd.to_numeric(count_val, errors="coerce")
                    if pd.isna(count):
                        continue

                    count_int = int(count)
                    if count_int <= 0:
                        continue

                    result.append(
                        {
                            "room": str(room_val).strip(),
                            "subject": subject_name,
                            "count": count_int,
                        }
                    )
            
            if not result:
                raise Exception("未提取到有效的考试数据（人数>0）")
                
            return result
            
        except Exception as e:
            raise Exception(f"读取试卷袋数据失败: {str(e)}")

    @staticmethod
    def load_student_info_data(file_path, mapping):
        mapping = dict(mapping or {})
        if "再选1" not in mapping and "选科1" in mapping:
            mapping["再选1"] = mapping.pop("选科1")
        if "再选2" not in mapping and "选科2" in mapping:
            mapping["再选2"] = mapping.pop("选科2")

        required_fields = [
            "考场号",
            "考场",
            "座位号",
            "考生姓名",
            "考生考号",
            "班级",
            "学号",
        ]
        optional_fields = ["首选", "再选1", "再选2"]

        field_config = {
            "考场号": "str",
            "考场": "str",
            "座位号": "str",
            "考生姓名": "str",
            "考生考号": "str",
            "班级": "int",
            "学号": "int",
            "首选": "str",
            "再选1": "str",
            "再选2": "str",
        }

        try:
            wb = openpyxl.load_workbook(file_path, read_only=True, data_only=True)
            ws = wb.active

            header_map = {}
            headers = []
            for cell in ws[1]:
                headers.append(cell.value)
            for col_idx, header in enumerate(headers):
                if header:
                    header_map[str(header)] = col_idx

            col_indices = {}
            for target_field, excel_col in mapping.items():
                if excel_col not in header_map:
                    raise Exception(f"找不到列: {excel_col}")
                col_indices[target_field] = header_map[excel_col]

            data_list = []
            rows = ws.iter_rows(min_row=2, values_only=True)

            for row_idx, row_values in enumerate(rows, start=2):
                all_empty = True
                for field in required_fields:
                    col_idx = col_indices[field]
                    if col_idx < len(row_values) and row_values[col_idx] is not None and str(row_values[col_idx]).strip() != "":
                        all_empty = False
                        break
                if all_empty:
                    continue

                row_data = {}
                for field in required_fields:
                    dtype = field_config[field]
                    col_idx = col_indices[field]
                    val = row_values[col_idx] if col_idx < len(row_values) else None
                    if val is None or str(val).strip() == "":
                        raise Exception(f"第 {row_idx} 行数据错误: '{field}' 不能为空")

                    if dtype == "str":
                        if isinstance(val, float) and val.is_integer():
                            row_data[field] = str(int(val))
                        else:
                            row_data[field] = str(val)
                    else:
                        row_data[field] = int(float(val))

                for field in optional_fields:
                    if field not in col_indices:
                        continue
                    col_idx = col_indices[field]
                    val = row_values[col_idx] if col_idx < len(row_values) else None
                    if val is None:
                        row_data[field] = ""
                    elif isinstance(val, float) and val.is_integer():
                        row_data[field] = str(int(val))
                    else:
                        row_data[field] = str(val)

                data_list.append(row_data)

            wb.close()

            if not data_list:
                raise Exception("未读取到有效数据")

            return data_list

        except Exception as e:
            raise e
