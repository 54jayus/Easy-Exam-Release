"""Shared utilities for generating Excel template instruction sheets.

All templates should use :func:`write_instruction_sheet` to produce a
consistent "填写说明" sheet with uniform formatting:
- Required columns: red background (#FFC7CE)
- Conditional-required columns: orange background (#FFE699)
- Optional columns: white background
"""
from __future__ import annotations

from typing import Any, Sequence


def write_instruction_sheet(
    writer: Any,
    columns: Sequence[str],
    instructions: dict[str, str],
    *,
    required_cols: set[str],
    conditional_cols: set[str] | None = None,
    sheet_name: str = "填写说明",
    row_height: int = 110,
) -> None:
    """Write a "填写说明" sheet into an open ``xlsxwriter`` ExcelWriter.

    Parameters
    ----------
    writer : xlsxwriter.ExcelWriter
        The open Excel writer instance.
    columns : sequence of str
        Column names in display order.
    instructions : dict
        Mapping of column name -> instruction text.
    required_cols : set
        Column names that are mandatory (red highlight).
    conditional_cols : set, optional
        Column names that are conditionally required (orange highlight).
    sheet_name : str
        Name of the instruction sheet.
    row_height : int
        Height of the instruction text row.
    """
    conditional_cols = conditional_cols or set()
    optional_cols = {c for c in columns if c not in required_cols and c not in conditional_cols}

    import pandas as pd
    instruction_row = [{col: instructions.get(col, "") for col in columns}]
    df_desc = pd.DataFrame(instruction_row)
    df_desc.to_excel(writer, sheet_name=sheet_name, index=False)

    workbook = writer.book
    ws = writer.sheets[sheet_name]

    # Formats
    wrap_left = workbook.add_format({"text_wrap": True, "align": "left", "valign": "top"})
    required_cell = workbook.add_format(
        {"text_wrap": True, "align": "left", "valign": "top", "bg_color": "#FFC7CE"}
    )
    conditional_cell = workbook.add_format(
        {"text_wrap": True, "align": "left", "valign": "top", "bg_color": "#FFE699"}
    )

    required_header = workbook.add_format(
        {"text_wrap": True, "align": "center", "valign": "vcenter",
         "bg_color": "#FFC7CE", "bold": True, "border": 1}
    )
    conditional_header = workbook.add_format(
        {"text_wrap": True, "align": "center", "valign": "vcenter",
         "bg_color": "#FFE699", "bold": True, "border": 1}
    )
    normal_header = workbook.add_format(
        {"text_wrap": True, "align": "center", "valign": "vcenter",
         "bold": True, "border": 1}
    )

    # Write headers and instruction cells
    for idx, col in enumerate(columns):
        value = instructions.get(col, "")
        if col in required_cols:
            ws.write(0, idx, col, required_header)
            ws.write(1, idx, value, required_cell)
        elif col in conditional_cols:
            ws.write(0, idx, col, conditional_header)
            ws.write(1, idx, value, conditional_cell)
        else:
            ws.write(0, idx, col, normal_header)
            ws.write(1, idx, value, wrap_left)

    # Set column widths (approximate: 3 * max(header_len, instruction_lines))
    for idx, col in enumerate(columns):
        width = max(len(col) * 3, 20)
        ws.set_column(idx, idx, min(width, 50), wrap_left)

    ws.set_row(1, row_height)


def write_instruction_sheet_openpyxl(
    workbook: Any,
    columns: Sequence[str],
    instructions: dict[str, str],
    *,
    required_cols: set[str],
    conditional_cols: set[str] | None = None,
    sheet_name: str = "填写说明",
    row_height: int = None,
) -> None:
    """Write a "填写说明" sheet into an open ``openpyxl`` Workbook.

    This variant is for templates that use openpyxl instead of xlsxwriter.

    Parameters
    ----------
    workbook : openpyxl.Workbook
        The open workbook instance.
    columns : sequence of str
        Column names in display order.
    instructions : dict
        Mapping of column name -> instruction text.
    required_cols : set
        Column names that are mandatory (red highlight).
    conditional_cols : set, optional
        Column names that are conditionally required (orange highlight).
    sheet_name : str
        Name of the instruction sheet.
    row_height : int, optional
        Height of the instruction text row; auto-calculated if None.
    """
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
    from openpyxl.utils import get_column_letter

    conditional_cols = conditional_cols or set()

    ws = workbook.create_sheet(sheet_name)

    thin = Side(style="thin")
    header_border = Border(left=thin, right=thin, top=thin, bottom=thin)
    header_align = Alignment(horizontal="center", vertical="center", wrap_text=True)

    required_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
    conditional_fill = PatternFill(start_color="FFE699", end_color="FFE699", fill_type="solid")
    normal_header_font = Font(bold=True)
    wrap_left = Alignment(horizontal="left", vertical="top", wrap_text=True)

    for idx, col in enumerate(columns, start=1):
        cell = ws.cell(row=1, column=idx, value=col)
        cell.font = normal_header_font
        cell.alignment = header_align
        cell.border = header_border
        if col in required_cols:
            cell.fill = required_fill
        elif col in conditional_cols:
            cell.fill = conditional_fill

    for idx, col in enumerate(columns, start=1):
        value = instructions.get(col, "")
        cell = ws.cell(row=2, column=idx, value=value)
        cell.alignment = wrap_left
        if col in required_cols:
            cell.fill = required_fill
        elif col in conditional_cols:
            cell.fill = conditional_fill

    # Auto-calculate row height
    if row_height is None:
        max_lines = max((instructions.get(col, "").count("\n") + 1) for col in columns) if columns else 1
        row_height = max_lines * 25 + 8
    ws.row_dimensions[2].height = row_height

    for idx in range(1, len(columns) + 1):
        letter = get_column_letter(idx)
        ws.column_dimensions[letter].width = 20
