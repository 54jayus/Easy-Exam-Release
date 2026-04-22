from __future__ import annotations

from datetime import date

import pandas as pd
from openpyxl import load_workbook

from backend.examroom.core.arrangement import ExamArrangement
from backend.examroom.core.subject_strategy import (
    assign_large_groups,
    assign_remaining_students,
    group_and_sort_subjects,
    initialize_rooms,
    reduce_mixed_rooms,
)


def test_get_room_capacity_supports_zero_padded_and_int_keys() -> None:
    arrangement = ExamArrangement(
        "students.xlsx",
        max_students_per_room=30,
        room_capacities={"001": 35, 2: 28, "003": 40},
    )

    assert arrangement.get_room_capacity("001") == 35
    assert arrangement.get_room_capacity("1") == 35
    assert arrangement.get_room_capacity("002") == 28
    assert arrangement.get_room_capacity(3) == 40
    assert arrangement.get_room_capacity("999") == 30


def test_validate_subject_column_normalizes_full_names_and_separators() -> None:
    arrangement = ExamArrangement("students.xlsx", arrangement_mode="subject_mode")
    arrangement.students = pd.DataFrame(
        [
            {"班级": "1", "学号": "01", "考号": "240001", "姓名": "张三", "选科": "物理+化学+生物"},
            {"班级": "1", "学号": "02", "考号": "240002", "姓名": "李四", "选科": "历史 地理 政治"},
            {"班级": "1", "学号": "03", "考号": "240003", "姓名": "王五", "选科": "理化生"},
        ]
    )

    ok, message = arrangement.validate_subject_column()

    assert ok is True
    assert message == "选科列校验通过"
    assert arrangement.students["选科"].tolist() == ["物化生", "史地政", "物化生"]


def test_check_required_columns_rejects_non_numeric_class_and_student_no() -> None:
    arrangement = ExamArrangement("students.xlsx", arrangement_mode="normal_mode")
    arrangement.students = pd.DataFrame(
        [{"班级": "高一1班", "学号": "01A", "考号": "240001", "姓名": "张三"}]
    )

    ok, message = arrangement.check_required_columns()

    assert ok is False
    assert "只能填写数字" in message


def test_parse_subject_combination_accepts_compact_and_delimited_forms() -> None:
    arrangement = ExamArrangement("students.xlsx")

    assert arrangement.parse_subject_combination("物化生") == ["物理", "化学", "生物"]
    assert arrangement.parse_subject_combination("历史/地理/政治") == ["历史", "地理", "政治"]


def test_format_subject_time_uses_default_settings_and_handles_invalid_values() -> None:
    arrangement = ExamArrangement("students.xlsx", arrangement_mode="gaokao_mode")
    today = date.today()
    expected_prefix = f"{today.month}月{today.day}日"

    assert arrangement._format_subject_time("语文") == f"{expected_prefix}09:00-11:30"
    assert arrangement._format_subject_time("化学", is_self_study=True) == f"{expected_prefix}08:30-09:45"

    arrangement.gaokao_time_settings = {
        "examTimes": {"语文": {"date": "", "startTime": "09:00", "endTime": "11:30"}},
        "selfStudyTimes": {},
    }
    assert arrangement._format_subject_time("语文") == ""


def test_get_room_name_prefers_room_setting_dataframe() -> None:
    arrangement = ExamArrangement("students.xlsx")
    arrangement.room_setting_df = pd.DataFrame(
        [
            {"考场号": "001", "考场": "第一考场"},
            {"考场号": "002", "考场": "第二考场"},
        ]
    )

    assert arrangement._get_room_name("001") == "第一考场"
    assert arrangement._get_room_name("2") == "第二考场"
    assert arrangement._get_room_name("003") == "第003考场"


def test_get_room_list_prefers_dataframe_then_falls_back_to_sequence() -> None:
    arrangement = ExamArrangement("students.xlsx", total_rooms=3)
    arrangement.room_setting_df = pd.DataFrame([{"考场号": "003"}, {"考场号": "001"}])

    assert arrangement._get_room_list() == ["003", "001"]

    delattr(arrangement, "room_setting_df")
    arrangement.room_setting_data = pd.DataFrame([{"考场号": "005"}, {"考场号": "004"}])
    assert arrangement._get_room_list() == ["005", "004"]

    arrangement.room_setting_data = None
    assert arrangement._get_room_list() == ["1", "2", "3"]


def test_fill_rooms_sequential_respects_capacity_and_start_index() -> None:
    arrangement = ExamArrangement(
        "students.xlsx",
        total_rooms=3,
        room_capacities={"001": 2, "002": 1, "003": 2},
    )
    arrangement.room_setting_df = pd.DataFrame([{"考场号": "001"}, {"考场号": "002"}, {"考场号": "003"}])
    students = pd.DataFrame(
        [
            {"考号": "240001", "姓名": "张三"},
            {"考号": "240002", "姓名": "李四"},
            {"考号": "240003", "姓名": "王五"},
        ]
    )

    rooms, last_index = arrangement._fill_rooms_sequential(students, start_room_index=0)

    assert [room["room_num"] for room in rooms] == ["001", "002"]
    assert [len(room["students"]) for room in rooms] == [2, 1]
    assert last_index == 1


def test_fill_rooms_sequential_raises_when_rooms_are_insufficient() -> None:
    arrangement = ExamArrangement(
        "students.xlsx",
        total_rooms=1,
        room_capacities={"001": 1},
    )
    arrangement.room_setting_df = pd.DataFrame([{"考场号": "001"}])
    students = pd.DataFrame(
        [
            {"考号": "240001", "姓名": "张三"},
            {"考号": "240002", "姓名": "李四"},
        ]
    )

    try:
        arrangement._fill_rooms_sequential(students, start_room_index=0)
    except ValueError as exc:
        assert "考场数量不足" in str(exc)
    else:
        raise AssertionError("expected ValueError when rooms are insufficient")


def test_extract_subject_from_combination_checks_subject_abbreviation() -> None:
    arrangement = ExamArrangement("students.xlsx")

    assert arrangement._extract_subject_from_combination("物化生", "化") is True
    assert arrangement._extract_subject_from_combination("史地政", "化") is False


def test_arrange_normal_mode_assigns_students_by_room_capacity() -> None:
    arrangement = ExamArrangement(
        "students.xlsx",
        arrangement_mode="normal_mode",
        total_rooms=2,
        room_capacities={"001": 2, "002": 2},
    )
    arrangement.room_setting_data = {"001": "第一考场", "002": "第二考场"}
    arrangement.students = pd.DataFrame(
        [
            {"班级": "1", "学号": "01", "考号": "240001", "姓名": "张三"},
            {"班级": "1", "学号": "02", "考号": "240002", "姓名": "李四"},
            {"班级": "1", "学号": "03", "考号": "240003", "姓名": "王五"},
        ]
    )

    ok, message = arrangement.arrange_normal_mode()

    assert ok is True
    assert "顺序编排完成" in message
    assert arrangement.arranged_students["考场号"].tolist() == ["001", "001", "002"]
    assert arrangement.arranged_students["座位号"].tolist() == ["01", "02", "01"]
    assert arrangement.arranged_students["考场"].tolist() == ["第一考场", "第一考场", "第二考场"]


def test_arrange_subject_mode_keeps_subject_groups_and_generates_split_columns() -> None:
    arrangement = ExamArrangement(
        "students.xlsx",
        arrangement_mode="subject_mode",
        total_rooms=3,
        room_capacities={"001": 2, "002": 2, "003": 2},
    )
    arrangement.room_setting_data = {"001": "第一考场", "002": "第二考场", "003": "第三考场"}
    arrangement.students = pd.DataFrame(
        [
            {"班级": "1", "学号": "01", "考号": "240001", "姓名": "张三", "选科": "物化生"},
            {"班级": "1", "学号": "02", "考号": "240002", "姓名": "李四", "选科": "物化生"},
            {"班级": "1", "学号": "03", "考号": "240003", "姓名": "王五", "选科": "物化生"},
            {"班级": "1", "学号": "04", "考号": "240004", "姓名": "赵六", "选科": "物生地"},
            {"班级": "1", "学号": "05", "考号": "240005", "姓名": "孙七", "选科": "史地政"},
            {"班级": "1", "学号": "06", "考号": "240006", "姓名": "周八", "选科": "史地政"},
        ]
    )

    ok, message = arrangement.arrange_subject_mode()

    assert ok is True
    assert "考场编排完成" in message
    assert arrangement.arranged_students["考场号"].tolist() == ["001", "001", "002", "002", "003", "003"]
    assert arrangement.arranged_students["座位号"].tolist() == ["01", "02", "01", "02", "01", "02"]
    assert arrangement.arranged_students["考场"].tolist() == ["第一考场", "第一考场", "第二考场", "第二考场", "第三考场", "第三考场"]
    assert arrangement.arranged_students["考场选科组合"].tolist() == [
        "物化生",
        "物化生",
        "物化生, 物生地",
        "物化生, 物生地",
        "史地政",
        "史地政",
    ]
    assert arrangement.arranged_students["首选"].tolist() == ["物理", "物理", "物理", "物理", "历史", "历史"]


def test_reduce_mixed_rooms_allows_underfilled_room_to_reduce_mixed_room_count() -> None:
    arrangement = ExamArrangement(
        "students.xlsx",
        arrangement_mode="subject_mode",
        total_rooms=3,
        room_capacities={"001": 5, "002": 5, "003": 5},
    )
    arrangement.room_setting_data = {"001": "第一考场", "002": "第二考场", "003": "第三考场"}

    students = []
    specs = [("物化生", 5), ("物生地", 4), ("物化地", 2), ("物政生", 2)]
    student_index = 1
    for subject, count in specs:
        for _ in range(count):
            students.append(
                {
                    "班级": "1",
                    "学号": f"{student_index:02d}",
                    "考号": f"2400{student_index:02d}",
                    "姓名": f"学生{student_index}",
                    "选科": subject,
                }
            )
            student_index += 1

    arrangement.students = pd.DataFrame(students)

    rooms = initialize_rooms(arrangement)
    physics_subjects, history_subjects = group_and_sort_subjects(arrangement)
    current_room_index, remaining_students = assign_large_groups(arrangement, rooms, physics_subjects, history_subjects)
    assign_remaining_students(arrangement, rooms, remaining_students, current_room_index)

    mixed_before = sum(1 for room in rooms if len(room["subjects"]) > 1)
    assert mixed_before == 2

    reduce_mixed_rooms(arrangement, rooms)

    mixed_after = sum(1 for room in rooms if len(room["subjects"]) > 1)
    room_sizes = [len(room["students"]) for room in rooms if room["students"]]

    assert mixed_after == 1
    assert sorted(room_sizes) == [4, 4, 5]

    ok, message = arrangement._generate_results(rooms)

    assert ok is True
    assert "考场编排完成" in message
    assert len(arrangement.arranged_students) == len(students)
    assert int((arrangement.arranged_students.groupby("考场号")["选科"].nunique() > 1).sum()) == 1


def test_save_results_exports_student_sheet_and_stats_sheet(tmp_path) -> None:
    arrangement = ExamArrangement("students.xlsx", arrangement_mode="subject_mode")
    arrangement.arranged_students = pd.DataFrame(
        [
            {
                "班级": "1",
                "学号": "01",
                "姓名": "张三",
                "考号": "240001",
                "选科": "物化生",
                "考场": "第一考场",
                "考场号": "001",
                "座位号": "01",
                "考场选科组合": "物化生",
            },
            {
                "班级": "1",
                "学号": "02",
                "姓名": "李四",
                "考号": "240002",
                "选科": "史地政",
                "考场": "第二考场",
                "考场号": "002",
                "座位号": "01",
                "考场选科组合": "史地政",
            },
        ]
    )
    arrangement.room_setting_df = pd.DataFrame([{"考场": "第一考场"}, {"考场": "第二考场"}])
    output_path = tmp_path / "subject-results.xlsx"

    ok, message = arrangement.save_results(str(output_path))

    assert ok is True
    assert "编排结果已保存到" in message
    workbook = load_workbook(output_path, read_only=True)
    assert workbook.sheetnames == ["学生编排结果", "考场选科统计"]
