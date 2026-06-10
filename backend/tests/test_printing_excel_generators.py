from __future__ import annotations

import openpyxl

from backend.printing.core.config import DeskLabelConfig, ExamBagLabelConfig, StudentInfoTableConfig
from backend.printing.core.generators.excel.desk_label import DeskLabelGenerator
from backend.printing.core.generators.excel.exam_bag_label import ExamBagLabelGenerator
from backend.printing.core.generators.excel.student_info_table import StudentInfoTableGenerator


def test_student_info_blank_examroom_template_includes_subject_columns(tmp_path) -> None:
    path = tmp_path / "student-info-blank.xlsx"

    StudentInfoTableGenerator(
        StudentInfoTableConfig(
            output_path=str(path),
            include_subject_fields=True,
            group_mode="examroom",
            title="名单",
        )
    ).generate()

    wb = openpyxl.load_workbook(path)
    ws = wb["总表"]

    assert wb.sheetnames == ["总表"]
    assert ws.print_area == "'总表'!$A$1:$J$44"
    assert [ws.cell(2, col).value for col in range(1, 11)] == [
        "班级",
        "学号",
        "考生姓名",
        "考生考号",
        "首选",
        "选科1",
        "选科2",
        "考场",
        "考场号",
        "座位号",
    ]


def test_student_info_helpers_sanitize_sheet_names_and_sort_examroom_rows(tmp_path) -> None:
    generator = StudentInfoTableGenerator(
        StudentInfoTableConfig(output_path=str(tmp_path / "student-info.xlsx"), group_mode="examroom")
    )

    assert generator._safe_sheet_name("Room/1:*?[]") == "Room 1"

    wb = openpyxl.Workbook()
    wb.active.title = "Room 1"
    assert generator._unique_sheet_name(wb, "Room/1") == "Room 1(2)"

    grouped = generator._group_by_examroom(
        [
            {"考场号": "001", "座位号": "02", "班级": "2", "学号": "02"},
            {"考场号": "001", "座位号": "01", "班级": "1", "学号": "03"},
            {"考场号": "001", "座位号": "01", "班级": "1", "学号": "02"},
            {"考场号": "002", "座位号": "01", "班级": "1", "学号": "01"},
        ]
    )

    assert [item["学号"] for item in grouped["001"]] == ["02", "03", "02"]


def test_exam_bag_blank_generation_creates_empty_3x3_grid(tmp_path) -> None:
    path = tmp_path / "exam-bag-blank.xlsx"

    ExamBagLabelGenerator(ExamBagLabelConfig(output_path=str(path))).generate()

    wb = openpyxl.load_workbook(path)
    ws = wb["试卷袋标签"]

    assert wb.sheetnames == ["试卷袋标签"]
    assert ws.max_row == 3
    assert ws.max_column == 3
    assert ws["A1"].value is None
    assert ws["C3"].value is None


def test_exam_bag_generation_adds_page_breaks_between_chunks_and_subjects(tmp_path) -> None:
    path = tmp_path / "exam-bag.xlsx"
    data = [{"subject": "SUB1", "room": f"R{i}", "count": i} for i in range(10)]
    data.append({"subject": "SUB2", "room": "RX", "count": 99})

    ExamBagLabelGenerator(
        ExamBagLabelConfig(output_path=str(path), school_name="School", student_data_list=data)
    ).generate()

    wb = openpyxl.load_workbook(path)
    ws = wb["试卷袋标签"]

    assert [b.id for b in ws.row_breaks.brk] == [3, 6]
    assert "科目：SUB1" in ws["A1"].value
    assert "考场：R9（9人）" in ws["A4"].value
    assert "科目：SUB2" in ws["A7"].value


def test_desk_label_blank_generation_adds_page_breaks(tmp_path) -> None:
    path = tmp_path / "desk-label.xlsx"

    DeskLabelGenerator(
        DeskLabelConfig(output_path=str(path), total_count=5, layout_rows=2, layout_cols=2)
    ).generate()

    wb = openpyxl.load_workbook(path)
    ws = wb["桌角纸（批量打印）"]

    assert [b.id for b in ws.row_breaks.brk] == [2]
    assert "姓名：" in ws["A1"].value
    assert "座位号：" in ws["B1"].value


def test_desk_label_seat_mapping_respects_start_position_and_custom_column_counts() -> None:
    default_generator = DeskLabelGenerator(
        DeskLabelConfig(output_path="unused.xlsx", layout_rows=2, layout_cols=2)
    )
    custom_generator = DeskLabelGenerator(
        DeskLabelConfig(output_path="unused.xlsx", layout_rows=3, layout_cols=2, custom_col_counts=[3, 1])
    )

    assert default_generator._get_seat_mapping(
        2, 2, default_generator.config.layout_pattern, default_generator.config.start_pos
    ) == {
        0: (0, 1),
        1: (0, 0),
        2: (1, 0),
        3: (1, 1),
    }
    assert custom_generator._get_seat_mapping(
        3, 2, custom_generator.config.layout_pattern, custom_generator.config.start_pos
    ) == {
        0: (0, 1),
        1: (0, 0),
        2: (1, 0),
        3: (2, 0),
    }
