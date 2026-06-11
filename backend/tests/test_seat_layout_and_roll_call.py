from __future__ import annotations

from types import SimpleNamespace

import openpyxl
import pandas as pd
import pytest

from backend.printing.core.adapters.roll_call import (
    build_blank_roll_call_groups,
    build_roll_call_groups,
    build_roll_call_groups_from_students,
)
from backend.printing.core.config import RollCallConfig
from backend.printing.core.generators.excel.roll_call import RollCallGenerator
from backend.printing.core.generators.pdf.roll_call_pdf import RollCallPDFGenerator
from backend.printing.core.generators.roll_call_text import format_class_name
from backend.printing.core.seat_layout import (
    get_seat_mapping,
    layout_capacity,
    layout_for_room,
    mirror_layout_start_pos,
    normalize_seat_layout,
)


@pytest.mark.parametrize(
    ("pattern", "expected"),
    [
        ("Z型横排", {1: (0, 0), 2: (0, 1), 3: (1, 0), 4: (1, 1)}),
        ("S型横排", {1: (0, 0), 2: (0, 1), 3: (1, 1), 4: (1, 0)}),
        ("Z型竖排", {1: (0, 0), 2: (1, 0), 3: (0, 1), 4: (1, 1)}),
        ("S型竖排", {1: (0, 0), 2: (1, 0), 3: (1, 1), 4: (0, 1)}),
    ],
)
def test_seat_mapping_supports_all_patterns(pattern, expected) -> None:
    layout = {"layoutRows": 2, "layoutCols": 2, "layoutPattern": pattern, "startPos": "right"}
    assert get_seat_mapping(layout) == expected


def test_roll_call_class_name_has_label() -> None:
    assert format_class_name("1班") == "班级：1班"
    assert format_class_name(3) == "班级：3"
    assert format_class_name("") == ""


def test_seat_mapping_honors_left_start_and_custom_columns() -> None:
    layout = {
        "layoutRows": 3,
        "layoutCols": 2,
        "layoutPattern": "Z型竖排",
        "startPos": "left",
        "customColCounts": [2, 3],
    }

    assert get_seat_mapping(layout) == {1: (0, 1), 2: (1, 1), 3: (2, 1), 4: (0, 0), 5: (1, 0)}
    assert layout_capacity(layout) == 5


def test_room_override_falls_back_to_default() -> None:
    config = normalize_seat_layout({
        "defaultLayout": {"layoutRows": 5, "layoutCols": 6},
        "roomOverrides": {"002": {"layoutRows": 7, "layoutCols": 6}},
    })

    assert layout_for_room(config, "001")["layoutRows"] == 5
    assert layout_for_room(config, "002")["layoutRows"] == 7


def test_mirror_layout_start_pos_flips_view() -> None:
    layout = {"layoutRows": 2, "layoutCols": 2, "layoutPattern": "Z型横排", "startPos": "left"}

    mirrored = mirror_layout_start_pos(layout, True)

    assert mirrored["startPos"] == "right"
    assert get_seat_mapping(mirrored)[1] == (0, 0)


def test_roll_call_regular_builds_one_group_per_subject_and_room() -> None:
    arrangement = SimpleNamespace(
        arrangement_mode="normal_mode",
        arranged_students=pd.DataFrame([
            {"姓名": "张三", "考号": "1", "班级": "1", "考场": "一考场", "考场号": "001", "座位号": "1"},
            {"姓名": "李四", "考号": "2", "班级": "1", "考场": "二考场", "考场号": "002", "座位号": "1"},
        ]),
    )
    seat_layout = {"defaultLayout": {"layoutRows": 5, "layoutCols": 6}, "roomOverrides": {}}

    groups = build_roll_call_groups(arrangement, [{"name": "语文"}, {"name": "数学"}], seat_layout)

    assert [(item["subject"], item["roomNo"]) for item in groups] == [
        ("语文", "001"), ("语文", "002"), ("数学", "001"), ("数学", "002")
    ]


def test_roll_call_subject_mode_groups_mixed_rooms_by_exam_session() -> None:
    arrangement = SimpleNamespace(
        arrangement_mode="subject_mode",
        arranged_students=pd.DataFrame([
            {"姓名": "张三", "考号": "1", "考场": "一考场", "考场号": "001", "座位号": "1", "首选": "物理", "选科1": "化学", "选科2": "生物"},
            {"姓名": "李四", "考号": "2", "考场": "一考场", "考场号": "001", "座位号": "2", "首选": "历史", "选科1": "政治", "选科2": "地理"},
        ]),
    )

    groups = build_roll_call_groups(arrangement, [], {"defaultLayout": {"layoutRows": 5, "layoutCols": 6}})
    assert [item["subject"] for item in groups] == [
        "语文", "数学", "英语", "首选科目场次", "再选科目一场次", "再选科目二场次",
    ]
    first_choice = groups[3]
    assert first_choice["subjectLabel"] == "场次"
    assert [(student["name"], student["examSubject"]) for student in first_choice["students"]] == [
        ("张三", "物理"), ("李四", "历史"),
    ]
    assert [(student["name"], student["examSubject"]) for student in groups[4]["students"]] == [
        ("张三", "化学"), ("李四", "政治"),
    ]


def test_roll_call_rejects_duplicate_and_out_of_range_seats() -> None:
    arrangement = SimpleNamespace(
        arrangement_mode="normal_mode",
        arranged_students=pd.DataFrame([
            {"姓名": "张三", "考场": "一考场", "考场号": "001", "座位号": "2"},
            {"姓名": "李四", "考场": "一考场", "考场号": "001", "座位号": "2"},
        ]),
    )

    with pytest.raises(ValueError, match="重复"):
        build_roll_call_groups(arrangement, [{"name": "语文"}], {"defaultLayout": {"layoutRows": 1, "layoutCols": 1}})


def test_roll_call_import_groups_students_by_room() -> None:
    groups = build_roll_call_groups_from_students(
        [
            {"考生姓名": "张三", "考生考号": "240001", "班级": 1, "考场": "第一考场", "考场号": "001", "座位号": "2"},
            {"考生姓名": "李四", "考生考号": "240002", "班级": 2, "考场": "第一考场", "考场号": "001", "座位号": "1"},
            {"考生姓名": "王五", "考生考号": "240003", "班级": 3, "考场": "第二考场", "考场号": "002", "座位号": "1"},
        ],
        {"defaultLayout": {"layoutRows": 5, "layoutCols": 6}},
    )

    assert [group["roomNo"] for group in groups] == ["001", "002"]
    assert [student["name"] for student in groups[0]["students"]] == ["李四", "张三"]
    assert groups[0]["students"][0]["examNo"] == "240002"


def test_blank_roll_call_groups_use_current_layout() -> None:
    groups = build_blank_roll_call_groups(
        2,
        {"defaultLayout": {"layoutRows": 7, "layoutCols": 6, "layoutPattern": "S型竖排", "startPos": "left"}},
    )

    assert len(groups) == 2
    assert groups[0]["students"] == []
    assert groups[0]["seatLayout"]["layoutRows"] == 7


def test_roll_call_generators_create_valid_files(tmp_path) -> None:
    groups = [{
        "subject": "语文",
        "roomName": "第一考场",
        "roomNo": "001",
        "students": [{"name": "张三", "examNo": "240001", "className": "1班", "seatNo": 1}],
        "seatLayout": {"layoutRows": 5, "layoutCols": 6, "layoutPattern": "S型竖排", "startPos": "left"},
    }]
    xlsx_path = tmp_path / "点名表.xlsx"
    pdf_path = tmp_path / "点名表.pdf"
    base = dict(
        exam_name="期末考试点名表",
        school_name="第一中学",
        groups=groups,
        instructions="1.缺考打勾",
        show_class=True,
    )

    RollCallGenerator(RollCallConfig(output_path=str(xlsx_path), export_xlsx=True, **base)).generate()
    RollCallPDFGenerator(RollCallConfig(output_path=str(pdf_path), export_xlsx=False, export_pdf=True, **base)).generate()

    wb = openpyxl.load_workbook(xlsx_path)
    assert wb.sheetnames == ["语文"]
    ws = wb["语文"]
    assert ws["A1"].value == "期末考试点名表"
    assert ws["A1"].font.sz == 18
    assert ws.row_dimensions[4].height > 70
    footer_gap_row = 4 + groups[0]["seatLayout"]["layoutRows"]
    box_start_row = footer_gap_row + 1
    box_end_row = box_start_row + 2
    assert ws.row_dimensions[footer_gap_row].height == 28
    assert ws.page_setup.fitToWidth == 1
    assert ws.page_setup.fitToHeight == 0
    assert ws.page_margins.left == pytest.approx(12 / 25.4)
    assert ws.sheet_view.showGridLines is False
    assert ws.cell(box_start_row, 1).border.left.style == "thin"
    assert ws.cell(box_end_row, 4).border.right.style == "thin"
    assert ws.cell(box_end_row, 4).border.bottom.style == "thin"
    assert ws.cell(box_start_row, 5).border.left.style is None
    assert ws.cell(box_start_row, 5).border.top.style is None
    assert ws.cell(box_end_row, 6).border.right.style is None
    assert ws.cell(box_end_row, 6).border.bottom.style is None
    student_cells = [
        str(cell.value)
        for row in ws.iter_rows(min_row=4, max_row=10, min_col=1, max_col=6)
        for cell in row
        if cell.value and "张三" in str(cell.value)
    ]
    assert student_cells == ["1. 张三\n240001\n班级：1班\n□ 缺考"]
    assert xlsx_path.read_bytes()[:2] == b"PK"
    assert pdf_path.read_bytes()[:5] == b"%PDF-"


def test_roll_call_generator_can_mirror_view(tmp_path) -> None:
    groups = [{
        "subject": "语文",
        "roomName": "第一考场",
        "roomNo": "001",
        "students": [{"name": "张三", "examNo": "240001", "className": "1班", "seatNo": 1}],
        "seatLayout": {"layoutRows": 2, "layoutCols": 2, "layoutPattern": "Z型横排", "startPos": "left"},
    }]
    xlsx_path = tmp_path / "点名表-镜像.xlsx"

    RollCallGenerator(
        RollCallConfig(output_path=str(xlsx_path), export_xlsx=True, groups=groups, mirror_view=True)
    ).generate()

    wb = openpyxl.load_workbook(xlsx_path)
    assert "张三" in str(wb["语文"]["A4"].value)


def test_roll_call_generator_prints_session_and_student_subjects(tmp_path) -> None:
    groups = [{
        "subject": "首选科目场次",
        "subjectLabel": "场次",
        "roomName": "第一考场",
        "roomNo": "001",
        "students": [
            {"name": "张三", "examNo": "240001", "className": "1班", "examSubject": "物理", "seatNo": 1},
            {"name": "李四", "examNo": "240002", "className": "2班", "examSubject": "历史", "seatNo": 2},
        ],
        "seatLayout": {"layoutRows": 7, "layoutCols": 6, "layoutPattern": "Z型横排", "startPos": "right"},
    }]
    xlsx_path = tmp_path / "点名表-混合考场.xlsx"

    RollCallGenerator(RollCallConfig(output_path=str(xlsx_path), groups=groups, show_class=True)).generate()

    ws = openpyxl.load_workbook(xlsx_path)["首选科目场次"]
    assert "场次：首选科目场次" in str(ws["A2"].value)
    assert "科目：物理" in str(ws["A4"].value)
    assert "科目：历史" in str(ws["B4"].value)
    assert ws["A4"].font.sz <= 9


def test_roll_call_excel_keeps_each_room_on_its_own_print_page(tmp_path) -> None:
    groups = [
        {
            "subject": "语文",
            "roomName": f"第{room_no}考场",
            "roomNo": f"{room_no:03d}",
            "students": [{"name": f"学生{room_no}", "examNo": f"24000{room_no}", "seatNo": 1}],
            "seatLayout": {"layoutRows": 7, "layoutCols": 6, "layoutPattern": "S型竖排", "startPos": "left"},
        }
        for room_no in (1, 2)
    ]
    xlsx_path = tmp_path / "点名表-多考场.xlsx"

    RollCallGenerator(RollCallConfig(output_path=str(xlsx_path), groups=groups)).generate()

    ws = openpyxl.load_workbook(xlsx_path)["语文"]
    assert len(ws.row_breaks.brk) == 1
    assert ws.page_setup.fitToWidth == 1
    assert ws.page_setup.fitToHeight == 0
    assert ws.print_area == "'语文'!$A$1:$F$28"


def test_roll_call_generator_honors_portrait_orientation(tmp_path) -> None:
    groups = [{
        "subject": "语文",
        "roomName": "第一考场",
        "roomNo": "001",
        "students": [{"name": "张三", "examNo": "240001", "className": "1班", "seatNo": 1}],
        "seatLayout": {"layoutRows": 6, "layoutCols": 9, "layoutPattern": "S型竖排", "startPos": "left"},
    }]
    xlsx_path = tmp_path / "点名表-纵向.xlsx"

    RollCallGenerator(
        RollCallConfig(output_path=str(xlsx_path), export_xlsx=True, groups=groups, orientation="portrait")
    ).generate()

    wb = openpyxl.load_workbook(xlsx_path)
    assert wb["语文"].page_setup.orientation == "portrait"
