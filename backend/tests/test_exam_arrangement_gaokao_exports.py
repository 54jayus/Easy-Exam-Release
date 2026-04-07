from __future__ import annotations

from copy import deepcopy

import pandas as pd
from openpyxl import load_workbook

from backend.examroom.core.arrangement import ExamArrangement
from backend.examroom.core.gaokao_defaults import GAOKAO_TIME_DEFAULTS


def _build_gaokao_arrangement() -> ExamArrangement:
    arrangement = ExamArrangement("students.xlsx", total_rooms=2, arrangement_mode="gaokao_mode")
    arrangement.subject_column = "选科"
    arrangement.room_setting_df = pd.DataFrame(
        [
            {"考场号": "001", "考场": "第一考场"},
            {"考场号": "002", "考场": "第二考场"},
        ]
    )
    arrangement.gaokao_time_settings = deepcopy(GAOKAO_TIME_DEFAULTS)
    arrangement.gaokao_results = {
        "unified": pd.DataFrame(
            [
                {
                    "班级": "1",
                    "学号": "01",
                    "姓名": "张三",
                    "考号": "240001",
                    "选科": "物化生",
                    "考场号": "001",
                    "考场": "第一考场",
                    "座位号": "01",
                },
                {
                    "班级": "1",
                    "学号": "02",
                    "姓名": "李四",
                    "考号": "240002",
                    "选科": "史地政",
                    "考场号": "001",
                    "考场": "第一考场",
                    "座位号": "02",
                },
            ]
        ),
        "electives": {
            "化学": pd.DataFrame(
                [
                    {"班级": "1", "学号": "01", "姓名": "张三", "考号": "240001", "考场号": "005", "考场": "第五考场", "座位号": "12", "科目类型": "化学"},
                    {"班级": "1", "学号": "02", "姓名": "李四", "考号": "240002", "考场号": "015", "考场": "第十五考场", "座位号": "08", "科目类型": "自习"},
                ]
            ),
            "地理": pd.DataFrame(
                [
                    {"班级": "1", "学号": "01", "姓名": "张三", "考号": "240001", "考场号": "025", "考场": "第二十五考场", "座位号": "18", "科目类型": "自习"},
                    {"班级": "1", "学号": "02", "姓名": "李四", "考号": "240002", "考场号": "006", "考场": "第六考场", "座位号": "06", "科目类型": "地理"},
                ]
            ),
            "政治": pd.DataFrame(
                [
                    {"班级": "1", "学号": "01", "姓名": "张三", "考号": "240001", "考场号": "035", "考场": "第三十五考场", "座位号": "28", "科目类型": "自习"},
                    {"班级": "1", "学号": "02", "姓名": "李四", "考号": "240002", "考场号": "007", "考场": "第七考场", "座位号": "09", "科目类型": "政治"},
                ]
            ),
            "生物": pd.DataFrame(
                [
                    {"班级": "1", "学号": "01", "姓名": "张三", "考号": "240001", "考场号": "008", "考场": "第八考场", "座位号": "03", "科目类型": "生物"},
                    {"班级": "1", "学号": "02", "姓名": "李四", "考号": "240002", "考场号": "045", "考场": "第四十五考场", "座位号": "11", "科目类型": "自习"},
                ]
            ),
        },
    }
    arrangement.arranged_students = arrangement._merge_gaokao_results()
    return arrangement


def test_merge_gaokao_results_creates_student_centered_view() -> None:
    arrangement = _build_gaokao_arrangement()

    merged = arrangement._merge_gaokao_results()

    assert merged["物理历史科目"].tolist() == ["物理", "历史"]
    assert merged["语文考场号"].tolist() == ["001", "001"]
    assert merged["化学科目"].tolist() == ["化学", "自习"]
    assert merged["地理科目"].tolist() == ["自习", "地理"]
    assert merged["政治座位号"].tolist() == ["28", "09"]
    assert merged["生物考场"].tolist() == ["第八考场", "第四十五考场"]


def test_save_gaokao_results_exports_expected_workbook(tmp_path) -> None:
    arrangement = _build_gaokao_arrangement()
    output_path = tmp_path / "gaokao-results.xlsx"

    ok, message = arrangement.save_gaokao_results(str(output_path))

    assert ok is True
    assert "高考编排结果已保存至" in message
    assert output_path.exists()

    workbook = load_workbook(output_path, read_only=True)
    assert workbook.sheetnames == [
        "考场安排（学生）",
        "考场安排（座位）",
        "统考编排结果",
        "化学编排结果",
        "地理编排结果",
        "政治编排结果",
        "生物编排结果",
        "考场人数统计",
    ]

    student_sheet = workbook["考场安排（学生）"]
    stats_sheet = workbook["考场人数统计"]
    assert student_sheet.max_row >= 3
    assert stats_sheet.max_row >= 2
