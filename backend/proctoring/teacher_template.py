from __future__ import annotations

import pandas as pd


def write_teacher_template_xlsx(file_path: str) -> None:
    template_data = [
        {
            "姓名": "张三",
            "性别": "男",
            "是否本校": "是",
            "最大监考段数": 3,
            "不监考科目": "1,3",
            "历次监考时长": 0,
            "预设监考考场": "1",
        },
        {
            "姓名": "李四",
            "性别": "女",
            "是否本校": "是",
            "最大监考段数": 2,
            "不监考科目": "2",
            "历次监考时长": 0,
            "预设监考考场": "2",
        },
        {
            "姓名": "王五",
            "性别": "男",
            "是否本校": "否",
            "最大监考段数": 4,
            "不监考科目": "",
            "历次监考时长": 0,
            "预设监考考场": "",
        },
        {
            "姓名": "赵六",
            "性别": "女",
            "是否本校": "否",
            "最大监考段数": 3,
            "不监考科目": "1,2,4",
            "历次监考时长": 0,
            "预设监考考场": "",
        },
    ]

    df = pd.DataFrame(template_data)

    with pd.ExcelWriter(file_path, engine="openpyxl") as writer:
        sheet_name = "Sheet1"
        df.to_excel(writer, sheet_name=sheet_name, index=False)
        workbook = writer.book
        worksheet = workbook[sheet_name]

        for column in worksheet.columns:
            header_cell = column[0]
            header_text = str(header_cell.value) if header_cell.value is not None else ""
            adjusted_width = max(len(header_text), 0) * 3
            worksheet.column_dimensions[header_cell.column_letter].width = min(adjusted_width, 50)

        from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
        from openpyxl.utils import get_column_letter

        instructions = {
            "姓名": "必填。\n示例：张三、李四。",
            "性别": "选填。\n若启用“双教师监考”且设置“男女搭配”约束，则必填。\n填写要求：男/女。",
            "是否本校": "选填。\n若启用“双教师监考”且设置“本校外搭配”约束，则必填。\n填写要求：是/否。",
            "最大监考段数": "选填。\n留空表示不设置监考段数限制。\n填写要求：为非负整数，且不超过科目数。",
            "不监考科目": "选填。\n支持科目编号或科目名称；多个项可用英文/中文逗号、分号、中文分号、顿号或空格分隔，例如：\n1,3\n语文、数学\n科目1 科目2\n科目名称需与已导入科目信息一致。",
            "历次监考时长": "选填。\n单位：分钟；可留空（默认0）。\n支持填写负数，用于历史监考时长补偿。",
            "预设监考考场": "选填。\n数字范围：1..考场数；可留空；越界或非法将被忽略。",
        }

        headers = list(df.columns)
        required_cols = {"姓名"}

        desc_ws = workbook.create_sheet("填写说明")

        thin = Side(style="thin")
        header_border = Border(left=thin, right=thin, top=thin, bottom=thin)
        header_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
        required_header_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
        normal_header_font = Font(bold=True)

        wrap_left = Alignment(horizontal="left", vertical="top", wrap_text=True)
        required_cell_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")

        for idx, col in enumerate(headers, start=1):
            cell = desc_ws.cell(row=1, column=idx, value=col)
            cell.font = normal_header_font
            cell.alignment = header_align
            cell.border = header_border
            if col in required_cols:
                cell.fill = required_header_fill

        for idx, col in enumerate(headers, start=1):
            val = instructions.get(col, "")
            cell = desc_ws.cell(row=2, column=idx, value=val)
            cell.alignment = wrap_left
            if col in required_cols:
                cell.fill = required_cell_fill

        max_lines = max((instructions.get(col, "").count("\n") + 1) for col in headers)
        line_height = 25
        padding = 8
        desc_ws.row_dimensions[2].height = max_lines * line_height + padding

        for idx in range(1, len(headers) + 1):
            letter = get_column_letter(idx)
            desc_ws.column_dimensions[letter].width = 20

