import pandas as pd


def create_stats_sheet_with_formulas(arrangement, writer, export_df=None):
    """使用公式创建考场选科统计工作表"""
    from openpyxl.utils import get_column_letter
    from openpyxl.styles import Alignment

    if export_df is not None:
        columns = export_df.columns.tolist()
        try:
            room_col_idx = columns.index("考场") + 1
            first_choice_col_idx = columns.index("首选") + 1
            sub1_col_idx = columns.index("选科1") + 1
            sub2_col_idx = columns.index("选科2") + 1

            room_col_letter = get_column_letter(room_col_idx)
            first_choice_col_letter = get_column_letter(first_choice_col_idx)
            sub1_col_letter = get_column_letter(sub1_col_idx)
            sub2_col_letter = get_column_letter(sub2_col_idx)
        except ValueError:
            room_col_letter = "F"
            first_choice_col_letter = "I"
            sub1_col_letter = "J"
            sub2_col_letter = "K"
    else:
        room_col_letter = "H"
        first_choice_col_letter = "K"
        sub1_col_letter = "L"
        sub2_col_letter = "M"

    if hasattr(arrangement, "room_setting_df") and arrangement.room_setting_df is not None:
        all_room_names_in_data = set(arrangement.arranged_students["考场"].unique())
        room_names = [room for room in arrangement.room_setting_df["考场"].tolist() if room in all_room_names_in_data]
    else:
        room_names = sorted(arrangement.arranged_students["考场"].unique())
    abbr_map = {"化": "化学", "生": "生物", "政": "政治", "地": "地理"}
    allowed = ["化学", "生物", "政治", "地理"]
    priority_order = getattr(arrangement, "subject_priority_order", None)
    if not isinstance(priority_order, list):
        priority_order = list(allowed)
    cleaned = [str(v or "").strip() for v in priority_order]
    filtered = [v for v in cleaned if v in allowed]
    dedup = []
    for v in filtered:
        if v not in dedup:
            dedup.append(v)
    for v in allowed:
        if v not in dedup:
            dedup.append(v)
    priority_order = dedup[: len(allowed)]

    subjects_to_count = ["物理", "历史"] + priority_order

    # 增加 "考试顺序建议" 列
    headers = ["考场"] + subjects_to_count + ["考试顺序建议"]

    stats_data = []
    stats_data.append(headers)

    for room_name in room_names:
        # 增加一个占位符
        row_data = [room_name] + [None] * len(subjects_to_count) + [None]
        stats_data.append(row_data)

    stats_df = pd.DataFrame(stats_data[1:], columns=headers)
    stats_df.to_excel(writer, sheet_name="考场选科统计", index=False)

    workbook = writer.book
    stats_sheet = workbook["考场选科统计"]

    # 设置列宽
    suggestion_col_letter = get_column_letter(len(headers))
    stats_sheet.column_dimensions[suggestion_col_letter].width = 50  # 设置建议列宽度

    try:
        room_combo_counts = arrangement.arranged_students.groupby("考场")[arrangement.subject_column].nunique().to_dict()
    except Exception:
        room_combo_counts = {room: 1 for room in room_names}

    def format_seat_ranges(seat_numbers):
        nums = sorted(set(seat_numbers))
        if not nums:
            return ""
        ranges = []
        start = prev = nums[0]
        for n in nums[1:]:
            if n == prev + 1:
                prev = n
            else:
                ranges.append(str(start) if start == prev else f"{start}-{prev}")
                start = prev = n
        ranges.append(str(start) if start == prev else f"{start}-{prev}")
        return "、".join(ranges)

    def calc_seat_string(room_name, subject):
        try:
            df_room = arrangement.arranged_students[arrangement.arranged_students["考场"] == room_name]
            if subject in ["物理", "历史"]:
                seats_series = df_room.loc[df_room["首选"] == subject, "座位号"]
            else:
                seats_series = df_room.loc[(df_room["选科1"] == subject) | (df_room["选科2"] == subject), "座位号"]
            seat_numbers = []
            for s in seats_series.tolist():
                try:
                    n = int(str(s))
                except Exception:
                    try:
                        n = int(str(s).lstrip("0") or "0")
                    except Exception:
                        continue
                seat_numbers.append(n)
            return format_seat_ranges(seat_numbers)
        except Exception:
            return ""
            
    def get_exam_order_suggestion(room_name):
        """根据优先级计算考试顺序建议"""
        try:
            df_room = arrangement.arranged_students[arrangement.arranged_students["考场"] == room_name]
            if df_room.empty:
                return ""
            priority_map = {subj: i for i, subj in enumerate(priority_order)}
            
            seg1_data = {s: [] for s in priority_order}
            seg2_data = {s: [] for s in priority_order}
            
            for _, row in df_room.iterrows():
                # 从"选科"列（如"物化生"）提取再选科目
                full_sub_str = str(row[arrangement.subject_column]).strip()
                
                subjects = []
                # 处理缩写 "物化生" 或 "理化生"
                # 跳过第一个字符（首选），取后面所有字符
                if len(full_sub_str) >= 2:
                    for c in full_sub_str[1:]:
                        if c in abbr_map:
                            subjects.append(abbr_map[c])
                
                if len(subjects) < 2:
                    continue
                    
                # 只取前两个再选科目（防止异常数据）
                subjects = subjects[:2]
                
                # 按优先级排序
                subjects.sort(key=lambda x: priority_map.get(x, 99))
                
                sub1 = subjects[0] # 高优先级 -> 第一段
                sub2 = subjects[1] # 低优先级 -> 第二段
                
                # 获取座位号
                seat_num = row["座位号"]
                try:
                    seat_num_int = int(str(seat_num).lstrip("0") or "0")
                except:
                    continue
                
                seg1_data[sub1].append(seat_num_int)
                seg2_data[sub2].append(seat_num_int)
            
            def format_segment(data_dict):
                # 将科目和对应的座位号列表收集起来
                subject_seat_pairs = []
                for subj in priority_order:
                    seats = data_dict.get(subj, [])
                    if seats:
                        seats.sort()
                        # 记录 (最小座位号, 科目名, 格式化的座位范围字符串)
                        # 使用最小座位号作为排序依据
                        min_seat = seats[0] if seats else 9999
                        range_str = format_seat_ranges(seats)
                        subject_seat_pairs.append((min_seat, subj, range_str))
                
                # 按最小座位号升序排序
                subject_seat_pairs.sort(key=lambda x: x[0])
                
                # 生成最终字符串
                parts = [f"{subj}（{range_str}）" for _, subj, range_str in subject_seat_pairs]
                return " ".join(parts)

            str1 = format_segment(seg1_data)
            str2 = format_segment(seg2_data)
            
            result = []
            if str1:
                result.append(f"第一段：{str1}")
            if str2:
                result.append(f"第二段：{str2}")
                
            return "\n".join(result) # 使用换行符分隔段

        except Exception as e:
            return f"Error: {str(e)}"

    for row_idx, room_name in enumerate(room_names, start=2):
        is_mixed_room = room_combo_counts.get(room_name, 1) > 1
        
        # 写入统计公式
        for col_idx, subject in enumerate(subjects_to_count, start=2):
            col_letter = get_column_letter(col_idx)

            if subject in ["物理", "历史"]:
                base_formula = f'=COUNTIFS(学生编排结果!${room_col_letter}:${room_col_letter},$A{row_idx},学生编排结果!${first_choice_col_letter}:${first_choice_col_letter},{col_letter}$1)'
            else:
                base_formula = (
                    f'=COUNTIFS(学生编排结果!${room_col_letter}:${room_col_letter},$A{row_idx},学生编排结果!${sub1_col_letter}:${sub1_col_letter},{col_letter}$1)'
                    f'+COUNTIFS(学生编排结果!${room_col_letter}:${room_col_letter},$A{row_idx},学生编排结果!${sub2_col_letter}:${sub2_col_letter},{col_letter}$1)'
                )

            if is_mixed_room:
                seat_str = calc_seat_string(room_name, subject)
                if seat_str:
                    formula = base_formula + f'&" ("&"{seat_str}"&")"'
                else:
                    formula = base_formula
            else:
                formula = base_formula

            stats_sheet[f"{col_letter}{row_idx}"] = formula
        
        # 写入考试顺序建议
        suggestion = get_exam_order_suggestion(room_name)
        sugg_col_idx = len(headers)
        sugg_col_letter = get_column_letter(sugg_col_idx)
        cell = stats_sheet[f"{sugg_col_letter}{row_idx}"]
        cell.value = suggestion
        # 启用自动换行和顶对齐
        cell.alignment = Alignment(wrap_text=True, vertical='top')

    total_row_idx = len(room_names) + 2
    stats_sheet[f"A{total_row_idx}"] = "总计"
    for col_idx, subject in enumerate(subjects_to_count, start=2):
        col_letter = get_column_letter(col_idx)
        if subject in ["物理", "历史"]:
            total_formula = f"=SUM(COUNTIF(学生编排结果!${first_choice_col_letter}:${first_choice_col_letter},{col_letter}$1))"
        else:
            total_formula = (
                f"=SUM(COUNTIF(学生编排结果!${sub1_col_letter}:${sub1_col_letter},{col_letter}$1),"
                f"COUNTIF(学生编排结果!${sub2_col_letter}:${sub2_col_letter},{col_letter}$1))"
            )
        stats_sheet[f"{col_letter}{total_row_idx}"] = total_formula

    return True
