import pandas as pd


def create_stats_sheet_with_formulas(arrangement, writer, export_df=None):
    """使用公式创建考场选科统计工作表"""
    from openpyxl.utils import get_column_letter

    # 动态获取列号
    if export_df is not None:
        # 获取列名列表
        columns = export_df.columns.tolist()
        try:
            # 获取各关键列的索引（从0开始，转换为Excel列号时需要+1）
            room_col_idx = columns.index("考场") + 1
            first_choice_col_idx = columns.index("首选") + 1
            sub1_col_idx = columns.index("选科1") + 1
            sub2_col_idx = columns.index("选科2") + 1

            # 转换为字母
            room_col_letter = get_column_letter(room_col_idx)
            first_choice_col_letter = get_column_letter(first_choice_col_idx)
            sub1_col_letter = get_column_letter(sub1_col_idx)
            sub2_col_letter = get_column_letter(sub2_col_idx)
        except ValueError:
            # 如果找不到列，回退到默认值（兼容旧逻辑）
            room_col_letter = "F"  # 修改后的默认位置
            first_choice_col_letter = "I"
            sub1_col_letter = "J"
            sub2_col_letter = "K"
    else:
        # 默认值
        room_col_letter = "H"
        first_choice_col_letter = "K"
        sub1_col_letter = "L"
        sub2_col_letter = "M"

    # 获取所有考场名称，按照导入考场设置的原始顺序
    if hasattr(arrangement, "room_setting_df") and arrangement.room_setting_df is not None:
        # 如果有考场设置DataFrame，按照原始顺序获取考场名称
        # 只包含实际在编排结果中存在的考场
        all_room_names_in_data = set(arrangement.arranged_students["考场"].unique())
        room_names = [room for room in arrangement.room_setting_df["考场"].tolist() if room in all_room_names_in_data]
    else:
        # 如果没有考场设置，按照考场名称排序
        room_names = sorted(arrangement.arranged_students["考场"].unique())
    subjects_to_count = ["物理", "历史", "化学", "生物", "地理", "政治"]

    # 创建表头数据
    headers = ["考场"] + subjects_to_count

    # 创建工作表数据结构
    stats_data = []

    # 添加表头
    stats_data.append(headers)

    # 为每个考场添加一行，但只填入考场名称，其他列将使用公式
    for room_name in room_names:
        row_data = [room_name] + [None] * len(subjects_to_count)
        stats_data.append(row_data)

    # 创建DataFrame并保存到Excel
    stats_df = pd.DataFrame(stats_data[1:], columns=headers)
    stats_df.to_excel(writer, sheet_name="考场选科统计", index=False)

    # 获取工作簿和工作表对象，添加公式
    workbook = writer.book
    stats_sheet = workbook["考场选科统计"]

    # 预计算各考场是否为混合考场：以房间内“选科”组合的种类数判定
    # 若同一考场存在多种“选科”字符串，则视为混合考场
    try:
        room_combo_counts = arrangement.arranged_students.groupby("考场")[arrangement.subject_column].nunique().to_dict()
    except Exception:
        room_combo_counts = {room: 1 for room in room_names}

    # 工具函数：将座位号列表压缩为连续区间字符串，如 "1-14、33-34"
    def format_seat_ranges(seat_numbers):
        nums = sorted(set(seat_numbers))
        if not nums:
            return ""
        ranges = []
        start = prev = nums[0]
        for n in nums[1:]:
            if n == prev + 1:
                prev = n
            else:
                ranges.append(str(start) if start == prev else f"{start}-{prev}")
                start = prev = n
        ranges.append(str(start) if start == prev else f"{start}-{prev}")
        return "、".join(ranges)

    # 工具函数：计算指定考场、指定科目的座位号字符串
    def calc_seat_string(room_name, subject):
        try:
            df_room = arrangement.arranged_students[arrangement.arranged_students["考场"] == room_name]
            if subject in ["物理", "历史"]:
                seats_series = df_room.loc[df_room["首选"] == subject, "座位号"]
            else:
                seats_series = df_room.loc[(df_room["选科1"] == subject) | (df_room["选科2"] == subject), "座位号"]
            seat_numbers = []
            for s in seats_series.tolist():
                try:
                    # 兼容 "01" 等文本座位号
                    n = int(str(s))
                except Exception:
                    try:
                        n = int(str(s).lstrip("0") or "0")
                    except Exception:
                        continue
                seat_numbers.append(n)
            return format_seat_ranges(seat_numbers)
        except Exception:
            return ""

    # 为每个考场的每个科目添加COUNTIFS公式，并在混合考场拼接座位号
    for row_idx, room_name in enumerate(room_names, start=2):  # 从第2行开始（第1行是表头）
        is_mixed_room = room_combo_counts.get(room_name, 1) > 1
        for col_idx, subject in enumerate(subjects_to_count, start=2):  # 从第2列开始（第1列是考场）
            col_letter = get_column_letter(col_idx)

            # 根据科目类型构建不同的COUNTIFS公式
            if subject in ["物理", "历史"]:
                # 物理/历史只统计首选科目
                base_formula = f'=COUNTIFS(学生编排结果!${room_col_letter}:${room_col_letter},$A{row_idx},学生编排结果!${first_choice_col_letter}:${first_choice_col_letter},{col_letter}$1)'
            else:
                # 其他科目统计选科1和选科2
                base_formula = (
                    f'=COUNTIFS(学生编排结果!${room_col_letter}:${room_col_letter},$A{row_idx},学生编排结果!${sub1_col_letter}:${sub1_col_letter},{col_letter}$1)'
                    f'+COUNTIFS(学生编排结果!${room_col_letter}:${room_col_letter},$A{row_idx},学生编排结果!${sub2_col_letter}:${sub2_col_letter},{col_letter}$1)'
                )

            # 如果是混合考场，则计算并拼接座位号字符串
            if is_mixed_room:
                seat_str = calc_seat_string(room_name, subject)
                if seat_str:
                    # 使用Excel的&拼接：人数公式 & " (" & 座位号串 & ")"
                    formula = base_formula + f'&" ("&"{seat_str}"&")"'
                else:
                    formula = base_formula
            else:
                formula = base_formula

            # 设置公式到对应单元格
            stats_sheet[f"{col_letter}{row_idx}"] = formula

    # 在最后一行添加“总计”并对每一科目进行求和（使用SUM公式）
    total_row_idx = len(room_names) + 2  # 数据起始行2 + 房间数
    stats_sheet[f"A{total_row_idx}"] = "总计"
    for col_idx, subject in enumerate(subjects_to_count, start=2):
        col_letter = get_column_letter(col_idx)
        if subject in ["物理", "历史"]:
            # 总计=所有房间的首选为该科的人数总和
            total_formula = f"=SUM(COUNTIF(学生编排结果!${first_choice_col_letter}:${first_choice_col_letter},{col_letter}$1))"
        else:
            # 总计=选科1为该科 + 选科2为该科 的人数总和
            total_formula = (
                f"=SUM(COUNTIF(学生编排结果!${sub1_col_letter}:${sub1_col_letter},{col_letter}$1),"
                f"COUNTIF(学生编排结果!${sub2_col_letter}:${sub2_col_letter},{col_letter}$1))"
            )
        stats_sheet[f"{col_letter}{total_row_idx}"] = total_formula

    return True

