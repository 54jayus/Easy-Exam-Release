#!/usr/bin/env python3
# -*- coding: utf-8 -*-

"""
教师监考安排功能页面
"""

import os
from models import Exam
from PyQt5.QtWidgets import (QWidget, QVBoxLayout, QHBoxLayout, 
                             QPushButton, QLabel, QFileDialog, QMessageBox,
                             QGroupBox, QRadioButton, QCheckBox, QSpinBox,
                             QTabWidget, QTableWidget, QTableWidgetItem,
                             QComboBox, QTextEdit, QLineEdit,QButtonGroup,
                             QApplication, QProgressBar)
from PyQt5.QtCore import Qt
from PyQt5.QtWidgets import QDialog
# 设置字体和字号
from PyQt5.QtGui import QFont
import pandas as pd

from data_import import DataImporter
from models import Schedule

class ProctorPage(QWidget):
    """
    智能考务系统主窗口
    """
    
    def __init__(self,parent=None):
        super().__init__(parent)
        
        # 数据存储
        self.teachers = []
        self.schedule = None
        self.selected_cells = []
        self.subject_count = 1
        self.adjust_mode = False
        self.subject_names = []  # 添加科目名称列表
        self.exam_times = []  # 添加考试时间列表

        # 二次均衡优化行为参数（默认：完成后自动执行；按钮默认不可见）
        self.auto_optimize_postprocess = True
        self.show_optimize_button_flag = False
        self.log_optimization_swaps = False

        # 初始化UI
        self.init_ui()
        
    def init_ui(self):
        """
        初始化用户界面
        """
        
        # 创建主布局
        main_layout = QVBoxLayout(self)
        main_layout.setContentsMargins(10, 10, 10, 0)  # 设置底部边距为0
        # 创建顶部工具栏
        self.create_toolbar(main_layout)
        
        # 创建参数设置区域
        self.create_parameter_settings(main_layout)
        
        # 创建约束条件设置区域
        self.create_constraints_settings(main_layout)
        
        # 创建操作按钮区域
        self.create_action_buttons(main_layout)
        
        # 创建结果显示区域
        self.create_result_tabs(main_layout)
        
    def create_toolbar(self, parent_layout):
        """
        创建工具栏
        """
        toolbar_layout = QHBoxLayout()
        
        # 添加生成教师信息模板按钮
        self.generate_template_btn = QPushButton('生成模板文件')
        self.generate_template_btn.clicked.connect(self.generate_teacher_template)
        toolbar_layout.addWidget(self.generate_template_btn)
        
        # 导入按钮
        self.import_btn = QPushButton('导入教师信息')
        self.import_btn.clicked.connect(self.import_teachers)
        toolbar_layout.addWidget(self.import_btn)
        
        # 预设监考安排按钮
        self.preset_btn = QPushButton('预设监考安排')
        self.preset_btn.setToolTip('生成空监考表或导入预设安排')
        self.preset_btn.setEnabled(False)
        self.preset_btn.clicked.connect(self.open_preset_dialog)
        toolbar_layout.addWidget(self.preset_btn)
        
        # 导入监考安排按钮
        self.import_schedule_btn = QPushButton('导入监考安排')
        self.import_schedule_btn.clicked.connect(self.import_schedule)
        self.import_schedule_btn.setEnabled(False)
        toolbar_layout.addWidget(self.import_schedule_btn)
        
        # 导出按钮
        self.export_btn = QPushButton('导出安排结果')
        self.export_btn.clicked.connect(self.export_schedule)
        self.export_btn.setEnabled(False)
        toolbar_layout.addWidget(self.export_btn)
        
        toolbar_layout.addStretch()
        parent_layout.addLayout(toolbar_layout)
        
    def create_parameter_settings(self, parent_layout):
        """
        创建参数设置区域
        """
        param_group = QGroupBox('参数设置')
        param_layout = QHBoxLayout()

        # 科目数设置（只读）
        subject_layout = QHBoxLayout()
        subject_label = QLabel('考试科目数:')
        subject_label.setFixedWidth(90)
        subject_label.setToolTip("科目数由科目设置页面确定，无法在此处修改")
        subject_layout.addWidget(subject_label)
        self.subject_spin = QLineEdit('1')  # 使用QLineEdit替代QLabel
        self.subject_spin.setStyleSheet("color:grey;font-size:16px;")  # 设置字体颜色为灰色
        self.subject_spin.setReadOnly(True)  # 设置为只读
        self.subject_spin.setMaximumWidth(40)  # 设置最大宽度为40像素
        self.subject_spin.setToolTip("科目数由科目设置页面确定，无法在此处修改")
        subject_layout.addWidget(self.subject_spin)
        
        param_layout.addLayout(subject_layout)
        
        # 考场数设置
        room_layout = QHBoxLayout()
        room_layout.addWidget(QLabel('考试场室数:'))
        self.room_spin = QSpinBox()
        # 设置为可为空的状态：0 表示未设置，显示为空文本
        self.room_spin.setRange(0, 100)
        self.room_spin.setSpecialValueText("")
        self.room_spin.setValue(0)
        room_layout.addWidget(self.room_spin)
        param_layout.addLayout(room_layout)
        
        # 监考模式设置
        mode_layout = QHBoxLayout()
        mode_layout.addWidget(QLabel('监考模式:'))
        self.single_mode_radio = QRadioButton('单教师监考')
        self.single_mode_radio.setChecked(True)
        self.double_mode_radio = QRadioButton('双教师监考')
        mode_layout.addWidget(self.single_mode_radio)
        mode_layout.addWidget(self.double_mode_radio)
        param_layout.addLayout(mode_layout)

        # 为监考模式创建按钮组
        self.mode_group = QButtonGroup()
        self.mode_group.addButton(self.single_mode_radio)
        self.mode_group.addButton(self.double_mode_radio)

        # 均衡模式设置
        balance_layout = QHBoxLayout()
        balance_layout.addWidget(QLabel('均衡模式:'))
        self.session_balance_radio = QRadioButton('场次均衡')
        self.session_balance_radio.setChecked(True)
        self.duration_balance_radio = QRadioButton('时长均衡')
        balance_layout.addWidget(self.session_balance_radio)
        balance_layout.addWidget(self.duration_balance_radio)
        param_layout.addLayout(balance_layout)

        # 为均衡模式创建按钮组
        self.balance_group = QButtonGroup()
        self.balance_group.addButton(self.session_balance_radio)
        self.balance_group.addButton(self.duration_balance_radio)

        param_layout.addStretch()
        param_group.setLayout(param_layout)
        parent_layout.addWidget(param_group)
        
    def create_constraints_settings(self, parent_layout):
        """
        创建约束条件设置区域
        """
        constraint_group = QGroupBox('约束条件设置')
        constraint_layout = QHBoxLayout()
        
        # 男女搭配选项
        self.gender_mix_checkbox = QCheckBox('男女搭配')
        self.gender_mix_checkbox.setEnabled(False)  # 默认禁用，选择双教师模式时启用
        constraint_layout.addWidget(self.gender_mix_checkbox)
        
        # 本外校搭配选项
        self.internal_mix_checkbox = QCheckBox('本外校搭配')
        self.internal_mix_checkbox.setEnabled(False)  # 默认禁用，选择双教师模式时启用
        constraint_layout.addWidget(self.internal_mix_checkbox)

        # 锁定导入安排选项
        self.lock_imported_checkbox = QCheckBox('锁定导入安排')
        self.lock_imported_checkbox.setChecked(True)
        self.lock_imported_checkbox.setVisible(False)
        # self.lock_imported_checkbox.setToolTip('勾选后，导入的监考安排在二次均衡优化中不会被调整；导入项将以红色显示。')
        constraint_layout.addWidget(self.lock_imported_checkbox)
        # 监听锁定导入安排变化
        self.lock_imported_checkbox.toggled.connect(self.on_lock_imported_changed)
        
        # 监听监考模式变化
        self.single_mode_radio.toggled.connect(self.on_mode_changed) #当单选按钮（Radio Button）的选中状态发生变化时会触发此信号
        self.double_mode_radio.toggled.connect(self.on_mode_changed)
        
        constraint_layout.addStretch() #右边增加弹性空间，使得空间靠左对齐
        constraint_group.setLayout(constraint_layout)
        parent_layout.addWidget(constraint_group)
        
    def create_action_buttons(self, parent_layout):
        """
        创建操作按钮区域
        """
        button_layout = QHBoxLayout()
        
        # 安排监考按钮
        self.schedule_btn = QPushButton('安排监考')
        self.schedule_btn.clicked.connect(self.generate_schedule)
        self.schedule_btn.setEnabled(False)
        button_layout.addWidget(self.schedule_btn)
        
        # 手动调整按钮
        self.adjust_btn = QPushButton('手动调整')
        self.adjust_btn.clicked.connect(self.manual_adjust)
        self.adjust_btn.setEnabled(False)
        button_layout.addWidget(self.adjust_btn)
        
        # 添加补全监考安排按钮
        self.continue_schedule_btn = QPushButton('补全监考安排')
        self.continue_schedule_btn.clicked.connect(self.continue_schedule)
        self.continue_schedule_btn.setEnabled(False)
        button_layout.addWidget(self.continue_schedule_btn)

        # 二次均衡优化按钮
        self.optimize_btn = QPushButton('二次均衡优化')
        # 改用包装槽方法以支持静默/非静默参数
        self.optimize_btn.clicked.connect(self.on_optimize_button_clicked)
        self.optimize_btn.setEnabled(False)
        # 通过代码参数控制初始可见性
        self.optimize_btn.setVisible(self.show_optimize_button_flag)
        button_layout.addWidget(self.optimize_btn)
        
        button_layout.addStretch() #右边增加弹性空间，使得空间靠左对齐
        parent_layout.addLayout(button_layout)
        
    def create_result_tabs(self, parent_layout):
        """
        创建结果显示区域
        """
        # 创建水平布局来放置标签页和日志区域
        h_layout = QHBoxLayout()
        
        # 左侧采用垂直布局：上方是标签页，下方是局部进度区
        left_layout = QVBoxLayout()
        
        # 创建标签页
        self.result_tabs = QTabWidget()
        
        # 监考总览表
        self.overview_table = QTableWidget(0, 0)
        self.overview_table.setHorizontalHeaderLabels(['科目', '考场', '监考教师'])
        self.overview_table.cellClicked.connect(self.on_overview_cell_clicked)
        self.result_tabs.addTab(self.overview_table, '监考总览表')
        
        # 监考统计表
        self.statistics_table = QTableWidget(0, 2)
        self.statistics_table.setHorizontalHeaderLabels(['教师姓名', '监考次数'])
        self.result_tabs.addTab(self.statistics_table, '监考统计')
        
        # 分科目监考表
        subject_widget = QWidget()
        subject_layout = QVBoxLayout(subject_widget)
        
        self.subject_combo = QComboBox()
        self.subject_combo.currentIndexChanged.connect(self.update_subject_table)
        subject_layout.addWidget(self.subject_combo)
        
        self.subject_table = QTableWidget(0, 3)
        self.subject_table.setHorizontalHeaderLabels(['科目', '考场', '监考教师'])
        subject_layout.addWidget(self.subject_table)
        
        self.result_tabs.addTab(subject_widget, '分科目监考表')
        
        # 在标签页下方创建局部进度显示区域（默认隐藏）
        local_progress_container = QHBoxLayout()
        self.local_progress_label = QLabel("")
        self.local_progress_label.setStyleSheet("color: #666666; font-size: 12px;")
        self.local_progress_label.setVisible(False)
        self.local_progress = QProgressBar()
        self.local_progress.setTextVisible(False)
        self.local_progress.setMaximumHeight(12)
        self.local_progress.setVisible(False)
        local_progress_container.addWidget(self.local_progress_label)
        local_progress_container.addWidget(self.local_progress)
        
        left_layout.addWidget(self.result_tabs)
        left_layout.addLayout(local_progress_container)
        
        # 将左侧布局添加到水平布局
        h_layout.addLayout(left_layout)
        
        # 创建右侧日志区域
        self.log_text = QTextEdit()
        self.log_text.setFont(QFont("Microsoft YaHei", 10))
        self.log_text.setReadOnly(True)
        self.log_text.setFixedWidth(300)
        
        # 创建一个垂直布局来放置日志区域
        log_layout = QVBoxLayout()
    
        # 通过设置上边距来微调
        tab_height = self.result_tabs.tabBar().height()
        log_layout.setContentsMargins(0, tab_height - 5, 0, 3)
        log_layout.addWidget(self.log_text)
        
        # 将日志布局添加到水平布局
        h_layout.addLayout(log_layout)
        
        # 将整个水平布局添加到父布局
        parent_layout.addLayout(h_layout)

    def start_local_busy(self, message: str = "正在处理…"):
        """显示页面内的进度条为不确定模式，并展示说明文字。"""
        try:
            self.local_progress_label.setText(message)
            self.local_progress_label.setVisible(True)
            self.local_progress.setRange(0, 0)
            self.local_progress.setVisible(True)
        except Exception:
            pass

    def update_local_busy(self, message: str = None, percent: int = None):
        """更新页面内进度条的说明或百分比。"""
        try:
            if message is not None:
                self.local_progress_label.setText(message)
                self.local_progress_label.setVisible(True)
            if percent is not None:
                self.local_progress.setRange(0, 100)
                self.local_progress.setValue(max(0, min(100, percent)))
                self.local_progress.setVisible(True)
        except Exception:
            pass

    def stop_local_busy(self):
        """隐藏页面内进度条与说明文字，并重置状态。"""
        try:
            self.local_progress.setVisible(False)
            self.local_progress.setRange(0, 100)
            self.local_progress.setValue(0)
            self.local_progress_label.setVisible(False)
        except Exception:
            pass

    def make_progress_reporter(self, prefix: str = None, fixed_message: str = None):
        """创建一个进度回调，将模型层的进度同步到页面内进度条与状态栏。
        使用示例：self.schedule.set_constraint('progress_callback', self.make_progress_reporter('生成监考安排'))
        支持固定消息，当 prefix 以中文顿号或冒号结尾时，直接拼接，不加破折号。
        """
        parent = self.parent()
        def reporter(message: str, percent: int = None):
            display_message = fixed_message if fixed_message is not None else message
            if prefix:
                if prefix.endswith('：') or prefix.endswith(':'):
                    text = f"{prefix}{display_message or ''}"
                else:
                    text = f"{prefix} - {display_message}" if display_message else prefix
            else:
                text = display_message
            try:
                # 页面内进度条
                self.update_local_busy(message=text, percent=percent)
                # 状态栏进度（如果主窗口支持）
                if hasattr(parent, 'update_busy'):
                    parent.update_busy(message=text, percent=percent)
                QApplication.processEvents()
            except Exception:
                pass
        return reporter

    def update_continue_button_state(self):
        """
        根据当前安排是否完整，更新“补全监考安排”按钮状态。
        安排已完成（每个考场每科都有老师）则禁用；否则启用。
        """
        try:
            if hasattr(self, 'continue_schedule_btn'):
                enabled = bool(self.schedule) and (not self.schedule.is_schedule_complete())
                self.continue_schedule_btn.setEnabled(enabled)
        except Exception:
            pass

    def log(self, message):
        """        
        输出日志到日志区
        """
        if self.log_text:
            self.log_text.append(message)

    def on_overview_cell_clicked(self, row, column):
        """
        监考总览表单元格点击事件处理
        """
        # 只在调整模式下处理点击事件
        if not self.adjust_mode:
            return
            
        # 检查是否点击的是监考员单元格（非标题行和考场列）
        if row >= 0 and column >= 0:
            item = self.overview_table.item(row, column)
            if item:
                # 检查单元格是否已有选中状态
                is_selected = False
                selected_index = -1
                for i, (r, c) in enumerate(self.selected_cells):
                    if r == row and c == column:
                        is_selected = True
                        selected_index = i
                        break
                
                # 如果已选中，则取消选中
                if is_selected:
                    self.selected_cells.pop(selected_index)
                    item.setBackground(Qt.white)  # 恢复背景色
                else:
                    # 如果未选中，添加到选中列表
                    # 限制最多只能选中两个单元格
                    if len(self.selected_cells) >= 2:
                        # 取消最早选中的单元格
                        old_row, old_col = self.selected_cells.pop(0)
                        old_item = self.overview_table.item(old_row, old_col)
                        if old_item:
                            old_item.setBackground(Qt.white)
                    
                    self.selected_cells.append((row, column))
                    item.setBackground(Qt.yellow)  # 设置选中背景色
                
                # 如果选中了两个单元格，尝试交换
                if len(self.selected_cells) == 2:
                    self.attempt_swap()
                    
    def manual_adjust(self):
        """
        手动调整监考安排
        """
        # 进入调整模式
        self.adjust_mode = True
        
        # 更新按钮文本和功能
        self.adjust_btn.setText("退出调整")
        self.adjust_btn.clicked.disconnect()
        self.adjust_btn.clicked.connect(self.exit_adjust_mode)
        
        # 显示提示信息
        QMessageBox.information(self, '手动调整', '已进入手动调整模式。\n请在监考总览表中点击要交换的两个监考教师姓名。')
        
    def exit_adjust_mode(self):
        """
        退出调整模式
        """
        self.adjust_mode = False
        self.clear_selection()
        
        # 恢复按钮文本和功能
        self.adjust_btn.setText("手动调整")
        self.adjust_btn.clicked.disconnect()
        self.adjust_btn.clicked.connect(self.manual_adjust)
        
        # 显示提示信息
        QMessageBox.information(self, '手动调整', '已退出手动调整模式。')
        
    def attempt_swap(self):
        """
        尝试交换两个选中的监考安排
        """
        if len(self.selected_cells) != 2 or not self.schedule:
            return
            
        # 获取选中的两个单元格
        row1, col1 = self.selected_cells[0]
        row2, col2 = self.selected_cells[1]
        
        # 获取考场号和科目信息
        room1 = int(self.overview_table.verticalHeaderItem(row1).text()[2:])  # 去掉"考场"前缀
        room2 = int(self.overview_table.verticalHeaderItem(row2).text()[2:])  # 去掉"考场"前缀
        
        # 根据监考模式计算科目ID和教师索引
        if self.double_mode_radio.isChecked():
            # 双教师模式：每科目有两列
            subject1 = col1 // 2 + 1  # 每两列对应一个科目
            subject2 = col2 // 2 + 1  # 每两列对应一个科目
            
            # 确定是监考员1还是监考员2
            is_second_teacher1 = col1 % 2  # 0表示监考员1，1表示监考员2
            is_second_teacher2 = col2 % 2  # 0表示监考员1，1表示监考员2
        else:
            # 单教师模式：每科目一列
            subject1 = col1 + 1
            subject2 = col2 + 1
            
            # 单教师模式下，只有一个教师，索引始终为0
            is_second_teacher1 = 0
            is_second_teacher2 = 0
        
        # 锁定导入安排时，禁止涉及预设位置的交换
        if self.schedule.get_constraint('lock_imported'):
            if (self.schedule.is_position_imported(subject1, room1, is_second_teacher1) or
                self.schedule.is_position_imported(subject2, room2, is_second_teacher2)):
                QMessageBox.warning(self, '交换失败', '选中的位置包含预设安排（已锁定），无法交换。')
                self.clear_selection()
                return
        
        # 尝试交换
        success, message = self.schedule.swap_teachers((subject1, room1, is_second_teacher1), 
                                                      (subject2, room2, is_second_teacher2))
        
        if success:
            # 交换成功，更新界面
            self.display_results(self.schedule.exams)
            QMessageBox.information(self, '交换成功', '监考安排已成功交换')
        else:
            # 交换失败，显示错误信息
            QMessageBox.warning(self, '交换失败', message)
            
        # 清除选中状态
        self.clear_selection()
        
    def clear_selection(self):
        """
        清除所有选中状态
        """
        # 恢复所有选中单元格的背景色
        for row, col in self.selected_cells:
            item = self.overview_table.item(row, col)
            if item:
                item.setBackground(Qt.white)
                
        # 清空选中列表
        self.selected_cells = []
        
    def update_subject_table(self, index):
        """
        根据科目选择更新分科目监考表
        """
        selected_subject = index + 1 if index >= 0 else 1
        
        self.subject_table.clearContents()
        self.subject_table.setRowCount(0)
        
        # 根据监考模式设置列数和标题
        if self.double_mode_radio.isChecked():
            # 双教师模式：监考员1（姓名）、监考员1性别、监考员1来源、监考员2（姓名）、监考员2性别、监考员2来源
            self.subject_table.setColumnCount(6)
            self.subject_table.setHorizontalHeaderLabels([
                '监考员1（姓名）', '监考员1性别', '监考员1来源', 
                '监考员2（姓名）', '监考员2性别', '监考员2来源'
            ])
        else:
            # 单教师模式：监考教师（姓名）、性别、来源
            self.subject_table.setColumnCount(3)
            self.subject_table.setHorizontalHeaderLabels(['监考教师（姓名）', '性别', '来源'])
        
        if not self.schedule or not self.schedule.exams:
            return
        
        # 获取选定科目的考场列表
        target_exam = None
        for exam in self.schedule.exams:
            if exam.subject_id == selected_subject:
                target_exam = exam
                break
        
        if not target_exam:
            return
            
        # 按考场号排序添加行
        sorted_rooms = sorted(target_exam.rooms)
        self.subject_table.setRowCount(len(sorted_rooms))
        
        # 设置行标题为考场序号
        self.subject_table.setVerticalHeaderLabels([str(room) for room in sorted_rooms])
        
        # 填充数据
        for row, room in enumerate(sorted_rooms):
            if room in target_exam.schedule:
                teachers = target_exam.schedule[room]
                
                if self.double_mode_radio.isChecked() and len(teachers) >= 1:
                    # 双教师模式
                    # 监考员1信息（可能为None）
                    if len(teachers) >= 1 and teachers[0] is not None:
                        item = QTableWidgetItem(teachers[0].name)
                        item.setTextAlignment(Qt.AlignCenter)
                        # 根据导入锁定或性别设置颜色（移除导入红色标记，除非显式启用）
                        highlight_imported = bool(self.schedule and self.schedule.get_constraint('highlight_imported', False))
                        if highlight_imported and self.schedule and self.schedule.get_constraint('lock_imported') and self.schedule.is_position_imported(selected_subject, room, 0):
                            item.setForeground(Qt.red)
                        else:
                            # 科目视图不进行预设绿色标记，仅按性别着色
                            if teachers[0].gender == 'M':
                                item.setForeground(Qt.blue)
                            elif teachers[0].gender == 'F':
                                item.setForeground(Qt.magenta)
                        self.subject_table.setItem(row, 0, item)
                        
                        gender_text = ''
                        if teachers[0].gender == 'M':
                            gender_text = '男'
                        elif teachers[0].gender == 'F':
                            gender_text = '女'
                        item = QTableWidgetItem(gender_text)
                        item.setTextAlignment(Qt.AlignCenter)
                        self.subject_table.setItem(row, 1, item)
                        
                        internal_text = ''
                        if teachers[0].is_internal is True:
                            internal_text = '本校'
                        elif teachers[0].is_internal is False:
                            internal_text = '外校'
                        item = QTableWidgetItem(internal_text)
                        item.setTextAlignment(Qt.AlignCenter)
                        self.subject_table.setItem(row, 2, item)
                    # 如果teachers[0]为None，单元格保持空白
                    
                    # 监考员2信息
                    if len(teachers) >= 2 and teachers[1] is not None:
                        item = QTableWidgetItem(teachers[1].name)
                        item.setTextAlignment(Qt.AlignCenter)
                        # 根据导入锁定或性别设置颜色（移除导入红色标记，除非显式启用）
                        highlight_imported = bool(self.schedule and self.schedule.get_constraint('highlight_imported', False))
                        if highlight_imported and self.schedule and self.schedule.get_constraint('lock_imported') and self.schedule.is_position_imported(selected_subject, room, 1):
                            item.setForeground(Qt.red)
                        else:
                            # 科目视图不进行预设绿色标记，仅按性别着色
                            if teachers[1].gender == 'M':
                                item.setForeground(Qt.blue)
                            elif teachers[1].gender == 'F':
                                item.setForeground(Qt.magenta)
                        self.subject_table.setItem(row, 3, item)
                        
                        gender_text = ''
                        if teachers[1].gender == 'M':
                            gender_text = '男'
                        elif teachers[1].gender == 'F':
                            gender_text = '女'
                        item = QTableWidgetItem(gender_text)
                        item.setTextAlignment(Qt.AlignCenter)
                        self.subject_table.setItem(row, 4, item)
                        
                        internal_text = ''
                        if teachers[1].is_internal is True:
                            internal_text = '本校'
                        elif teachers[1].is_internal is False:
                            internal_text = '外校'
                        item = QTableWidgetItem(internal_text)
                        item.setTextAlignment(Qt.AlignCenter)
                        self.subject_table.setItem(row, 5, item)
                    # 处理只有一个教师但为监考员2的情况
                    elif len(teachers) == 1 and teachers[0] is None:
                        # 这种情况表示监考员1为空，但列表中有一个None占位符
                        pass
                elif not self.double_mode_radio.isChecked():
                    # 单教师模式
                    # 修改：过滤掉None值后再访问name属性
                    teacher_names = ', '.join([t.name for t in teachers if t is not None])
                    item = QTableWidgetItem(teacher_names)
                    item.setTextAlignment(Qt.AlignCenter)
                    if self.schedule and self.schedule.get_constraint('lock_imported') and self.schedule.is_position_imported(selected_subject, room, 0):
                        item.setForeground(Qt.red)
                    self.subject_table.setItem(row, 0, item)
                    
                    if len(teachers) > 0 and teachers[0] is not None:
                        teacher = teachers[0]
                        # 性别
                        gender_text = ''
                        if teacher.gender == 'M':
                            gender_text = '男'
                        elif teacher.gender == 'F':
                            gender_text = '女'
                        item = QTableWidgetItem(gender_text)
                        item.setTextAlignment(Qt.AlignCenter)
                        self.subject_table.setItem(row, 1, item)
                        
                        # 来源
                        internal_text = ''
                        if teacher.is_internal is True:
                            internal_text = '本校'
                        elif teacher.is_internal is False:
                            internal_text = '外校'
                        item = QTableWidgetItem(internal_text)
                        item.setTextAlignment(Qt.AlignCenter)
                        self.subject_table.setItem(row, 2, item)
        
        # 调整列宽
        self.subject_table.resizeColumnsToContents()
        
    def on_mode_changed(self):
        """
        监考模式改变时的处理
        """
        if self.double_mode_radio.isChecked():
            self.gender_mix_checkbox.setEnabled(True)
            self.internal_mix_checkbox.setEnabled(True)
            # 双教师模式下，默认不选中“男女搭配”和“本外校搭配”
            self.gender_mix_checkbox.setChecked(False)
            self.internal_mix_checkbox.setChecked(False)
        else:
            self.gender_mix_checkbox.setEnabled(False)
            self.internal_mix_checkbox.setEnabled(False)
            self.gender_mix_checkbox.setChecked(False)
            self.internal_mix_checkbox.setChecked(False)

    def on_lock_imported_changed(self, checked):
        """
        锁定导入安排开关变化时的处理
        """
        if self.schedule:
            self.schedule.set_constraint('lock_imported', bool(checked))
        # 重新渲染以刷新颜色提示
        if self.schedule and self.schedule.exams:
            self.display_results(self.schedule.exams)

    
            
    def import_teachers(self):
        """
        导入教师信息
        """
        # 校验“考试场室数”是否已设置（非空）
        if self.room_spin.value() <= 0:
            QMessageBox.warning(self, '提示', '请先设置考场试室数')
            return

        file_path, _ = QFileDialog.getOpenFileName(
            self, '选择教师信息文件', '', 'Excel Files (*.xlsx *.xls)')
        
        if not file_path:
            return
            
        try:
            # 导入教师数据
            self.teachers = DataImporter.import_teachers_from_excel(file_path)
            
            # 验证数据（增强版）
            mode = "double" if self.double_mode_radio.isChecked() else "single"
            gender_mix = self.gender_mix_checkbox.isChecked()
            internal_mix = self.internal_mix_checkbox.isChecked()

            # 校验前检查科目信息是否已设置
            if not hasattr(self, 'subject_count') or not self.subject_count or self.subject_count <= 0:
                QMessageBox.warning(self, '数据验证失败', '请先在“科目设置”中导入或设置科目信息，再导入教师信息')
                self.teachers = []
                self.schedule_btn.setEnabled(False)
                return

            # 传入房间数用于预设考场范围校验
            num_rooms = self.room_spin.value()
            errors, warnings = DataImporter.validate_teachers(
                self.teachers,
                mode,
                gender_mix,
                internal_mix,
                subject_count=self.subject_count,
                subject_names=self.subject_names,
                source_file_path=file_path,
                num_rooms=num_rooms,
            )

            if errors:
                error_msg = '\n'.join(errors)
                self.log("数据验证失败")
                QMessageBox.warning(self, '数据验证失败', error_msg)
                self.teachers = []
                self.schedule_btn.setEnabled(False)
                return
            # 非致命的校验警告提示（例如预设监考考场越界或无法识别）
            if warnings:
                warn_msg = '\n'.join(warnings)
                self.log("存在校验警告，不影响导入")
                QMessageBox.information(self, '提示', f"存在以下警告（不影响导入）：\n{warn_msg}")
            
            self.log(f'导入成功，共导入{len(self.teachers)}名教师')
            QMessageBox.information(self, '导入成功', f'成功导入 {len(self.teachers)} 名教师')
            
            # 如果已有安排对象，更新教师列表
            if self.schedule:
                self.schedule.teachers = self.teachers

            self.schedule_btn.setEnabled(True)
            self.import_schedule_btn.setEnabled(True)  # 启用导入监考安排按钮
            if hasattr(self, 'preset_btn'):
                self.preset_btn.setEnabled(True)
            
        except Exception as e:
            QMessageBox.critical(self, '导入失败', str(e))
            
    def import_schedule(self):
        """
        导入监考安排
        """
        if not self.teachers:
            QMessageBox.warning(self, '警告', '请先导入教师信息')
            return
            
        file_path, _ = QFileDialog.getOpenFileName(
            self, '选择监考安排文件', '', 'Excel Files (*.xlsx *.xls)')
        
        if not file_path:
            return
            
        try:
            # 读取Excel文件中的监考总览表
            df = pd.read_excel(file_path, sheet_name='监考总览表')
            
            # 获取科目数和考场数
            # 修复：subject_spin是QLineEdit，应该使用text()方法并转换为整数
            num_subjects = int(self.subject_spin.text())
            num_rooms = self.room_spin.value()
            mode = "double" if self.double_mode_radio.isChecked() else "single"
            
            # 创建新的安排对象
            self.schedule = Schedule(self.teachers, num_subjects, num_rooms, mode)
            
            # 设置约束条件
            self.schedule.set_constraint('gender_mix', self.gender_mix_checkbox.isChecked())
            self.schedule.set_constraint('internal_mix', self.internal_mix_checkbox.isChecked())
            self.schedule.set_constraint('lock_imported', self.lock_imported_checkbox.isChecked())
            # 导入监考安排不启用红色标记
            self.schedule.set_constraint('highlight_imported', False)
            # 行为参数（仅代码控制）
            self.schedule.set_constraint('auto_postprocess_optimize', self.auto_optimize_postprocess)
            self.schedule.set_constraint('show_optimize_button', self.show_optimize_button_flag)
            
            # 设置科目时长约束（从subject_page获取实际时长）
            subject_durations = []
            
            # 尝试获取subject_page对象（与generate_schedule方法相同的逻辑）
            subject_page = None
            parent = self.parent()
            
            # 方式1: 从parent.subject_page获取
            if hasattr(parent, 'subject_page'):
                subject_page = parent.subject_page
            
            # 方式2: 从parent.parent().subject_page获取
            elif hasattr(parent, 'parent') and hasattr(parent.parent(), 'subject_page'):
                subject_page = parent.parent().subject_page
            
            # 方式3: 从parent.subject获取
            elif hasattr(parent, 'subject'):
                subject_page = parent.subject
            
            # 方式4: 尝试作为QStackedWidget的子页面获取
            elif hasattr(parent, 'count') and hasattr(parent, 'widget'):
                page_count = parent.count()
                for i in range(page_count):
                    page = parent.widget(i)
                    page_type = type(page).__name__
                    if 'SubjectPage' in page_type:
                        subject_page = page
                        break
            
            if subject_page and hasattr(subject_page, 'subjects') and hasattr(subject_page, 'get_subject_duration'):
                subject_count = len(subject_page.subjects)
                for i in range(subject_count):
                    duration = subject_page.get_subject_duration(i)
                    subject_durations.append(duration)
            else:
                # 无法获取subject_page，使用默认时长
                subject_durations = [120] * num_subjects  # 默认每科120分钟
            
            # 设置科目时长约束
            self.schedule.set_constraint('subject_durations', subject_durations)
            
            # 解析Excel数据并更新安排
            validation_errors = self.parse_schedule_from_excel(df)
            
            if validation_errors:
                error_msg = '\n'.join(validation_errors)
                QMessageBox.warning(self, '验证失败', f'导入的监考安排不满足约束条件：\n{error_msg}')
                return
                
            # 显示结果
            self.display_results(self.schedule.exams)

            # 启用相关按钮
            self.update_continue_button_state()
            self.adjust_btn.setEnabled(True)
            self.export_btn.setEnabled(True)
            # 根据是否全部安排完成控制优化按钮使能
            unassigned_count = 0
            for exam in self.schedule.exams:
                for room in exam.rooms:
                    if room not in exam.schedule:
                        unassigned_count += 1
                    elif self.double_mode_radio.isChecked():
                        teachers = exam.schedule[room]
                        if len(teachers) < 2 or None in teachers:
                            unassigned_count += 1
            self.optimize_btn.setEnabled(unassigned_count == 0)
            # 根据行为参数控制“二次均衡优化”按钮可见性
            self.optimize_btn.setVisible(self.schedule.get_constraint('show_optimize_button', False))
              
            
            QMessageBox.information(self, '导入成功', '监考安排已成功导入并更新')
            
        except Exception as e:
            QMessageBox.critical(self, '导入失败', f'导入监考安排时出错:\n{str(e)}')
            
    def parse_schedule_from_excel(self, df):
        """
        从Excel数据解析监考安排
        
        Args:
            df: DataFrame，包含监考总览表数据
            
        Returns:
            list: 验证错误信息列表
        """
        errors = []
        
        # 获取考场列表
        room_headers = df['考场'].tolist() if '考场' in df.columns else []
        if not room_headers:
            # 尝试从行索引获取考场信息
            room_headers = [f"考场{r}" for r in range(1, len(df) + 1)]
            
        rooms = []
        for header in room_headers:
            if isinstance(header, str) and header.startswith('考场'):
                try:
                    room_num = int(header[2:])  # 去掉"考场"前缀
                    rooms.append(room_num)
                except ValueError:
                    pass
            else:
                try:
                    room_num = int(header)
                    rooms.append(room_num)
                except (ValueError, TypeError):
                    pass
        
        # 初始化考试安排
        self.schedule.exams = []
        for subject_id in range(1, self.schedule.num_subjects + 1):
            exam = Exam(subject_id, list(range(1, self.schedule.num_rooms + 1)))
            self.schedule.exams.append(exam)
            
        # 清空教师已分配的场次和监考时长
        for teacher in self.schedule.teachers:
            teacher.assigned_sessions = []
            teacher.supervision_duration = 0
            
        # 解析数据
        if self.schedule.mode == "double":
            # 双教师模式：每科目两列
            for subject_id in range(1, self.schedule.num_subjects + 1):
                # 使用实际科目名称而不是"科目{subject_id}"
                subject_name = self.subject_names[subject_id-1] if subject_id-1 < len(self.subject_names) and self.subject_names[subject_id-1] else f"科目{subject_id}"
                exam_time = self.exam_times[subject_id-1] if subject_id-1 < len(self.exam_times) and self.exam_times[subject_id-1] else ""
                # 获取监考员1和监考员2列
                col1_name = f"{subject_name}-监考员1\n{exam_time}"
                col2_name = f"{subject_name}-监考员2\n{exam_time}"

                if col1_name not in df.columns or col2_name not in df.columns:
                    errors.append(f"科目{subject_name}缺少监考员信息")
                    continue
                    
                for i, room in enumerate(rooms):
                    if i >= len(df):
                        continue
                        
                    teacher1_name = str(df[col1_name].iloc[i]) if not pd.isna(df[col1_name].iloc[i]) else ""
                    teacher2_name = str(df[col2_name].iloc[i]) if not pd.isna(df[col2_name].iloc[i]) else ""
                    
                    # 处理空值情况
                    teacher1 = None
                    teacher2 = None
                    
                    if teacher1_name:
                        # 查找教师对象
                        teacher1 = next((t for t in self.schedule.teachers if t.name == teacher1_name), None)
                        if not teacher1:
                            errors.append(f"考场{room}科目{subject_name}中的监考员1({teacher1_name})未在教师信息中找到")
                            continue
                            
                        # 验证约束条件
                        # 1. 检查教师是否可以监考该科目
                        if not teacher1.can_supervise(subject_id):
                            errors.append(f"教师 {teacher1.name} 不能监考科目 {subject_name}")
                            continue
                            
                        # 2. 检查教师是否超过最大监考次数
                        if len(teacher1.assigned_sessions) >= teacher1.max_sessions:
                            errors.append(f"教师 {teacher1.name} 的监考次数已达到最大限制 ({teacher1.max_sessions})")
                            return errors  # 立即停止导入并报告错误
                        
                        # 获取科目时长
                        subject_durations = self.schedule.get_constraint('subject_durations', [])
                        duration = subject_durations[subject_id - 1] if (subject_id - 1) < len(subject_durations) else 0
                        # 将教师分配到考场
                        teacher1.assign((subject_id, room), duration)
                    
                    if teacher2_name:
                        # 查找教师对象
                        teacher2 = next((t for t in self.schedule.teachers if t.name == teacher2_name), None)
                        if not teacher2:
                            errors.append(f"考场{room}科目{subject_name}中的监考员2({teacher2_name})未在教师信息中找到")
                            continue
                            
                        # 验证约束条件
                        # 1. 检查教师是否可以监考该科目
                        if not teacher2.can_supervise(subject_id):
                            errors.append(f"教师 {teacher2.name} 不能监考科目 {subject_name}")
                            continue
                            
                        # 2. 检查教师是否超过最大监考次数
                        if len(teacher2.assigned_sessions) >= teacher2.max_sessions:
                            errors.append(f"教师 {teacher2.name} 的监考次数已达到最大限制 ({teacher2.max_sessions})")
                            return errors  # 立即停止导入并报告错误
                        
                        # 获取科目时长
                        subject_durations = self.schedule.get_constraint('subject_durations', [])
                        duration = subject_durations[subject_id - 1] if (subject_id - 1) < len(subject_durations) else 0
                        # 将教师分配到考场
                        teacher2.assign((subject_id, room), duration)
                    
                    # 检查搭配约束（只有当两个教师都存在时才检查）
                    if teacher1 and teacher2:
                        if not self.schedule.is_valid_pair(teacher1, teacher2):
                            error_detail = ""
                            if self.schedule.get_constraint('gender_mix') and teacher1.gender == teacher2.gender:
                                error_detail = "（性别不匹配）"
                            elif self.schedule.get_constraint('internal_mix') and teacher1.is_internal == teacher2.is_internal:
                                error_detail = "（本外校不匹配）"
                            errors.append(f"考场{room}科目{subject_name}的教师搭配不满足约束条件{error_detail}")
                            continue
                    
                    # 更新安排（即使只有一个教师或都没有教师）
                    exam = next((e for e in self.schedule.exams if e.subject_id == subject_id), None)
                    if exam:
                        teachers_list = []
                        if teacher1:
                            teachers_list.append(teacher1)
                        if teacher2:
                            # 确保在双教师模式下，即使teacher1为空也要保留位置
                            if not teacher1 and len(teachers_list) == 0:
                                teachers_list.append(None)  # 占位符，表示监考员1为空
                                teachers_list.append(teacher2)  # 监考员2
                            else:
                                teachers_list.append(teacher2)
                        exam.schedule[room] = teachers_list
                        # 标记导入位置
                        if teacher1:
                            self.schedule.mark_imported_position(subject_id, room, 0)
                        if teacher2:
                            self.schedule.mark_imported_position(subject_id, room, 1)
        else:
            # 单教师模式：每科目一列
            for subject_id in range(1, self.schedule.num_subjects + 1):
                # 使用实际科目名称而不是"科目{subject_id}"
                subject_name = self.subject_names[subject_id-1] if subject_id-1 < len(self.subject_names) and self.subject_names[subject_id-1] else f"科目{subject_id}"
                exam_time = self.exam_times[subject_id-1] if subject_id-1 < len(self.exam_times) and self.exam_times[subject_id-1] else ""
                col_name = f"{subject_name}\n{exam_time}"
                if col_name not in df.columns:
                    # 如果使用实际名称找不到列，尝试使用默认名称
                    col_name = f"科目{subject_id}"
                    if col_name not in df.columns:
                        errors.append(f"找不到科目 {subject_name} 对应的列")
                        continue
                    
                for i, room in enumerate(rooms):
                    if i >= len(df):
                        continue
                        
                    teacher_names = str(df[col_name].iloc[i]) if not pd.isna(df[col_name].iloc[i]) else ""
                    if not teacher_names:
                        # 空值表示未安排监考，直接跳过
                        continue
                        
                    # 查找教师对象
                    teacher = next((t for t in self.schedule.teachers if t.name == teacher_names), None)
                    if not teacher:
                        errors.append(f"考场{room}科目{subject_name}中的教师未在教师信息中找到")
                        continue
                        
                    # 验证约束条件
                    # 1. 检查教师是否可以监考该科目
                    if not teacher.can_supervise(subject_id):
                        errors.append(f"教师 {teacher.name} 不能监考科目 {subject_name}")
                        continue
                        
                    # 2. 检查教师是否超过最大监考次数
                    if len(teacher.assigned_sessions) >= teacher.max_sessions:
                        errors.append(f"教师 {teacher.name} 的监考次数已达到最大限制 ({teacher.max_sessions})")
                        return errors  # 立即停止导入并报告错误
                    
                    # 更新安排
                    exam = next((e for e in self.schedule.exams if e.subject_id == subject_id), None)
                    if exam:
                        exam.schedule[room] = [teacher]
                        # 获取科目时长
                        subject_durations = self.schedule.get_constraint('subject_durations', [])
                        duration = subject_durations[subject_id - 1] if (subject_id - 1) < len(subject_durations) else 0
                        teacher.assign((subject_id, room), duration)
                        # 标记导入位置
                        self.schedule.mark_imported_position(subject_id, room, 0)
                        
        # 检查教师监考次数是否超过限制（原来的检查保留作为最后的验证）
        for teacher in self.schedule.teachers:
            if len(teacher.assigned_sessions) > teacher.max_sessions:
                errors.append(f"教师 {teacher.name} 的监考次数超过限制 ({len(teacher.assigned_sessions)} > {teacher.max_sessions})")
                
        return errors
            
    def generate_schedule(self):
        """
        生成监考安排
        """
        if not self.teachers:
            QMessageBox.warning(self, '警告', '请先导入教师信息')
            return
            
        # 检查是否已有完整安排，如果有的话提示用户是否重新生成
        if self.schedule and self.schedule.is_schedule_complete():
            reply = QMessageBox.question(self, '确认', '当前监考安排已经完整，是否要重新生成安排？',
                                         QMessageBox.Yes | QMessageBox.No, QMessageBox.No)
            if reply == QMessageBox.No:
                return
            
        # 获取参数
        num_subjects = self.subject_count
        num_rooms = self.room_spin.value()
        mode = "double" if self.double_mode_radio.isChecked() else "single"
        
        # 如果还没有安排对象，则创建一个新的
        if not self.schedule:
            self.schedule = Schedule(self.teachers, num_subjects, num_rooms, mode)
        else:
            # 如果已存在安排对象，检查模式是否改变
            if self.schedule.mode != mode:
                # 模式改变，创建新的安排对象
                self.schedule = Schedule(self.teachers, num_subjects, num_rooms, mode)
            else:
                # 模式未改变，更新参数和教师列表
                self.schedule.num_subjects = num_subjects
                self.schedule.num_rooms = num_rooms
                self.schedule.teachers = self.teachers
                # 重新初始化考试安排
                self.schedule.exams = []
            
        # 设置约束条件
        self.schedule.set_constraint('gender_mix', self.gender_mix_checkbox.isChecked())
        self.schedule.set_constraint('internal_mix', self.internal_mix_checkbox.isChecked())
        self.schedule.set_constraint('lock_imported', self.lock_imported_checkbox.isChecked())
        # 设置均衡模式
        balance_mode = 'session' if self.session_balance_radio.isChecked() else 'duration'
        self.schedule.set_constraint('balance_mode', balance_mode)
        # 设置二次优化行为参数（仅代码控制）
        self.schedule.set_constraint('auto_postprocess_optimize', self.auto_optimize_postprocess)
        self.schedule.set_constraint('show_optimize_button', self.show_optimize_button_flag)

        # 获取科目时长信息
        subject_durations = []
        parent = self.parent()
        
        # 尝试多种方式获取subject_page
        subject_page = None
        
        # 方式1: 直接从parent获取subject_page属性
        if hasattr(parent, 'subject_page'):
            subject_page = parent.subject_page
        
        # 方式2: 从parent.pages['subject']获取
        elif hasattr(parent, 'pages') and 'subject' in parent.pages:
            subject_page = parent.pages['subject']
        
        # 方式3: 从parent.subject获取
        elif hasattr(parent, 'subject'):
            subject_page = parent.subject
        
        # 方式4: 尝试作为QStackedWidget的子页面获取
        elif hasattr(parent, 'count') and hasattr(parent, 'widget'):
            page_count = parent.count()
            for i in range(page_count):
                page = parent.widget(i)
                page_type = type(page).__name__
                if 'SubjectPage' in page_type:
                    subject_page = page
                    break
        
        if subject_page and hasattr(subject_page, 'subjects') and hasattr(subject_page, 'get_subject_duration'):
            subject_count = len(subject_page.subjects)
            for i in range(subject_count):
                duration = subject_page.get_subject_duration(i)
                subject_durations.append(duration)
        else:
            # 无法获取subject_page，不设置默认时长
            subject_durations = []
        
        # 输出获取到的科目时长，用于调试
        # print(f"获取到的科目时长列表: {subject_durations}")
        
        # 设置科目时长约束
        self.schedule.set_constraint('subject_durations', subject_durations)

        # 注入精确进度回调，使进度与后端步骤完全同步
        self.schedule.set_constraint('progress_callback', self.make_progress_reporter('生成监考安排'))
        
        # 双监考且启用性别/本外校约束时，先进行可行性预判，避免无谓长时间尝试
        if mode == "double" and (self.schedule.get_constraint('gender_mix') or self.schedule.get_constraint('internal_mix')):
            feasible, reason = self.schedule.check_feasibility()
            if not feasible:
                QMessageBox.warning(self, '不可行', f'当前参数下双教师监考不可行：{reason}\n请调整教师资源或放宽约束后重试。')
                return
        
        # 将双监考尝试次数降到30以减少UI阻塞；单监考保留为100
        max_attempts = 30 if mode == "double" else 100  # 最大尝试次数
        attempt = 0
        unassigned_count = float('inf')
        best_schedule = None
        best_exams = None
        
        # 在状态栏与页面内显示忙碌进度
        if hasattr(parent, 'start_busy'):
            parent.start_busy('正在生成监考安排…')
        try:
            self.start_local_busy('正在生成监考安排…')
        except Exception:
            pass

        # 重试机制：最多尝试max_attempts次
        while attempt < max_attempts and unassigned_count > 0:
            attempt += 1
            
            # 更新外部与页面内的进度显示
            try:
                if hasattr(parent, 'update_busy'):
                    parent.update_busy(message=f'正在生成监考安排（尝试 {attempt}/{max_attempts}）', percent=None)
                self.update_local_busy(message=f'正在生成监考安排（尝试 {attempt}/{max_attempts}）', percent=None)
                QApplication.processEvents()
            except Exception:
                pass

            try:
                # 生成安排（在现有安排基础上继续安排）
                exams, unassigned_count = self.schedule.generate_schedule()
                
                # 保存当前最好的安排（未分配考场数最少的）
                if unassigned_count < float('inf') and unassigned_count >= 0:
                    if not best_schedule or unassigned_count < best_schedule[0]:
                        best_schedule = (unassigned_count, self.schedule)
                        best_exams = exams
                        
            except Exception as e:
                if attempt == max_attempts:
                    QMessageBox.critical(self, '安排失败', f'生成监考安排时出错:\n{str(e)}')
                continue
        
        # 使用最好的安排结果
        if best_schedule:
            unassigned_count, self.schedule = best_schedule
            exams = best_exams
        else:
            QMessageBox.critical(self, '安排失败', '无法生成监考安排')
            # 结束忙碌显示（在用户关闭对话框后再隐藏进度）
            try:
                if hasattr(parent, 'stop_busy'):
                    parent.stop_busy('当前任务：监考编排')
                self.stop_local_busy()
            except Exception:
                pass
            return
            
        # 在自动优化开启且全部安排完成的情况下，先进行静默优化，再展示结果
        if unassigned_count == 0 and self.schedule.get_constraint('auto_postprocess_optimize', True):
            # 静默执行二次均衡优化，不弹窗不日志，优化完成后再展示
            self.run_postprocess_optimization(silent=True)
            self.display_results(self.schedule.exams)
        else:
            # 未全部完成或未开启自动优化，直接展示当前结果
            self.display_results(exams)
        
        # 启用相关按钮
        self.export_btn.setEnabled(True)
        self.adjust_btn.setEnabled(True)
        # 二次均衡优化按钮仅在全部安排完成后才允许点击
        self.optimize_btn.setEnabled(unassigned_count == 0)
        # 根据行为参数控制“二次均衡优化”按钮可见性
        self.optimize_btn.setVisible(self.schedule.get_constraint('show_optimize_button', False))
        # 更新“补全监考安排”按钮状态（完成则禁用，未完成则启用）
        self.update_continue_button_state()
        
        # 检查是否所有考场都已安排
        total_rooms = num_subjects * num_rooms
        if unassigned_count == 0:
            QMessageBox.information(self, '安排完成', f'监考安排已生成，所有科目和考场均已成功安排监考人员。')
            # 自动优化已在展示前处理，此处不再重复执行
        else:
            assigned_count = total_rooms - unassigned_count
            QMessageBox.warning(self, '安排完成但不完整', 
                              f'监考安排已生成，但只有{assigned_count}个考场成功安排了监考人员，仍有{unassigned_count}个考场未能安排。\n\n'
                              f'这可能是因为监考人员不足或约束条件过于严格导致的。\n尝试次数: {attempt}/{max_attempts}')
        # 结束忙碌显示（在用户关闭对话框后再隐藏进度）
        try:
            if hasattr(parent, 'stop_busy'):
                parent.stop_busy('当前任务：监考编排')
            self.stop_local_busy()
        except Exception:
            pass

    def continue_schedule(self):
        """
        继续为未安排的考场分配监考教师
        """
        if not self.schedule or not self.schedule.exams:
            QMessageBox.warning(self, '错误', '请先导入或生成监考安排!')
            return
        # 补全前，刷新一次“约束条件设置”，以便沿用当前勾选状态
        try:
            self.schedule.set_constraint('gender_mix', self.gender_mix_checkbox.isChecked())
            self.schedule.set_constraint('internal_mix', self.internal_mix_checkbox.isChecked())
            self.schedule.set_constraint('lock_imported', self.lock_imported_checkbox.isChecked())
            balance_mode = 'session' if self.session_balance_radio.isChecked() else 'duration'
            self.schedule.set_constraint('balance_mode', balance_mode)
            # 同步最新科目时长，用于补全与时长均衡
            subject_durations = []
            subject_page = None
            parent = self.parent()
            if hasattr(parent, 'subject_page'):
                subject_page = parent.subject_page
            elif hasattr(parent, 'parent') and hasattr(parent.parent(), 'subject_page'):
                subject_page = parent.parent().subject_page
            elif hasattr(parent, 'subject'):
                subject_page = parent.subject
            elif hasattr(parent, 'count') and hasattr(parent, 'widget'):
                page_count = parent.count()
                for i in range(page_count):
                    page = parent.widget(i)
                    page_type = type(page).__name__
                    if 'SubjectPage' in page_type:
                        subject_page = page
                        break
            if subject_page and hasattr(subject_page, 'subjects') and hasattr(subject_page, 'get_subject_duration'):
                subject_count = len(subject_page.subjects)
                for i in range(subject_count):
                    duration = subject_page.get_subject_duration(i)
                    subject_durations.append(duration)
            else:
                subject_durations = [120] * self.schedule.num_subjects
            self.schedule.set_constraint('subject_durations', subject_durations)
        except Exception:
            pass
            
        # 统计未完成的考场数
        unassigned_count = 0
        for exam in self.schedule.exams:
            for room in exam.rooms:
                if room not in exam.schedule:
                    unassigned_count += 1
                elif self.double_mode_radio.isChecked():
                    teachers = exam.schedule[room]
                    if len(teachers) < 2 or None in teachers:
                        unassigned_count += 1
        
        if unassigned_count == 0:
            # 已全部完成：若启用自动优化，先静默优化，再展示结果
            if self.schedule.get_constraint('auto_postprocess_optimize', True):
                self.run_postprocess_optimization(silent=True)
            self.display_results(self.schedule.exams)
            QMessageBox.information(self, '提示', '所有考场都已安排完成!')
            # 根据行为参数更新按钮可见性与使能
            self.optimize_btn.setVisible(self.schedule.get_constraint('show_optimize_button', False))
            self.optimize_btn.setEnabled(True)
            # 安排已完成，禁用“补全监考安排”按钮
            self.update_continue_button_state()
            return
        
        # 尝试继续安排
        retry_count = 0
        max_retries = 100
        # 显示忙碌进度
        parent = self.parent()
        if hasattr(parent, 'start_busy'):
            parent.start_busy('正在补全监考安排…')
        try:
            self.start_local_busy('正在补全监考安排…')
        except Exception:
            pass
        # 注入精确进度回调，使进度与后端步骤完全同步
        if self.schedule:
            self.schedule.set_constraint('progress_callback', self.make_progress_reporter('补全监考安排'))
        while retry_count < max_retries:
            # 更新进度
            try:
                if hasattr(parent, 'update_busy'):
                    parent.update_busy(message=f'正在补全监考安排（尝试 {retry_count+1}/{max_retries}）', percent=None)
                self.update_local_busy(message=f'正在补全监考安排（尝试 {retry_count+1}/{max_retries}）', percent=None)
                QApplication.processEvents()
            except Exception:
                pass
            success, message = self.schedule.continue_schedule()
            if success:
                # 若启用自动优化，先静默优化，再展示结果
                if self.schedule.get_constraint('auto_postprocess_optimize', True):
                    self.run_postprocess_optimization(silent=True)
                self.display_results(self.schedule.exams)
                QMessageBox.information(self, '成功', '剩余考场监考安排完成!')
                # 补全完成后，更新按钮可见性与使能（仅当全部完成时允许点击）
                self.optimize_btn.setVisible(self.schedule.get_constraint('show_optimize_button', False))
                self.optimize_btn.setEnabled(True)
                # 补全完成后，禁用“补全监考安排”按钮
                self.update_continue_button_state()
                # 结束忙碌显示
                try:
                    if hasattr(parent, 'stop_busy'):
                        parent.stop_busy('当前任务：监考编排')
                    self.stop_local_busy()
                except Exception:
                    pass
                return
            retry_count += 1
        
        QMessageBox.warning(self, '警告', 
                        f'尝试{max_retries}次后仍有考场未能安排。\n'
                        f'原因: {message}\n'
                        '建议调整参数或约束条件后重试。')
        # 结束忙碌显示（在用户关闭对话框后再隐藏进度）
        try:
            if hasattr(parent, 'stop_busy'):
                parent.stop_busy('当前任务：监考编排')
            self.stop_local_busy()
        except Exception:
            pass

    def on_optimize_button_clicked(self):
        """
        优化按钮点击事件：执行非静默的二次均衡优化（保留弹窗与日志）。
        """
        self.run_postprocess_optimization(silent=False)

    def run_postprocess_optimization(self, silent=False):
        """
        运行二次均衡优化（后处理均衡）。
        当 silent=True 时：不弹出任何对话框、不输出日志，也不主动刷新表格显示；由调用方在优化完成后再统一展示。
        """
        if not self.schedule or not self.schedule.exams:
            if not silent:
                QMessageBox.warning(self, '提示', '请先生成或导入监考安排')
            return

        # 显示忙碌进度（仅非静默模式显示优化进度，静默模式保持调用方的进度不变）
        parent = self.parent()
        if not silent:
            if hasattr(parent, 'start_busy'):
                parent.start_busy('生成监考安排 - 均衡优化中：')
            try:
                self.start_local_busy('生成监考安排 - 均衡优化中：')
            except Exception:
                pass

        try:
            # 注入设置：日志开关始终传递；进度回调仅非静默模式
            if self.schedule:
                self.schedule.set_constraint('log_optimization_swaps', self.log_optimization_swaps)
                # 新增：优化阶段尊重预设房间，避免把已在预设的教师换走
                self.schedule.set_constraint('respect_preset_on_swap', True)
            if self.schedule and not silent:
                # 固定提示文本，忽略后端具体步骤信息
                self.schedule.set_constraint('progress_callback', self.make_progress_reporter('生成监考安排 - 均衡优化中：', fixed_message=''))
            # 先执行二次均衡优化
            report = self.schedule.optimize_duration_postprocess(max_passes=40)
            # 优化完成后执行预设房间修复
            preset_report = self.schedule.enforce_preset_room_postprocess()
        except Exception as e:
            if not silent:
                QMessageBox.critical(self, '优化失败', f'运行二次均衡优化时出错:\n{str(e)}')
            return

        # 非静默模式下，刷新显示以反映交换与预设修复后的结果
        if not silent:
            self.display_results(self.schedule.exams)

        # 汇总提示信息（仅非静默模式）
        before_overall = report.get('before', {}).get('max_overall', 0)
        after_overall = report.get('after', {}).get('max_overall', 0)
        before_current = report.get('before', {}).get('max_current', 0)
        after_current = report.get('after', {}).get('max_current', 0)
        swap_count = report.get('swap_count', 0)
        early_reason = report.get('early_stop_reason')
        if not silent:
            if swap_count > 0:
                msg = (
                    f"已完成二次均衡优化，共交换 {swap_count} 次。\n"
                    f"最大总时长：{before_overall} → {after_overall} 分钟。\n"
                    f"最大本次时长：{before_current} → {after_current} 分钟。"
                )
                # 追加预设房间修复提示
                try:
                    preset_moves = preset_report.get('moves', 0) if 'preset_report' in locals() else 0
                except Exception:
                    preset_moves = 0
                if preset_moves > 0:
                    msg += f"\n已应用预设房间修复 {preset_moves} 次。"
                if early_reason:
                    msg += "\n提示：本次优化已提前结束（判定为当前算法均衡瓶颈）。"
                QMessageBox.information(self, '二次均衡优化', msg)
            else:
                msg = '没有可行的改善；当前安排已较为均衡。'
                try:
                    preset_moves = preset_report.get('moves', 0) if 'preset_report' in locals() else 0
                except Exception:
                    preset_moves = 0
                if preset_moves > 0:
                    msg += f"\n已应用预设房间修复 {preset_moves} 次。"
                if early_reason:
                    msg += "\n提示：本次优化已提前结束（判定为当前算法均衡瓶颈）。"
                QMessageBox.information(self, '二次均衡优化', msg)

            # 如果有提前结束原因，输出到日志区
            if early_reason:
                self.log(f"[二次均衡] 提前结束原因：{early_reason}")

            # 写入详细日志（不设上限，逐次输出关键指标与交换对象）
            self.log('二次均衡优化详情（每次交换后的指标与对象）：')
        if not silent:
            # 先输出基线
            base = report.get('before', {})
            try:
                self.log(
                    f"初始指标：最大总时长={base.get('max_overall', 0)}, "
                    f"最大本次时长={base.get('max_current', 0)}, "
                    f"总时长方差={base.get('var_overall', 0):.4f}, "
                    f"本次时长方差={base.get('var_current', 0):.4f}"
                )
            except Exception:
                self.log(
                    f"初始指标：最大总时长={base.get('max_overall', 0)}, 最大本次时长={base.get('max_current', 0)}, "
                    f"总时长方差={base.get('var_overall', 0)}, 本次时长方差={base.get('var_current', 0)}"
                )

            swaps = report.get('swaps', [])
            for i, s in enumerate(swaps):
                heavy = s.get('heavy', '')
                light = s.get('light', '')
                frm = s.get('from', {})
                to = s.get('to', {})
                mo = s.get('max_overall')
                mc = s.get('max_current')
                vo = s.get('var_overall')
                vc = s.get('var_current')
                prefix = (
                    f"{i+1}. 最大总时长={mo}, 最大本次时长={mc}, "
                    + (f"总时长方差={vo:.4f}, 本次时长方差={vc:.4f}" if isinstance(vo, (int, float)) and isinstance(vc, (int, float)) else f"总时长方差={vo}, 本次时长方差={vc}")
                )
                detail = (
                    f"；重载教师={heavy} 的科目{frm.get('subject')}考场{frm.get('room')}({frm.get('duration')}分钟) ? "
                    f"轻载教师={light} 的科目{to.get('subject')}考场{to.get('room')}({to.get('duration')}分钟)"
                )
                self.log(prefix + detail)

        # 结束忙碌显示（仅非静默模式；静默模式由调用方在对话框关闭后统一隐藏）
        if not silent:
            try:
                if hasattr(parent, 'stop_busy'):
                    parent.stop_busy('当前任务：监考编排')
                self.stop_local_busy()
            except Exception:
                pass

    def update_subject_names(self, subject_names, exam_times):
        """
        更新科目名称列表
        """
        self.subject_names = subject_names
        self.exam_times = exam_times
        # 如果已经生成了安排结果，需要更新显示
        if self.schedule and self.schedule.exams:
            self.display_results(self.schedule.exams)
            
    def update_subject_count(self, count):
        """
        更新科目数
        """
        self.subject_count = count
        self.subject_spin.setText(str(count))  # 更新UI显示
        # 确保科目名称列表长度与科目数一致
        while len(self.subject_names) < count:
            self.subject_names.append(f"科目{len(self.subject_names) + 1}")
        self.subject_names = self.subject_names[:count]
        
    def display_results(self, exams):
        """
        显示安排结果
        """
        # 清空现有数据
        # 初始化监考总览表
        self.overview_table.setRowCount(0)
        self.overview_table.setColumnCount(0)
        
        # 初始化监考统计表
        self.statistics_table.setRowCount(0)
        self.statistics_table.setColumnCount(0)
        
        # 初始化分科目监考表
        self.subject_table.setRowCount(0)
        self.subject_table.setColumnCount(0)
        
        # 更新科目选择下拉框
        self.subject_combo.clear()
        if exams:
            num_subjects = max([exam.subject_id for exam in exams])
            for i in range(1, num_subjects + 1):
                # 使用实际科目名称
                subject_name = self.subject_names[i-1] if i-1 < len(self.subject_names) and self.subject_names[i-1] else f"科目{i}"
                self.subject_combo.addItem(subject_name)
            # 默认选择第一个科目（科目1）
            if num_subjects >= 1:
                self.subject_combo.setCurrentIndex(0)
        
        # 显示监考总览表（修改为行列格式：行=考场，列=科目）
        # 获取科目数和考场数
        num_subjects = max([exam.subject_id for exam in exams]) if exams else 0
        all_rooms = set()
        for exam in exams:
            all_rooms.update(exam.rooms) 
        all_rooms = sorted(list(all_rooms))
        num_rooms = len(all_rooms)
        
        # 设置表格行列数和标题
        self.overview_table.setRowCount(num_rooms)
        
        # 根据监考模式设置列数和列标题
        if self.double_mode_radio.isChecked():
            # 双教师模式：每科目两列（监考员1和监考员2）
            self.overview_table.setColumnCount(num_subjects * 2)
            column_labels = []
            for i in range(1, num_subjects + 1):
                # 使用实际科目名称
                subject_name = self.subject_names[i-1] if i-1 < len(self.subject_names) and self.subject_names[i-1] else f"科目{i}"
                exam_time = self.exam_times[i-1] if i-1 < len(self.exam_times) and self.exam_times[i-1] else ""
                column_labels.extend([f"{subject_name}-监考员1\n{exam_time}", f"{subject_name}-监考员2\n{exam_time}"]) 
            self.overview_table.setHorizontalHeaderLabels(column_labels)
        else:
            # 单教师模式：每科目一列
            self.overview_table.setColumnCount(num_subjects)
            # 使用实际科目名称
            column_labels = []
            for i in range(1, num_subjects + 1):
                subject_name = self.subject_names[i-1] if i-1 < len(self.subject_names) and self.subject_names[i-1] else f"科目{i}"
                exam_time = self.exam_times[i-1] if i-1 < len(self.exam_times) and self.exam_times[i-1] else ""
                column_labels.append(f"{subject_name}\n{exam_time}")
            self.overview_table.setHorizontalHeaderLabels(column_labels)
        
        # 设置行标题（考场）
        self.overview_table.setVerticalHeaderLabels([f"考场{r}" for r in all_rooms])
        
        # 填充数据
        room_index_map = {room: idx for idx, room in enumerate(all_rooms)}
        for exam in exams:
            for room in exam.rooms:
                if room in exam.schedule:
                    teachers = exam.schedule[room]
                    room_row = room_index_map[room]
                    
                    if self.double_mode_radio.isChecked():
                        # 双教师模式：分别填充监考员1和监考员2
                        subject_col_start = (exam.subject_id - 1) * 2
                        
                        # 处理教师列表，确保正确显示空值位置
                        teacher1 = None
                        teacher2 = None
                        
                        if len(teachers) >= 2:
                            teacher1 = teachers[0]
                            teacher2 = teachers[1]
                        elif len(teachers) == 1:
                            # 判断是监考员1还是监考员2
                            if teachers[0] is not None:
                                teacher1 = teachers[0]
                            else:
                                teacher2 = teachers[0]  # 这种情况是监考员2有值但监考员1为空
                        
                        # 监考员1
                        if teacher1:
                            # 根据状态添加标记并设置颜色：红色+[锁]（导入锁定且高亮启用）> 绿色+[预]（在预设房间）> 性别颜色
                            highlight_imported = bool(self.schedule and self.schedule.get_constraint('highlight_imported', False))
                            display_name1 = teacher1.name
                            if highlight_imported and self.schedule and self.schedule.get_constraint('lock_imported') and self.schedule.is_position_imported(exam.subject_id, room, 0):
                                display_name1 += '[锁]'
                                item1 = QTableWidgetItem(display_name1)
                                item1.setTextAlignment(Qt.AlignCenter)
                                item1.setForeground(Qt.red)
                            else:
                                # 仅当教师当前显示在其预设考场时，绿色并标注[预]
                                preset_room1 = getattr(teacher1, 'preset_room', None)
                                if preset_room1 is not None and int(preset_room1) == int(room):
                                    display_name1 += '[预]'
                                    item1 = QTableWidgetItem(display_name1)
                                    item1.setTextAlignment(Qt.AlignCenter)
                                    item1.setForeground(Qt.green)
                                else:
                                    item1 = QTableWidgetItem(display_name1)
                                    item1.setTextAlignment(Qt.AlignCenter)
                                    if teacher1.gender == 'M':
                                        item1.setForeground(Qt.blue)
                                    elif teacher1.gender == 'F':
                                        item1.setForeground(Qt.magenta)
                            self.overview_table.setItem(room_row, subject_col_start, item1)
                        # 监考员2
                        if teacher2:
                            # 根据状态添加标记并设置颜色：红色+[锁]（导入锁定且高亮启用）> 绿色+[预]（在预设房间）> 性别颜色
                            highlight_imported = bool(self.schedule and self.schedule.get_constraint('highlight_imported', False))
                            display_name2 = teacher2.name
                            if highlight_imported and self.schedule and self.schedule.get_constraint('lock_imported') and self.schedule.is_position_imported(exam.subject_id, room, 1):
                                display_name2 += '[锁]'
                                item2 = QTableWidgetItem(display_name2)
                                item2.setTextAlignment(Qt.AlignCenter)
                                item2.setForeground(Qt.red)
                            else:
                                preset_room2 = getattr(teacher2, 'preset_room', None)
                                if preset_room2 is not None and int(preset_room2) == int(room):
                                    display_name2 += '[预]'
                                    item2 = QTableWidgetItem(display_name2)
                                    item2.setTextAlignment(Qt.AlignCenter)
                                    item2.setForeground(Qt.green)
                                else:
                                    item2 = QTableWidgetItem(display_name2)
                                    item2.setTextAlignment(Qt.AlignCenter)
                                    if teacher2.gender == 'M':
                                        item2.setForeground(Qt.blue)
                                    elif teacher2.gender == 'F':
                                        item2.setForeground(Qt.magenta)
                            self.overview_table.setItem(room_row, subject_col_start + 1, item2)
                    else:
                        # 单教师模式
                        # 修改：过滤掉None值后再访问name属性
                        # 单教师模式：添加标记和颜色（红色+[锁]优先；绿色+[预]其预设房间；否则默认/性别颜色）
                        subject_col = exam.subject_id - 1  # 转为0基索引
                        t0 = next((t for t in teachers if t is not None), None)
                        if t0 is not None:
                            highlight_imported = bool(self.schedule and self.schedule.get_constraint('highlight_imported', False))
                            display_name = t0.name
                            if highlight_imported and self.schedule and self.schedule.get_constraint('lock_imported') and self.schedule.is_position_imported(exam.subject_id, room, 0):
                                display_name += '[锁]'
                                item = QTableWidgetItem(display_name)
                                item.setTextAlignment(Qt.AlignCenter)
                                item.setForeground(Qt.red)
                            else:
                                preset_room0 = getattr(t0, 'preset_room', None)
                                if preset_room0 is not None and int(preset_room0) == int(room):
                                    display_name += '[预]'
                                    item = QTableWidgetItem(display_name)
                                    item.setTextAlignment(Qt.AlignCenter)
                                    item.setForeground(Qt.green)
                                else:
                                    item = QTableWidgetItem(display_name)
                                    item.setTextAlignment(Qt.AlignCenter)
                                    # 单教师模式原本不按性别着色，这里保持不着色以避免冲突
                            self.overview_table.setItem(room_row, subject_col, item)
                        else:
                            # 无教师信息时显示空白
                            item = QTableWidgetItem('')
                            item.setTextAlignment(Qt.AlignCenter)
                            self.overview_table.setItem(room_row, subject_col, item)
        
        # 显示监考统计表
        stats = self.schedule.get_statistics()
        # 修改统计表列数和标题，添加更多教师信息和科目监考情况
        max_subject_id = max([exam.subject_id for exam in exams]) if exams else 0
        # 使用实际科目名称
        stat_columns = ['教师姓名', '性别', '是否本校', '最大监考段数', '剩余监考次数', '不监考科目']
        for i in range(1, max_subject_id+1):
            subject_name = self.subject_names[i-1] if i-1 < len(self.subject_names) and self.subject_names[i-1] else f"科目{i}"
            stat_columns.append(subject_name)
        stat_columns.append('监考次数')
        
        # 添加历次、本次与总时长列（本次已存在为“监考时长(分钟)”）
        stat_columns.append('监考时长(分钟)')
        stat_columns.append('历次监考时长（分钟）')
        stat_columns.append('总监考时长(分钟)')
        self.statistics_table.setColumnCount(len(stat_columns))
        self.statistics_table.setHorizontalHeaderLabels(stat_columns)
        self.statistics_table.setRowCount(len(stats))
        
        for i, stat in enumerate(stats):
            # 查找教师对象
            teacher = next((t for t in self.schedule.teachers if t.name == stat['name']), None)
            if teacher:
                # 填充教师基本信息
                item = QTableWidgetItem(teacher.name)
                item.setTextAlignment(Qt.AlignCenter)
                self.statistics_table.setItem(i, 0, item)
                
                # 性别信息
                gender_text = ''
                if teacher.gender == 'M':
                    gender_text = '男'
                elif teacher.gender == 'F':
                    gender_text = '女'
                item = QTableWidgetItem(gender_text)
                item.setTextAlignment(Qt.AlignCenter)
                self.statistics_table.setItem(i, 1, item)
                
                # 是否本校信息
                internal_text = ''
                if teacher.is_internal is True:
                    internal_text = '是'
                elif teacher.is_internal is False:
                    internal_text = '否'
                item = QTableWidgetItem(internal_text)
                item.setTextAlignment(Qt.AlignCenter)
                self.statistics_table.setItem(i, 2, item)
                
                # 最大监考段数
                item = QTableWidgetItem(str(teacher.max_sessions))
                item.setTextAlignment(Qt.AlignCenter)
                self.statistics_table.setItem(i, 3, item)
                
                # 剩余监考次数 = 最大监考段数 - 目前已监考次数
                remaining_sessions = teacher.max_sessions - stat['count']
                item = QTableWidgetItem(str(remaining_sessions))
                item.setTextAlignment(Qt.AlignCenter)
                self.statistics_table.setItem(i, 4, item)
                
                # 不监考科目
                unavailable_subjects_str = ','.join(map(str, teacher.unavailable_subjects))
                item = QTableWidgetItem(unavailable_subjects_str)
                item.setTextAlignment(Qt.AlignCenter)
                self.statistics_table.setItem(i, 5, item)
                
                # 各科目监考情况 (1表示监考，0表示不监考)
                for j in range(1, max_subject_id+1):
                    is_assigned = 1 if teacher.is_assigned_to_subject(j) else 0
                    item = QTableWidgetItem(str(is_assigned))
                    item.setTextAlignment(Qt.AlignCenter)
                    self.statistics_table.setItem(i, 5+j, item)
                
                # 监考次数
                item = QTableWidgetItem(str(stat['count']))
                item.setTextAlignment(Qt.AlignCenter)
                self.statistics_table.setItem(i, 5+max_subject_id+1, item)
                
                # 本次监考时长
                duration_item = QTableWidgetItem(str(teacher.supervision_duration))
                duration_item.setTextAlignment(Qt.AlignCenter)
                self.statistics_table.setItem(i, 5+max_subject_id+2, duration_item)

                # 历次监考时长
                prev_item = QTableWidgetItem(str(teacher.previous_supervision_duration))
                prev_item.setTextAlignment(Qt.AlignCenter)
                self.statistics_table.setItem(i, 5+max_subject_id+3, prev_item)

                # 总监考时长（历次 + 本次）
                total_duration = (teacher.supervision_duration or 0) + (teacher.previous_supervision_duration or 0)
                total_item = QTableWidgetItem(str(total_duration))
                total_item.setTextAlignment(Qt.AlignCenter)
                self.statistics_table.setItem(i, 5+max_subject_id+4, total_item)
            
                # 显示分科目监考表（默认显示科目1）
                # 默认会显示第一个科目（科目1）的数据，因为我们设置了currentIndex为0
                
                # 调整列宽
                for table in [self.overview_table, self.statistics_table, self.subject_table]:
                    table.resizeColumnsToContents()

    def open_preset_dialog(self):
        """
        打开“预设监考安排”对话框，包含“生成空监考表”和“导入预设安排”。
        """
        dialog = QDialog(self)
        dialog.setWindowTitle('预设监考安排')
        layout = QVBoxLayout(dialog)
        layout.setContentsMargins(16, 12, 16, 12)
        layout.setSpacing(10)

        gen_btn = QPushButton('生成空监考表')
        # 生成空监考表后不关闭对话框
        gen_btn.clicked.connect(self.export_empty_overview_template)
        layout.addWidget(gen_btn)

        import_btn = QPushButton('导入预设安排')
        # 仅在导入成功时关闭对话框
        import_btn.clicked.connect(lambda: (dialog.accept() if self.import_preset_schedule() else None))
        layout.addWidget(import_btn)

        dialog.setLayout(layout)
        dialog.adjustSize()
        dialog.setSizeGripEnabled(True)
        dialog.exec_()

    def export_empty_overview_template(self):
        """
        导出空的“监考总览表”模板，仅一个sheet，不填充教师姓名。
        列名依据当前模式与科目名称/时间。
        """
        try:
            # 使用与导出/解析一致的来源（与export_schedule/parse_schedule_from_excel对齐）
            num_subjects = getattr(self, 'subject_count', None) or int(self.subject_spin.text() or '0')
            num_rooms = self.room_spin.value()
            mode = 'double' if self.double_mode_radio.isChecked() else 'single'

            file_path, _ = QFileDialog.getSaveFileName(
                self, '保存模板', '预设监考模板.xlsx', 'Excel Files (*.xlsx)')
            if not file_path:
                return

            columns = ['考场']
            for subject_id in range(1, num_subjects + 1):
                name = self.subject_names[subject_id-1] if subject_id-1 < len(self.subject_names) and self.subject_names[subject_id-1] else f"科目{subject_id}"
                time = self.exam_times[subject_id-1] if subject_id-1 < len(self.exam_times) and self.exam_times[subject_id-1] else ''
                if mode == 'double':
                    # 与parse_schedule_from_excel一致
                    columns.append(f"{name}-监考员1\n{time}")
                    columns.append(f"{name}-监考员2\n{time}")
                else:
                    # 与parse_schedule_from_excel一致
                    columns.append(f"{name}\n{time}")

            data = []
            for room in range(1, num_rooms + 1):
                row = {'考场': f"考场{room}"}
                for col in columns[1:]:
                    row[col] = ''
                data.append(row)

            df = pd.DataFrame(data, columns=columns)
            with pd.ExcelWriter(file_path, engine='xlsxwriter') as writer:
                df.to_excel(writer, sheet_name='监考总览表', index=False)

            QMessageBox.information(self, '导出成功', '已生成空的“监考总览表”模板。')
            # 询问是否打开文件所在文件夹
            reply = QMessageBox.question(self, '打开文件夹', 
                                    '是否打开文件所在文件夹？',
                                    QMessageBox.Yes | QMessageBox.No,
                                    QMessageBox.No)
            if reply == QMessageBox.Yes:
                try:
                    folder = os.path.dirname(file_path)
                    if folder and os.path.exists(folder):
                        os.startfile(folder)
                except Exception:
                    pass
        except Exception as e:
            QMessageBox.critical(self, '导出失败', f'生成空监考表时出错:\n{str(e)}')

    def import_preset_schedule(self):
        """
        导入预设安排（锁定导入项，标红，不参与二次优化/手动交换），复用现有校验逻辑。
        """
        if not self.teachers:
            QMessageBox.warning(self, '警告', '请先导入教师信息')
            return False

        file_path, _ = QFileDialog.getOpenFileName(
            self, '选择预设监考安排文件', '', 'Excel Files (*.xlsx *.xls)')
        if not file_path:
            return False

        try:
            df = pd.read_excel(file_path, sheet_name='监考总览表')
            num_subjects = getattr(self, 'subject_count', None) or int(self.subject_spin.text() or '0')
            num_rooms = self.room_spin.value()
            mode = 'double' if self.double_mode_radio.isChecked() else 'single'

            self.schedule = Schedule(self.teachers, num_subjects, num_rooms, mode)
            # 继承当前约束并强制锁定导入
            self.schedule.set_constraint('gender_mix', self.gender_mix_checkbox.isChecked())
            self.schedule.set_constraint('internal_mix', self.internal_mix_checkbox.isChecked())
            self.schedule.set_constraint('lock_imported', True)
            self.lock_imported_checkbox.setChecked(True)
            # 预设导入启用红色标记
            self.schedule.set_constraint('highlight_imported', True)
            self.schedule.set_constraint('auto_postprocess_optimize', self.auto_optimize_postprocess)
            self.schedule.set_constraint('show_optimize_button', self.show_optimize_button_flag)

            # 科目时长
            subject_durations = []
            subject_page = None
            parent = self.parent()
            if hasattr(parent, 'subject_page'):
                subject_page = parent.subject_page
            elif hasattr(parent, 'parent') and hasattr(parent.parent(), 'subject_page'):
                subject_page = parent.parent().subject_page
            elif hasattr(parent, 'subject'):
                subject_page = parent.subject
            elif hasattr(parent, 'count') and hasattr(parent, 'widget'):
                page_count = parent.count()
                for i in range(page_count):
                    page = parent.widget(i)
                    page_type = type(page).__name__
                    if 'SubjectPage' in page_type:
                        subject_page = page
                        break
            if subject_page and hasattr(subject_page, 'subjects') and hasattr(subject_page, 'get_subject_duration'):
                subject_count = len(subject_page.subjects)
                for i in range(subject_count):
                    duration = subject_page.get_subject_duration(i)
                    subject_durations.append(duration)
            else:
                subject_durations = [120] * num_subjects
            self.schedule.set_constraint('subject_durations', subject_durations)

            # 解析与校验
            errors = self.parse_schedule_from_excel(df)
            if errors:
                QMessageBox.warning(self, '验证失败', '导入的预设安排存在问题：\n' + '\n'.join(errors))
                return False

            # 展示与按钮状态
            self.display_results(self.schedule.exams)
            self.update_continue_button_state()
            self.adjust_btn.setEnabled(True)
            self.export_btn.setEnabled(True)
            self.optimize_btn.setEnabled(False)
            self.optimize_btn.setVisible(self.schedule.get_constraint('show_optimize_button', False))

            QMessageBox.information(self, '导入成功', '预设安排已导入并锁定，红色标注。')
            return True
        except Exception as e:
            QMessageBox.critical(self, '导入失败', f'导入预设安排时出错:\n{str(e)}')
            return False

    def export_schedule(self):
        """
        导出安排结果
        """
        if not self.schedule:
            QMessageBox.warning(self, '警告', '没有可导出的安排结果')
            return
            
        file_path, _ = QFileDialog.getSaveFileName(
            self, '保存安排结果', '监考安排.xlsx', 'Excel Files (*.xlsx)')
            
        if not file_path:
            return
            
        try:
            # 创建要导出的数据
            exams = self.schedule.exams
            
            # 监考总览表数据（行列格式）
            # 获取科目数和考场数
            if exams:
                num_subjects = max([exam.subject_id for exam in exams])
                all_rooms = set()
                for exam in exams:
                    all_rooms.update(exam.rooms)
                all_rooms = sorted(list(all_rooms))
                
                # 创建监考总览表数据（行列格式）
                overview_data = []
                for room in all_rooms:
                    # 根据监考模式创建行数据
                    if self.double_mode_radio.isChecked():
                        # 双教师模式：每科目两列
                        row_data = {'考场': f"考场{room}"}
                        for subject_id in range(1, num_subjects + 1):
                            # 使用实际科目名称
                            subject_name = self.subject_names[subject_id-1] if subject_id-1 < len(self.subject_names) and self.subject_names[subject_id-1] else f"科目{subject_id}"
                            exam_time = self.exam_times[subject_id-1] if subject_id-1 < len(self.exam_times) and self.exam_times[subject_id-1] else ""
                            row_data[f"{subject_name}-监考员1\n{exam_time}"] = ""
                            row_data[f"{subject_name}-监考员2\n{exam_time}"] = ""
                        overview_data.append(row_data)
                    else:
                        # 单教师模式：每科目一列
                        row_data = {'考场': f"考场{room}"}
                        for subject_id in range(1, num_subjects + 1):
                            # 使用实际科目名称
                            subject_name = self.subject_names[subject_id-1] if subject_id-1 < len(self.subject_names) and self.subject_names[subject_id-1] else f"科目{subject_id}"
                            exam_time = self.exam_times[subject_id-1] if subject_id-1 < len(self.exam_times) and self.exam_times[subject_id-1] else ""
                            row_data[f"{subject_name}\n{exam_time}"] = ""
                        overview_data.append(row_data)
                
                # 填充数据
                room_index_map = {room: idx for idx, room in enumerate(all_rooms)}
                for exam in exams:
                    for room in exam.rooms:
                        if room in exam.schedule:
                            teachers = exam.schedule[room]
                            room_row = room_index_map[room]
                            
                            if self.double_mode_radio.isChecked() and len(teachers) >= 2:
                                # 双教师模式：分别填充监考员1和监考员2
                                # 使用实际科目名称
                                subject_name = self.subject_names[exam.subject_id-1] if exam.subject_id-1 < len(self.subject_names) and self.subject_names[exam.subject_id-1] else f"科目{exam.subject_id}"
                                exam_time = self.exam_times[exam.subject_id-1] if exam.subject_id-1 < len(self.exam_times) and self.exam_times[exam.subject_id-1] else ""
                                col1_name = f"{subject_name}-监考员1\n{exam_time}"
                                col2_name = f"{subject_name}-监考员2\n{exam_time}"
                                overview_data[room_row][col1_name] = teachers[0].name
                                overview_data[room_row][col2_name] = teachers[1].name
                            elif self.double_mode_radio.isChecked() and len(teachers) == 1:
                                # 双教师模式但只有一个教师
                                # 使用实际科目名称
                                subject_name = self.subject_names[exam.subject_id-1] if exam.subject_id-1 < len(self.subject_names) and self.subject_names[exam.subject_id-1] else f"科目{exam.subject_id}"
                                exam_time = self.exam_times[exam.subject_id-1] if exam.subject_id-1 < len(self.exam_times) and self.exam_times[exam.subject_id-1] else ""
                                col1_name = f"{subject_name}-监考员1\n{exam_time}"
                                overview_data[room_row][col1_name] = teachers[0].name
                            else:
                                # 单教师模式
                                # 使用实际科目名称
                                subject_name = self.subject_names[exam.subject_id-1] if exam.subject_id-1 < len(self.subject_names) and self.subject_names[exam.subject_id-1] else f"科目{exam.subject_id}"
                                exam_time = self.exam_times[exam.subject_id-1] if exam.subject_id-1 < len(self.exam_times) and self.exam_times[exam.subject_id-1] else ""
                                # 修改：过滤掉None值后再访问name属性
                                teacher_names = ', '.join([t.name for t in teachers if t is not None])
                                overview_data[room_row][f"{subject_name}\n{exam_time}"] = teacher_names
            
            # 监考统计表数据
            stats_data = []
            if self.schedule:
                max_subject_id = max([exam.subject_id for exam in exams]) if exams else 0
                stats = self.schedule.get_statistics()
                
                for stat in stats:
                    # 查找教师对象
                    teacher = next((t for t in self.schedule.teachers if t.name == stat['name']), None)
                    if teacher:
                        # 构建统计数据行
                        stat_row = {
                            '教师姓名': teacher.name,
                            '性别': '男' if teacher.gender == 'M' else ('女' if teacher.gender == 'F' else ''),
                            '是否本校': '是' if teacher.is_internal is True else ('否' if teacher.is_internal is False else ''),
                            '最大监考段数': str(teacher.max_sessions),
                            '剩余监考次数': str(teacher.max_sessions - stat['count']),
                            '不监考科目': ','.join(map(str, teacher.unavailable_subjects))
                        }
                        
                        # 添加各科目监考情况（使用实际科目名称）
                        for j in range(1, max_subject_id+1):
                            subject_name = self.subject_names[j-1] if j-1 < len(self.subject_names) and self.subject_names[j-1] else f"科目{j}"
                            is_assigned = 1 if teacher.is_assigned_to_subject(j) else 0
                            stat_row[subject_name] = str(is_assigned)
                        
                        # 添加监考次数
                        stat_row['监考次数'] = str(stat['count'])
                        
                        # 添加本次、历次与总监考时长
                        stat_row['监考时长(分钟)'] = str(teacher.supervision_duration)
                        stat_row['历次监考时长（分钟）'] = str(teacher.previous_supervision_duration)
                        total_duration = (teacher.supervision_duration or 0) + (teacher.previous_supervision_duration or 0)
                        stat_row['总监考时长(分钟)'] = str(total_duration)
                        stats_data.append(stat_row)

            # 分科目监考表数据（每个科目一个工作表）
            subject_sheets = {}
            if exams:
                for exam in exams:
                    # 使用实际科目名称
                    subject_key = self.subject_names[exam.subject_id-1] if exam.subject_id-1 < len(self.subject_names) and self.subject_names[exam.subject_id-1] else f"科目{exam.subject_id}"
                    if subject_key not in subject_sheets:
                        subject_sheets[subject_key] = []
                    
                    # 创建科目表头
                    if self.double_mode_radio.isChecked():
                        # 双教师模式表头
                        headers = ['考场', '监考员1（姓名）', '监考员1性别', '监考员1来源', 
                                  '监考员2（姓名）', '监考员2性别', '监考员2来源']
                    else:
                        # 单教师模式表头
                        headers = ['考场', '监考教师（姓名）', '性别', '来源']
                    
                    # 按考场排序处理数据
                    sorted_rooms = sorted(exam.rooms)
                    for room in sorted_rooms:
                        if room in exam.schedule:
                            teachers = exam.schedule[room]
                            row_data = dict.fromkeys(headers, '')  # 初始化行数据
                            row_data['考场'] = room
                            
                            if self.double_mode_radio.isChecked() and len(teachers) >= 2:
                                # 双教师模式
                                # 监考员1信息
                                row_data['监考员1（姓名）'] = teachers[0].name
                                row_data['监考员1性别'] = '男' if teachers[0].gender == 'M' else ('女' if teachers[0].gender == 'F' else '')
                                row_data['监考员1来源'] = '本校' if teachers[0].is_internal is True else ('外校' if teachers[0].is_internal is False else '')
                                
                                # 监考员2信息
                                row_data['监考员2（姓名）'] = teachers[1].name
                                row_data['监考员2性别'] = '男' if teachers[1].gender == 'M' else ('女' if teachers[1].gender == 'F' else '')
                                row_data['监考员2来源'] = '本校' if teachers[1].is_internal is True else ('外校' if teachers[1].is_internal is False else '')
                            elif self.double_mode_radio.isChecked() and len(teachers) == 1:
                                # 双教师模式但只有一个教师
                                row_data['监考员1（姓名）'] = teachers[0].name
                                row_data['监考员1性别'] = '男' if teachers[0].gender == 'M' else ('女' if teachers[0].gender == 'F' else '')
                                row_data['监考员1来源'] = '本校' if teachers[0].is_internal is True else ('外校' if teachers[0].is_internal is False else '')
                            else:
                                # 单教师模式
                                # 修改：过滤掉None值后再访问name属性
                                teacher_names = ', '.join([t.name for t in teachers if t is not None])
                                row_data['监考教师（姓名）'] = teacher_names
                                
                                if len(teachers) > 0:
                                    teacher = teachers[0]
                                    row_data['性别'] = '男' if teacher.gender == 'M' else ('女' if teacher.gender == 'F' else '')
                                    row_data['来源'] = '本校' if teacher.is_internal is True else ('外校' if teacher.is_internal is False else '')
                                    
                            subject_sheets[subject_key].append(row_data)
            
            # 写入Excel文件
            with pd.ExcelWriter(file_path, engine='openpyxl') as writer:
                # 导出监考总览表
                if overview_data:
                    df_overview = pd.DataFrame(overview_data)
                    df_overview.to_excel(writer, sheet_name='监考总览表', index=False)
                
                # 导出监考统计表
                if stats_data:
                    df_stats = pd.DataFrame(stats_data)
                    df_stats.to_excel(writer, sheet_name='监考统计', index=False)
                
                # 导出分科目监考表（每个科目一个工作表）
                sheet_dfs = {}
                for subject_key, subject_data in subject_sheets.items():
                    if subject_data:
                        df_subject = pd.DataFrame(subject_data)
                        df_subject.to_excel(writer, sheet_name=subject_key, index=False)
                        sheet_dfs[subject_key] = df_subject
                
                # 获取工作簿对象以调整列宽
                workbook = writer.book
                
                # 调整监考总览表首行样式与统一列宽
                if overview_data:
                    worksheet = workbook['监考总览表']
                    try:
                        from openpyxl.styles import Alignment
                        # 设置首行高度与居中、自动换行
                        worksheet.row_dimensions[1].height = 60
                        for cell in worksheet[1]:
                            cell.alignment = Alignment(wrap_text=True, horizontal="center", vertical="center")
                    except Exception:
                        # 样式设置失败不影响导出
                        pass
                    # 统一列宽为10字符
                    for column in worksheet.columns:
                        column_letter = column[0].column_letter
                        worksheet.column_dimensions[column_letter].width = 10
                
                # 调整监考统计表列宽
                if stats_data:
                    worksheet = workbook['监考统计']
                    for column in worksheet.columns:
                        max_length = 0
                        column_letter = column[0].column_letter
                        for cell in column:
                            try:
                                if len(str(cell.value)) > max_length:
                                    max_length = len(str(cell.value))
                            except:
                                pass
                        adjusted_width = (max_length*3)
                        worksheet.column_dimensions[column_letter].width = min(adjusted_width, 50)
                
                # 调整分科目监考表列宽
                for subject_key, subject_data in subject_sheets.items():
                    if subject_data:
                        worksheet = workbook[subject_key]
                        for column in worksheet.columns:
                            max_length = 0
                            column_letter = column[0].column_letter
                            for cell in column:
                                try:
                                    if len(str(cell.value)) > max_length:
                                        max_length = len(str(cell.value))
                                except:
                                    pass
                            adjusted_width = (max_length*3)
                            worksheet.column_dimensions[column_letter].width = min(adjusted_width, 50)
                
            QMessageBox.information(self, '导出成功', f'安排结果已导出到:\n{file_path}')
            
            # 询问是否打开文件所在文件夹
            reply = QMessageBox.question(self, '打开文件夹', 
                                    '是否打开文件所在文件夹？',
                                    QMessageBox.Yes | QMessageBox.No,
                                    QMessageBox.No)
            
            if reply == QMessageBox.Yes:
                # 获取文件所在文件夹路径
                folder_path = os.path.dirname(os.path.abspath(file_path))
                # 在Windows系统中打开文件夹
                os.startfile(folder_path)

        except Exception as e:
            QMessageBox.critical(self, '导出失败', f'导出安排结果时出错:\n{str(e)}')

    def generate_teacher_template(self):
        """
        生成教师信息模板文件
        """
        file_path, _ = QFileDialog.getSaveFileName(
            self, '保存教师信息模板文件', '教师信息模板.xlsx', 'Excel Files (*.xlsx)')
            
        if not file_path:
            return
            
        try:
            # 创建模板数据
            template_data = [
                {
                    '姓名': '张三',
                    '性别': '男',
                    '是否本校': '是',
                    '最大监考段数': 3,
                    '不监考科目': '1,3',
                    '历次监考时长': 0,
                    '预设监考考场': '1'
                },
                {
                    '姓名': '李四',
                    '性别': '女',
                    '是否本校': '是',
                    '最大监考段数': 2,
                    '不监考科目': '2',
                    '历次监考时长': 0,
                    '预设监考考场': '2'
                },
                {
                    '姓名': '王五',
                    '性别': '男',
                    '是否本校': '否',
                    '最大监考段数': 4,
                    '不监考科目': '',
                    '历次监考时长': 0,
                    '预设监考考场': ''
                },
                {
                    '姓名': '赵六',
                    '性别': '女',
                    '是否本校': '否',
                    '最大监考段数': 3,
                    '不监考科目': '1,2,4',
                    '历次监考时长': 0,
                    '预设监考考场': ''
                }
            ]
            
            # 创建DataFrame
            df = pd.DataFrame(template_data)
            
            # 保存到Excel文件并按列标题长度调整列宽
            with pd.ExcelWriter(file_path, engine='openpyxl') as writer:
                sheet_name = 'Sheet1'
                df.to_excel(writer, sheet_name=sheet_name, index=False)
                workbook = writer.book
                worksheet = workbook[sheet_name]

                # 根据列标题长度设置列宽（长度 * 3，最大不超过50）
                for column in worksheet.columns:
                    header_cell = column[0]
                    header_text = str(header_cell.value) if header_cell.value is not None else ''
                    adjusted_width = max(len(header_text), 0) * 3
                    column_letter = header_cell.column_letter
                    worksheet.column_dimensions[column_letter].width = min(adjusted_width, 50)

                # ===== 新增：填写说明 sheet =====
                from openpyxl.styles import Alignment, PatternFill, Font, Border, Side
                from openpyxl.utils import get_column_letter

                # 说明文本（依据最新规则）
                instructions = {
                    '姓名': '必填。\n示例：张三、李四。',
                    '性别': '选填。\n若启用“双教师监考”且设置“男女搭配”约束，则必填。\n填写要求：男/女。',
                    '是否本校': '选填。\n若启用“双教师监考”且设置“本校外搭配”约束，则必填。\n填写要求：是/否。',
                    '最大监考段数': '选填。\n留空表示不设置监考段数限制。\n填写要求：为非负整数，且不超过科目数。',
                    '不监考科目': '选填。\n支持科目编号或科目名称；多个项可用英文/中文逗号、分号、中文分号、顿号或空格分隔，例如：\n1,3\n语文、数学\n科目1 科目2\n科目名称需与已导入科目信息一致。',
                    '历次监考时长': '选填。\n单位：分钟；可留空（默认0）。',
                    '预设监考考场': '选填。\n数字范围：1..考场数；可留空；越界或非法将被忽略。'
                }

                headers = list(df.columns)

                # 必填列：仅“姓名”始终必填并高亮
                required_cols = {'姓名'}

                # 创建“填写说明”sheet
                desc_ws = workbook.create_sheet('填写说明')

                # 样式：标题行加粗、居中、有边框；必填列浅红底色
                thin = Side(style='thin')
                header_border = Border(left=thin, right=thin, top=thin, bottom=thin)
                header_align = Alignment(horizontal='center', vertical='center', wrap_text=True)
                required_header_fill = PatternFill(start_color='FFC7CE', end_color='FFC7CE', fill_type='solid')
                normal_header_font = Font(bold=True)

                # 说明行样式：左上对齐、自动换行；必填列浅红底色
                wrap_left = Alignment(horizontal='left', vertical='top', wrap_text=True)
                required_cell_fill = PatternFill(start_color='FFC7CE', end_color='FFC7CE', fill_type='solid')

                # 写入第一行标题并设置样式
                for idx, col in enumerate(headers, start=1):
                    cell = desc_ws.cell(row=1, column=idx, value=col)
                    cell.font = normal_header_font
                    cell.alignment = header_align
                    cell.border = header_border
                    if col in required_cols:
                        cell.fill = required_header_fill

                # 写入第二行说明并设置样式
                for idx, col in enumerate(headers, start=1):
                    val = instructions.get(col, '')
                    cell = desc_ws.cell(row=2, column=idx, value=val)
                    cell.alignment = wrap_left
                    if col in required_cols:
                        cell.fill = required_cell_fill

                # 第二行行高（按多行文本适配）
                max_lines = max((instructions.get(col, '').count('\n') + 1) for col in headers)
                line_height = 25  # 每行约18pt
                padding = 8
                desc_ws.row_dimensions[2].height = max_lines * line_height + padding

                # 设置“填写说明”sheet列宽为统一的20字符
                for idx in range(1, len(headers) + 1):
                    letter = get_column_letter(idx)
                    desc_ws.column_dimensions[letter].width = 20
                # ===== 新增结束 =====
            
            QMessageBox.information(self, '生成成功', f'教师信息模板文件已生成:\n{file_path}')
            
            # 询问是否打开文件所在文件夹
            reply = QMessageBox.question(self, '打开文件夹', 
                                    '是否打开文件所在文件夹？',
                                    QMessageBox.Yes | QMessageBox.No,
                                    QMessageBox.No)
            
            if reply == QMessageBox.Yes:
                # 获取文件所在文件夹路径
                folder_path = os.path.dirname(os.path.abspath(file_path))
                # 在Windows系统中打开文件夹
                os.startfile(folder_path)

        except Exception as e:
            QMessageBox.critical(self, '生成失败', f'生成教师信息模板文件时出错:\n{str(e)}')
