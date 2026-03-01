import openpyxl
from openpyxl.styles import Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.pagebreak import Break


class AdmissionTicketGenerator:
    """
    负责生成准考证 Excel 文件的核心逻辑类。
    """

    # 模板尺寸常量
    TEMPLATE_WIDTH = 7  # 单个模板占用的列数 (A-G 共7列)
    TEMPLATES_PER_ROW = 3  # 每一行放置的模板数量

    # ==========================================
    # 样式定义
    # ==========================================
    # 1. 边框样式
    THIN_BORDER = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )

    DASH_DOT_SIDE = Side(style="dashDot")

    # 2. 对齐样式
    CENTER_ALIGN = Alignment(horizontal="center", vertical="center", wrap_text=True)

    # 3. 字体样式
    FONT_TITLE = Font(name="宋体", size=14, bold=True)  # 大标题
    FONT_HEADER = Font(name="宋体", size=10, bold=True)  # 表头标签
    FONT_NORMAL = Font(name="宋体", size=10, bold=False)  # 普通内容

    def __init__(self, config):
        self.config = config
        self.subjects = config.subjects
        # 获取时间，如果 config 中没有该属性（兼容旧代码），则设为空列表
        self.subject_times = getattr(config, "subject_times", [""] * len(self.subjects))

        # 动态计算高度
        # 基础行: 标题(1) + 考号(1) + 姓名(1) + 表头(1) = 4 行
        # 科目行: len(subjects)
        # 间隔行: 2 行 (底部/顶部间隔)
        self.total_subject_rows = len(self.subjects)
        self.TEMPLATE_HEIGHT = 4 + self.total_subject_rows + 2

        # 动态计算每页纵向模板数量 (Templates Per Column per Page)
        # 根据科目数量进行判断，保持与台角纸一致的分页逻辑
        subject_count = len(self.subjects)

        if subject_count <= 3:
            self.TEMPLATES_PER_COL_PAGE = 5
        elif 4 <= subject_count <= 5:
            self.TEMPLATES_PER_COL_PAGE = 4
        elif 6 <= subject_count <= 9:
            self.TEMPLATES_PER_COL_PAGE = 3
        else:  # >= 10
            self.TEMPLATES_PER_COL_PAGE = 2

    def draw_template(self, ws, start_row, start_col, data=None, is_first_col=False):
        """
        在指定位置绘制一个完整的准考证模板。

        参数:
            ws: Worksheet 对象
            start_row: 模板起始行索引
            start_col: 模板起始列索引
            data: 考生数据字典
            is_first_col: 是否为该行的第一个模板（用于设置行高）
        """

        # 提取数据
        kaochang = data.get("考场", "") if data else ""
        kaochang_no = data.get("考场号", "") if data else ""
        seat_no = data.get("座位号", "") if data else ""
        name = data.get("考生姓名", "") if data else ""
        exam_no = data.get("考生考号", "") if data else ""
        class_no = data.get("班级", "") if data else ""
        student_no = data.get("学号", "") if data else ""

        # --- 辅助函数：相对坐标获取单元格 ---
        def cell(r_offset, c_offset):
            return ws.cell(row=start_row + r_offset, column=start_col + c_offset)

        # 设置行高 (仅在第一列时设置，避免重复操作)
        if is_first_col:
            # 标题行 18.75 (Row 0)
            ws.row_dimensions[start_row].height = 18.75
            # 内容行 13.5 (Row 1-N)
            # 内容行数 = 3 (Basic info) + N (Subjects)
            # Row 1, 2, 3 + Subjects
            content_rows_count = 3 + self.total_subject_rows
            for r_offset in range(1, 1 + content_rows_count):
                ws.row_dimensions[start_row + r_offset].height = 13.5

            # 间隔行 5 (最后两行)
            # 索引是 TEMPLATE_HEIGHT - 2 和 TEMPLATE_HEIGHT - 1
            last_idx = self.TEMPLATE_HEIGHT - 1
            ws.row_dimensions[start_row + last_idx - 1].height = 5
            ws.row_dimensions[start_row + last_idx].height = 5

        # --- 1. 绘制标题行 (Row 0) ---
        # 合并 A-E 列作为标题区域
        ws.merge_cells(
            start_row=start_row,
            start_column=start_col,
            end_row=start_row,
            end_column=start_col + 4,
        )
        c = cell(0, 0)
        c.value = self.config.title
        c.font = self.FONT_TITLE
        c.alignment = self.CENTER_ALIGN

        # 修正: 为标题行合并区域内的所有单元格设置边框
        # 顶部、左右、底部都需要 THIN_BORDER
        for c_offset in range(5):
            c_tmp = cell(0, c_offset)
            c_tmp.border = self.THIN_BORDER

        # --- 2. 考号行 (Row 1) ---
        # 布局: A(考号), B(值), C-D(班级), E(值)

        # A: 考号
        c = cell(1, 0)
        c.value = "考号"
        c.font = self.FONT_HEADER
        c.alignment = self.CENTER_ALIGN
        c.border = self.THIN_BORDER

        # B: 考号值
        c = cell(1, 1)
        c.value = exam_no
        c.font = self.FONT_NORMAL
        c.alignment = self.CENTER_ALIGN
        c.border = self.THIN_BORDER

        # C-D: 班级 (Label)
        ws.merge_cells(
            start_row=start_row + 1,
            start_column=start_col + 2,
            end_row=start_row + 1,
            end_column=start_col + 3,
        )
        c = cell(1, 2)
        c.value = "班级"
        c.font = self.FONT_HEADER
        c.alignment = self.CENTER_ALIGN
        c.border = self.THIN_BORDER
        cell(1, 3).border = self.THIN_BORDER

        # E: 班级值
        c = cell(1, 4)
        c.value = class_no
        c.font = self.FONT_NORMAL
        c.alignment = self.CENTER_ALIGN
        c.border = self.THIN_BORDER

        # --- 3. 姓名行 (Row 2) ---
        # 布局: A(姓名), B(值), C-D(学号), E(值)

        # A: 姓名
        c = cell(2, 0)
        c.value = "姓名"
        c.font = self.FONT_HEADER
        c.alignment = self.CENTER_ALIGN
        c.border = self.THIN_BORDER

        # B: 姓名值
        c = cell(2, 1)
        c.value = name
        c.font = self.FONT_NORMAL
        c.alignment = self.CENTER_ALIGN
        c.border = self.THIN_BORDER

        # C-D: 学号 (Label)
        ws.merge_cells(
            start_row=start_row + 2,
            start_column=start_col + 2,
            end_row=start_row + 2,
            end_column=start_col + 3,
        )
        c = cell(2, 2)
        c.value = "学号"
        c.font = self.FONT_HEADER
        c.alignment = self.CENTER_ALIGN
        c.border = self.THIN_BORDER
        cell(2, 3).border = self.THIN_BORDER

        # E: 学号值
        c = cell(2, 4)
        c.value = student_no
        c.font = self.FONT_NORMAL
        c.alignment = self.CENTER_ALIGN
        c.border = self.THIN_BORDER

        # --- 4. 列表头 (Row 3) ---
        # 布局: A(科目), B(时间), C(考场), D(考场号), E(座位号)

        headers = [(0, 0, "科目"), (1, 1, "时间"), (2, 2, "考场"), (3, 3, "考场号"), (4, 4, "座位号")]

        for c_start, c_end, text in headers:
            if c_start != c_end:
                ws.merge_cells(
                    start_row=start_row + 3,
                    start_column=start_col + c_start,
                    end_row=start_row + 3,
                    end_column=start_col + c_end,
                )

            c = cell(3, c_start)
            c.value = text
            c.font = self.FONT_HEADER
            c.alignment = self.CENTER_ALIGN
            c.border = self.THIN_BORDER

            if c_start != c_end:
                for i in range(c_start + 1, c_end + 1):
                    cell(3, i).border = self.THIN_BORDER

        # --- 5. 科目数据行 (Row 4 to 11) ---
        subjects = self.subjects
        # total_subject_rows 已在 __init__ 计算

        for i in range(self.total_subject_rows):
            r_current = 4 + i
            subject = subjects[i] if i < len(subjects) else ""
            subject_time = self.subject_times[i] if i < len(self.subject_times) else ""

            # A: 科目
            c = cell(r_current, 0)
            c.value = subject
            c.font = self.FONT_NORMAL
            c.alignment = self.CENTER_ALIGN
            c.border = self.THIN_BORDER

            # B: 时间 (单列)
            c = cell(r_current, 1)
            c.value = subject_time  # 填入时间
            c.font = self.FONT_NORMAL
            c.alignment = self.CENTER_ALIGN
            c.border = self.THIN_BORDER

            # C: 考场
            c = cell(r_current, 2)
            c.value = kaochang
            c.font = self.FONT_NORMAL
            c.alignment = self.CENTER_ALIGN
            c.border = self.THIN_BORDER

            # D: 考场号
            c = cell(r_current, 3)
            c.value = kaochang_no
            c.font = self.FONT_NORMAL
            c.alignment = self.CENTER_ALIGN
            c.border = self.THIN_BORDER

            # E: 座位号
            c = cell(r_current, 4)
            c.value = seat_no
            c.font = self.FONT_NORMAL
            c.alignment = self.CENTER_ALIGN
            c.border = self.THIN_BORDER

        # --- 6. 裁剪线与间隔列 (F, G) ---
        # F列 (Index 5): 右边框为虚线 (Gap 1.625)
        # G列 (Index 6): 左边框为虚线 (Gap 1.0)

        last_idx = self.TEMPLATE_HEIGHT - 1

        for r_off in range(self.TEMPLATE_HEIGHT):  # 遍历所有行
            # F Column (Index 5)
            c_f = cell(r_off, 5)

            # G Column (Index 6)
            c_g = cell(r_off, 6)

            # 默认边框
            f_border_style = {"right": self.DASH_DOT_SIDE}
            g_border_style = {"left": self.DASH_DOT_SIDE}

            # 特殊处理最后两行 (间隔行)的边框，形成闭合区域
            if r_off == last_idx - 1:  # 倒数第二行
                # 绘制 A-E 内容区域的底部边框
                for c_off in range(5):
                    c = cell(r_off, c_off)
                    cur = c.border
                    new_border = Border(
                        left=cur.left,
                        right=cur.right,
                        top=cur.top,
                        bottom=self.DASH_DOT_SIDE,
                    )
                    c.border = new_border

                f_border_style["bottom"] = self.DASH_DOT_SIDE
                g_border_style["bottom"] = self.DASH_DOT_SIDE

            elif r_off == last_idx:  # 最后一行
                # 绘制 A-E 内容区域的顶部边框（如果有下一个准考证，这里就是顶部）
                for c_off in range(5):
                    c = cell(r_off, c_off)
                    cur = c.border
                    new_border = Border(
                        left=cur.left,
                        right=cur.right,
                        top=self.DASH_DOT_SIDE,
                        bottom=cur.bottom,
                    )
                    c.border = new_border

                f_border_style["top"] = self.DASH_DOT_SIDE
                g_border_style["top"] = self.DASH_DOT_SIDE

            # 应用 F/G 边框
            c_f.border = Border(**f_border_style)
            c_g.border = Border(**g_border_style)

    def generate_sheet(
        self,
        ws,
        data_list,
        start_idx=0,
        progress_callback=None,
        total_steps=1,
        include_room_in_footer=True,
        footer_msg=None,
        force_break_by_class=False,
    ):
        """
        在给定的 Worksheet 上绘制一批数据。
        """
        num_templates = len(data_list)
        
        current_grid_row = 0
        current_col = 0
        last_class = None
        
        # 记录上一次分页的行号
        last_break_row_idx = 0
        
        for idx, data in enumerate(data_list):
            if idx % self.TEMPLATES_PER_ROW == 0 and progress_callback:
                current = start_idx + current_grid_row
                progress_callback(current, total_steps)

            # 强制分页逻辑
            cls = data.get("班级") if data else None
            
            if force_break_by_class and idx > 0 and cls != last_class:
                if current_col > 0:
                    current_grid_row += 1
                    current_col = 0
                
                current_start_row = current_grid_row * self.TEMPLATE_HEIGHT + 1
                ws.row_breaks.append(Break(id=current_start_row - 1))
                
                # 重置分页基准
                last_break_row_idx = current_grid_row
            
            if data:
                last_class = cls

            # 插入自然分页符 (基于相对行号)
            rows_since_break = current_grid_row - last_break_row_idx
            
            if rows_since_break > 0 and rows_since_break % self.TEMPLATES_PER_COL_PAGE == 0 and current_col == 0:
                current_start_row = current_grid_row * self.TEMPLATE_HEIGHT + 1
                ws.row_breaks.append(Break(id=current_start_row - 1))
                last_break_row_idx = current_grid_row

            start_row = current_grid_row * self.TEMPLATE_HEIGHT + 1
            start_col = current_col * self.TEMPLATE_WIDTH + 1

            # 优化：只在第一列设置行高
            is_first = current_col == 0
            self.draw_template(ws, start_row, start_col, data, is_first_col=is_first)
            
            current_col += 1
            if current_col >= self.TEMPLATES_PER_ROW:
                current_col = 0
                current_grid_row += 1

        # 2. 列宽设置
        col_widths_template = [8.125, 17.625, 8.125, 6.625, 6.625, 1.275, 1.275]
        
        max_grid_cols = self.TEMPLATES_PER_ROW
        if num_templates < self.TEMPLATES_PER_ROW and current_grid_row == 0:
             max_grid_cols = num_templates

        for grid_c in range(max_grid_cols):
            base_col_idx = grid_c * self.TEMPLATE_WIDTH
            for local_c, width in enumerate(col_widths_template):
                col_letter = get_column_letter(base_col_idx + local_c + 1)
                ws.column_dimensions[col_letter].width = width

        # 3. 页面设置 (A4 横向)
        ws.page_setup.paperSize = 9  # A4
        ws.page_setup.orientation = "landscape"

        ws.page_margins.left = 0.2
        ws.page_margins.right = 0
        ws.page_margins.top = 0.2
        ws.page_margins.bottom = 0.06
        ws.page_margins.header = 0
        ws.page_margins.footer = 0

        # 4. 页脚设置
        if footer_msg:
            footer_text = f"&\"宋体\"&8第 &P 页，共 &N 页，{footer_msg}"
        elif include_room_in_footer:
            first_kaochang = ""
            for d in data_list:
                if d and d.get("考场"):
                    first_kaochang = d.get("考场")
                    break
            footer_text = f"&\"宋体\"&8第 &P 页，共 &N 页，当前考场：{first_kaochang}"
        else:
            footer_text = f"&\"宋体\"&8第 &P 页，共 &N 页"

        ws.oddFooter.center.text = footer_text
        ws.evenFooter.center.text = footer_text

    def generate(self, progress_callback=None):
        """
        批量生成准考证。
        返回生成的 Excel 文件路径。
        """
        num_templates = self.config.num_templates
        output_path = self.config.output_path
        student_data_list = self.config.student_data_list

        wb = openpyxl.Workbook()

        if not student_data_list:
            # 模式 1: 生成空白模板
            ws = wb.active
            ws.title = "空白模板"
            dummy_data = [None] * num_templates
            self.generate_sheet(
                ws,
                dummy_data,
                0,
                progress_callback,
                (num_templates + 2) // 3,
                include_room_in_footer=False,
            )

        else:
            # 模式 2: 根据数据生成 (包含总表和分班级表)
            # 首先对所有数据进行排序：班级 (从小到大) -> 学号 (从小到大)
            # 辅助函数：尝试将字符串转换为数字进行排序
            def safe_int_sort_key(val):
                if isinstance(val, int):
                    return val
                if isinstance(val, str) and val.isdigit():
                    return int(val)
                # 尝试从字符串中提取数字 (例如 "1班" -> 1)
                import re

                match = re.search(r"(\d+)", str(val))
                if match:
                    return int(match.group(1))
                return 0  # 无法提取数字则排在最前或最后，这里设为0

            student_data_list.sort(key=lambda x: (safe_int_sort_key(x.get("班级", 0)), safe_int_sort_key(x.get("学号", 0))))

            # 分组依据：班级
            class_groups = {}
            class_order = []

            # 按班级分组
            for d in student_data_list:
                # 构造班级名称，确保统一
                # 这里假设 '班级' 已经是纯数字或统一格式，如果为了显示好看，可以加上"班"字
                # 但 Sheet 名最好简洁
                bj_val = d.get("班级", "未分类")
                class_name = f"{bj_val}班" if str(bj_val).isdigit() else str(bj_val)

                if class_name not in class_groups:
                    class_groups[class_name] = []
                    class_order.append(class_name)
                class_groups[class_name].append(d)

            total_templates = len(student_data_list)
            total_rows_summary = (total_templates + self.TEMPLATES_PER_ROW - 1) // self.TEMPLATES_PER_ROW
            total_steps = total_rows_summary * 2  # 简单估算进度步数

            processed_steps = 0

            # 1. 生成总表 (按班级顺序排列)
            ws_total = wb.active
            ws_total.title = "总表"
            self.generate_sheet(
                ws_total,
                student_data_list,
                processed_steps,
                progress_callback,
                total_steps,
                include_room_in_footer=False,
                force_break_by_class=True, # 开启强制分页
            )
            processed_steps += total_rows_summary

            # 2. 生成分班级表
            for cls_name in class_order:
                # 处理 Excel Sheet 名称非法字符
                safe_name = (
                    str(cls_name)[:30]
                    .replace(":", "")
                    .replace("\\", "")
                    .replace("/", "")
                    .replace("?", "")
                    .replace("*", "")
                    .replace("[", "")
                    .replace("]", "")
                )
                ws = wb.create_sheet(title=safe_name)

                cls_data = class_groups[cls_name]
                # 分班级表中，页脚显示班级信息而不是考场
                # 临时修改 generate_sheet 中的页脚逻辑有点麻烦，
                # 我们可以复用 include_room_in_footer 参数，但此时它显示的"当前考场"文案可能不合适
                # 最好修改 generate_sheet 的页脚逻辑
                self.generate_sheet(
                    ws,
                    cls_data,
                    processed_steps,
                    progress_callback,
                    total_steps,
                    include_room_in_footer=False,
                    footer_msg=f"班级：{cls_name}",
                )

                rows_in_cls = (len(cls_data) + self.TEMPLATES_PER_ROW - 1) // self.TEMPLATES_PER_ROW
                processed_steps += rows_in_cls

        wb.save(output_path)
        return output_path

