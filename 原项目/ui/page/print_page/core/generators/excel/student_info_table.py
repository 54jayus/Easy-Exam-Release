import os
from collections import defaultdict

import openpyxl
from openpyxl.styles import Font, Alignment, Border, Side
from openpyxl.worksheet.pagebreak import Break
from openpyxl.worksheet.page import PageMargins


class StudentInfoTableGenerator:
    def __init__(self, config):
        self.config = config
        self.include_subject_fields = bool(getattr(config, "include_subject_fields", False))
        self.group_mode = str(getattr(config, "group_mode", "class") or "class")
        
        # Row Heights
        self.title_header_row_height = 20
        if self.group_mode == "examroom":
            self.info_row_height = 16
        else:
            self.info_row_height = 13.5

        self._print_area_end_col = "J" if self.include_subject_fields else "G"

        # Define styles
        self.font_title = Font(name="宋体", size=14, bold=True)
        self.font_header = Font(name="宋体", size=11, bold=True)
        self.font_normal = Font(name="宋体", size=11)
        
        self.align_center = Alignment(horizontal="center", vertical="center", wrap_text=True)
        self.align_left = Alignment(horizontal="left", vertical="center", wrap_text=True)
        
        self.border_thin = Border(
            left=Side(style="thin"), 
            right=Side(style="thin"), 
            top=Side(style="thin"), 
            bottom=Side(style="thin")
        )

    def generate(self, progress_callback=None):
        #不再加载外部模板，直接创建新工作簿
        wb = openpyxl.Workbook()
        # 删除默认的 Sheet
        if "Sheet" in wb.sheetnames:
            del wb["Sheet"]
            
        title_value = getattr(self.config, "title", "") or ""
        data_list = getattr(self.config, "student_data_list", None) or []

        # 创建总表
        ws_total = wb.create_sheet("总表")
        self._prepare_sheet(ws_total, title_value)

        if not data_list:
            # 如果没有数据，填充一些空行示例
            self._fill_blank_rows(ws_total, 50)
        else:
            grouped = self._group_data(data_list)
            group_order = sorted(grouped.keys(), key=self._group_sort_key)
            
            # 填充总表
            self._fill_total_sheet(ws_total, grouped, group_order, progress_callback)
            
            # 为每个分组创建分表
            for idx, group_key in enumerate(group_order, start=1):
                sheet_name = self._group_sheet_name(group_key, grouped[group_key])
                unique_name = self._unique_sheet_name(wb, sheet_name)
                ws_group = wb.create_sheet(unique_name)
                self._prepare_sheet(ws_group, title_value)
                self._fill_single_group_sheet(ws_group, group_key, grouped[group_key], progress_callback, idx, len(group_order))

        out_path = self.config.output_path
        wb.save(out_path)
        wb.close()
        return out_path

    def _prepare_sheet(self, ws, title_value):
        # 设置页面边距 (Left/Right: 0, Top/Bottom: ~0.2)
        ws.page_margins = PageMargins(left=0, right=0, top=0.2, bottom=0.2, header=0, footer=0)
        
        # 设置列宽
        # 根据是否包含选科字段设置不同的列宽
        if self.include_subject_fields:
            # A:班级, B:学号, C:姓名, D:考号, E:首选, F:选1, G:选2, H:考场, I:考场号, J:座位号
            ws.column_dimensions['A'].width = 8.7
            ws.column_dimensions['B'].width = 10
            ws.column_dimensions['C'].width = 10
            ws.column_dimensions['D'].width = 12.7
            ws.column_dimensions['E'].width = 7.7
            ws.column_dimensions['F'].width = 7.7
            ws.column_dimensions['G'].width = 7.7
            ws.column_dimensions['H'].width = 10.7
            ws.column_dimensions['I'].width = 8.7
            ws.column_dimensions['J'].width = 8
            max_col_idx = 10
            merge_range = "A1:J1"
        else:
            # A:班级, B:学号, C:姓名, D:考号, E:考场, F:考场号, G:座位号
            ws.column_dimensions['A'].width = 8.7
            ws.column_dimensions['B'].width = 10
            ws.column_dimensions['C'].width = 10
            ws.column_dimensions['D'].width = 12.7
            ws.column_dimensions['E'].width = 10.7  # 考场
            ws.column_dimensions['F'].width = 8.7   # 考场号
            ws.column_dimensions['G'].width = 8     # 座位号
            max_col_idx = 7
            merge_range = "A1:G1"

        # 设置标题行 (Row 1)
        ws.merge_cells(merge_range)
        cell_title = ws.cell(row=1, column=1, value=title_value)
        cell_title.font = self.font_title
        cell_title.alignment = self.align_center
        # 标题四个方向都需要实线边框，即使合并了单元格，也需要对范围内所有单元格应用边框样式
        for c in range(1, max_col_idx + 1):
            ws.cell(row=1, column=c).border = self.border_thin
        
        ws.row_dimensions[1].height = self.title_header_row_height

        # 设置表头行 (Row 2)
        headers = ["班级", "学号", "考生姓名", "考生考号"]
        if self.include_subject_fields:
            headers.extend(["首选", "选科1", "选科2", "考场", "考场号", "座位号"])
        else:
            headers.extend(["考场", "考场号", "座位号"])

        for col_idx, header_text in enumerate(headers, start=1):
            cell = ws.cell(row=2, column=col_idx, value=header_text)
            cell.font = self.font_header
            cell.alignment = self.align_center
            cell.border = self.border_thin

        ws.row_dimensions[2].height = self.title_header_row_height
        
        # 设置打印标题行
        ws.print_title_rows = "1:2"
        ws.print_options.horizontalCentered = True
        
        # 清除大纲设置
        ws.sheet_properties.outlinePr.summaryBelow = False
        ws.sheet_properties.outlinePr.summaryRight = False

    def _fill_blank_rows(self, ws, blank_rows=50):
        start_row = 3
        max_col = 10 if self.include_subject_fields else 7
        
        for r in range(start_row, start_row + blank_rows):
            for c in range(1, max_col + 1):
                cell = ws.cell(row=r, column=c)
                cell.font = self.font_normal
                cell.alignment = self.align_center
                cell.border = self.border_thin
            ws.row_dimensions[r].height = self.info_row_height
        
        # 总计数行
        last_row = start_row + blank_rows
        ws.cell(row=last_row, column=1).value = "总计数"
        ws.cell(row=last_row, column=3).value = f"=SUBTOTAL(3,C{start_row}:C{last_row-1})"
        for c in range(1, max_col + 1):
            cell = ws.cell(row=last_row, column=c)
            cell.font = self.font_normal
            cell.alignment = self.align_center
            cell.border = self.border_thin
        ws.row_dimensions[last_row].height = self.info_row_height
        
        self._set_print_area(ws, last_row)

    def _fill_total_sheet(self, ws, grouped, class_order, progress_callback):
        current_row = 3
        total_steps = sum(len(grouped[c]) + 1 for c in class_order) if class_order else 1
        done = 0

        for class_name in class_order:
            students = grouped[class_name]
            label = self._group_count_label(class_name, students)
            start_row = current_row
            
            for item in students:
                self._write_student_row(ws, current_row, item)
                ws.row_dimensions[current_row].height = self.info_row_height
                current_row += 1
                done += 1
                if progress_callback:
                    progress_callback(done, total_steps)

            end_row = current_row - 1
            
            # 分组计数行
            self._write_count_row(ws, current_row, label, start_row, end_row)
            ws.row_dimensions[current_row].height = self.info_row_height
            ws.row_breaks.append(Break(id=current_row))
            current_row += 1
            done += 1
            if progress_callback:
                progress_callback(done, total_steps)

        self._set_print_area(ws, max(2, current_row - 1))

    def _fill_single_group_sheet(self, ws, group_key, students, progress_callback, index, total):
        current_row = 3
        start_row = current_row
        
        for item in students:
            self._write_student_row(ws, current_row, item)
            ws.row_dimensions[current_row].height = self.info_row_height
            current_row += 1
            
        end_row = current_row - 1
        
        # 分组计数行
        label = self._group_count_label(group_key, students)
        self._write_count_row(ws, current_row, label, start_row, end_row)
        ws.row_dimensions[current_row].height = self.info_row_height
        ws.row_breaks.append(Break(id=current_row))
        self._set_print_area(ws, current_row)

        if progress_callback and total > 0:
            progress_callback(index, total)

    def _write_student_row(self, ws, row_idx, item):
        # 写入数据
        ws.cell(row=row_idx, column=1).value = item.get("班级", "")
        ws.cell(row=row_idx, column=2).value = item.get("学号", "")
        ws.cell(row=row_idx, column=3).value = item.get("考生姓名", item.get("姓名", ""))
        ws.cell(row=row_idx, column=4).value = item.get("考生考号", item.get("考号", ""))

        if self.include_subject_fields:
            ws.cell(row=row_idx, column=5).value = item.get("首选", item.get("类别", ""))
            ws.cell(row=row_idx, column=6).value = item.get("选科1", item.get("选1", ""))
            ws.cell(row=row_idx, column=7).value = item.get("选科2", item.get("选2", ""))
            ws.cell(row=row_idx, column=8).value = item.get("考场", "")
            ws.cell(row=row_idx, column=9).value = item.get("考场号", "")
            ws.cell(row=row_idx, column=10).value = item.get("座位号", item.get("座位", ""))
            max_col = 10
        else:
            ws.cell(row=row_idx, column=5).value = item.get("考场", "")
            ws.cell(row=row_idx, column=6).value = item.get("考场号", "")
            ws.cell(row=row_idx, column=7).value = item.get("座位号", item.get("座位", ""))
            max_col = 7
            
        # 设置样式
        for col in range(1, max_col + 1):
            cell = ws.cell(row=row_idx, column=col)
            cell.font = self.font_normal
            cell.alignment = self.align_center
            cell.border = self.border_thin

    def _write_count_row(self, ws, row_idx, label, start_row, end_row):
        label_col = self._count_label_col()
        ws.cell(row=row_idx, column=label_col).value = f"{label} 计数"
        
        if end_row >= start_row:
            ws.cell(row=row_idx, column=3).value = f"=SUBTOTAL(3,C{start_row}:C{end_row})"
        else:
            ws.cell(row=row_idx, column=3).value = "=0"
            
        max_col = 10 if self.include_subject_fields else 7
        for col in range(1, max_col + 1):
            cell = ws.cell(row=row_idx, column=col)
            cell.font = self.font_normal
            cell.alignment = self.align_center
            cell.border = self.border_thin

    def _count_label_col(self):
        if self.group_mode != "examroom":
            return 1
        return 8 if self.include_subject_fields else 5

    def _set_print_area(self, ws, end_row):
        ws.print_area = f"A1:{self._print_area_end_col}{end_row}"

    # --- Grouping and Sorting Helpers (Unchanged) ---
    def _group_data(self, data_list):
        if self.group_mode == "examroom":
            return self._group_by_examroom(data_list)
        return self._group_by_class(data_list)

    def _group_by_class(self, data_list):
        grouped = defaultdict(list)
        for item in data_list:
            class_value = item.get("班级", "")
            grouped[str(class_value).strip()].append(item)
        for k, v in grouped.items():
            v.sort(key=lambda x: self._class_student_sort_key(x))
        return grouped

    def _group_by_examroom(self, data_list):
        grouped = defaultdict(list)
        for item in data_list:
            examroom_no = item.get("考场号", "")
            grouped[str(examroom_no).strip()].append(item)
        for _, v in grouped.items():
            v.sort(key=lambda x: self._examroom_student_sort_key(x))
        return grouped

    def _group_sort_key(self, group_key):
        if self.group_mode == "examroom":
            return self._examroom_sort_key(group_key)
        return self._class_sort_key(group_key)

    def _class_sort_key(self, class_name):
        s = str(class_name).strip()
        if s.isdigit():
            return (0, int(s))
        return (1, s)

    def _class_student_sort_key(self, item):
        sn = str(item.get("学号", "")).strip()
        exam_no = str(item.get("考生考号", "")).strip()
        if sn.isdigit():
            return (0, int(sn), exam_no)
        return (1, sn, exam_no)

    def _examroom_sort_key(self, examroom_no):
        s = str(examroom_no).strip()
        if s.isdigit():
            return (0, int(s))
        if s:
            return (1, s)
        return (2, "")

    def _seat_sort_key(self, seat_value):
        s = str(seat_value).strip()
        if s.isdigit():
            return (0, int(s))
        return (1, s)

    def _examroom_student_sort_key(self, item):
        examroom_no = str(item.get("考场号", "")).strip()
        seat = item.get("座位号", item.get("座位", ""))
        class_value = str(item.get("班级", "")).strip()
        sn = str(item.get("学号", "")).strip()
        return (self._examroom_sort_key(examroom_no), self._seat_sort_key(seat), self._class_sort_key(class_value), self._seat_sort_key(sn))

    def _group_sheet_name(self, group_key, students):
        if self.group_mode == "examroom":
            room = str((students[0] if students else {}).get("考场", "")).strip()
            if room:
                return room
            key = str(group_key).strip()
            return key or "考场"
        return f"{str(group_key).strip()}班"

    def _group_count_label(self, group_key, students):
        if self.group_mode == "examroom":
            room = str((students[0] if students else {}).get("考场", "")).strip()
            if room:
                return room
            key = str(group_key).strip()
            return key or "考场"
        return str(group_key).strip()

    def _unique_sheet_name(self, wb, name):
        base = self._safe_sheet_name(name)
        if base not in wb.sheetnames:
            return base
        i = 2
        while True:
            suffix = f"({i})"
            trimmed = base
            max_len = 31 - len(suffix)
            if len(trimmed) > max_len:
                trimmed = trimmed[:max_len]
            candidate = f"{trimmed}{suffix}"
            if candidate not in wb.sheetnames:
                return candidate
            i += 1

    def _safe_sheet_name(self, name):
        invalid = ["\\", "/", "*", "?", ":", "[", "]"]
        for ch in invalid:
            name = name.replace(ch, " ")
        name = name.strip()
        if not name:
            name = "Sheet"
        return name[:31]
