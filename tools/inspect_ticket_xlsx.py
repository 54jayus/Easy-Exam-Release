from __future__ import annotations

import json
from dataclasses import asdict, dataclass
from pathlib import Path
from typing import Any, Dict, List, Optional, Tuple

import openpyxl
from openpyxl.worksheet.worksheet import Worksheet


@dataclass
class BorderSideInfo:
    style: Optional[str]
    color: Optional[str]


@dataclass
class BorderInfo:
    left: BorderSideInfo
    right: BorderSideInfo
    top: BorderSideInfo
    bottom: BorderSideInfo


@dataclass
class FontInfo:
    name: Optional[str]
    size: Optional[float]
    bold: bool
    italic: bool
    underline: Optional[str]
    color: Optional[str]


@dataclass
class AlignmentInfo:
    horizontal: Optional[str]
    vertical: Optional[str]
    wrap_text: Optional[bool]
    shrink_to_fit: Optional[bool]


@dataclass
class CellInfo:
    address: str
    value: Any
    font: FontInfo
    alignment: AlignmentInfo
    border: BorderInfo
    number_format: Optional[str]


def _rgb(color_obj: Any) -> Optional[str]:
    if color_obj is None:
        return None
    rgb = getattr(color_obj, "rgb", None)
    if rgb:
        return str(rgb)
    indexed = getattr(color_obj, "indexed", None)
    if indexed is not None:
        return f"indexed:{indexed}"
    theme = getattr(color_obj, "theme", None)
    if theme is not None:
        return f"theme:{theme}"
    return None


def _side(side_obj: Any) -> BorderSideInfo:
    if side_obj is None:
        return BorderSideInfo(style=None, color=None)
    return BorderSideInfo(style=getattr(side_obj, "style", None), color=_rgb(getattr(side_obj, "color", None)))


def _border(border_obj: Any) -> BorderInfo:
    if border_obj is None:
        empty = BorderSideInfo(style=None, color=None)
        return BorderInfo(left=empty, right=empty, top=empty, bottom=empty)
    return BorderInfo(
        left=_side(getattr(border_obj, "left", None)),
        right=_side(getattr(border_obj, "right", None)),
        top=_side(getattr(border_obj, "top", None)),
        bottom=_side(getattr(border_obj, "bottom", None)),
    )


def _font(font_obj: Any) -> FontInfo:
    if font_obj is None:
        return FontInfo(name=None, size=None, bold=False, italic=False, underline=None, color=None)
    return FontInfo(
        name=getattr(font_obj, "name", None),
        size=float(getattr(font_obj, "size", 0) or 0) or None,
        bold=bool(getattr(font_obj, "bold", False)),
        italic=bool(getattr(font_obj, "italic", False)),
        underline=getattr(font_obj, "underline", None),
        color=_rgb(getattr(font_obj, "color", None)),
    )


def _alignment(alignment_obj: Any) -> AlignmentInfo:
    if alignment_obj is None:
        return AlignmentInfo(horizontal=None, vertical=None, wrap_text=None, shrink_to_fit=None)
    return AlignmentInfo(
        horizontal=getattr(alignment_obj, "horizontal", None),
        vertical=getattr(alignment_obj, "vertical", None),
        wrap_text=getattr(alignment_obj, "wrap_text", None),
        shrink_to_fit=getattr(alignment_obj, "shrink_to_fit", None),
    )


def _cell_info(ws: Worksheet, row: int, col: int) -> CellInfo:
    cell = ws.cell(row=row, column=col)
    return CellInfo(
        address=cell.coordinate,
        value=cell.value,
        font=_font(cell.font),
        alignment=_alignment(cell.alignment),
        border=_border(cell.border),
        number_format=getattr(cell, "number_format", None),
    )


def _find_title_block(ws: Worksheet) -> Tuple[Optional[str], Optional[Any]]:
    for rng in sorted(ws.merged_cells.ranges, key=lambda r: (r.min_row, r.min_col)):
        v = ws.cell(row=rng.min_row, column=rng.min_col).value
        if isinstance(v, str) and v.strip() and ("准考证" in v or "考试" in v):
            return v.strip(), rng
    return None, None


def inspect(path: Path) -> Dict[str, Any]:
    wb = openpyxl.load_workbook(path, data_only=False)
    ws = wb.active

    title_value, title_merge = _find_title_block(ws)
    if title_merge is None:
        r0, c0, r1, c1 = (1, 1, min(ws.max_row or 1, 60), min(ws.max_column or 1, 20))
    else:
        r0 = int(title_merge.min_row)
        c0 = int(title_merge.min_col)
        c1 = int(title_merge.max_col)
        next_title = None
        for rng in sorted(ws.merged_cells.ranges, key=lambda r: (r.min_row, r.min_col)):
            if int(rng.min_col) == c0 and int(rng.max_col) == c1 and int(rng.min_row) > r0:
                v = ws.cell(row=rng.min_row, column=rng.min_col).value
                if isinstance(v, str) and v.strip() and ("准考证" in v or "考试" in v):
                    next_title = rng
                    break
        if next_title is not None:
            r1 = int(next_title.min_row) - 1
        else:
            r1 = min((ws.max_row or r0), r0 + 80)

    col_widths: Dict[str, Any] = {}
    for col in range(c0, c1 + 1):
        letter = openpyxl.utils.get_column_letter(col)
        dim = ws.column_dimensions.get(letter)
        col_widths[letter] = {"width": getattr(dim, "width", None), "hidden": bool(getattr(dim, "hidden", False))}

    row_heights: Dict[str, Any] = {}
    for row in range(r0, r1 + 1):
        dim = ws.row_dimensions.get(row)
        row_heights[str(row)] = {"height": getattr(dim, "height", None), "hidden": bool(getattr(dim, "hidden", False))}

    merged_ranges = [str(rng) for rng in ws.merged_cells.ranges if not (rng.max_row < r0 or rng.min_row > r1)]

    cells: List[Dict[str, Any]] = []
    for row in range(r0, r1 + 1):
        for col in range(c0, c1 + 1):
            info = _cell_info(ws, row, col)
            b = info.border
            has_border = any(
                getattr(getattr(b, side, None), "style", None) for side in ("left", "right", "top", "bottom")
            )
            if info.value is None and not has_border:
                continue
            cells.append(asdict(info))

    return {
        "workbook": {"sheet": ws.title},
        "title": title_value,
        "ticket_guess": {"r0": r0, "c0": c0, "r1": r1, "c1": c1},
        "merged_ranges": merged_ranges,
        "column_dimensions": col_widths,
        "row_dimensions": row_heights,
        "cells_in_block": cells,
    }


def main() -> None:
    xlsx = Path(r"d:\百度网盘同步空间\BaiduSyncdisk\自动化办公\智能考务系统\测试文件\准考证_批量生成.xlsx")
    out = inspect(xlsx)
    print(json.dumps(out, ensure_ascii=False, indent=2))


if __name__ == "__main__":
    main()

